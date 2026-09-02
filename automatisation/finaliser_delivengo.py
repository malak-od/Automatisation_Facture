#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR DELIVENGO — produit "AAAA_MM_Delivengo_LPPAQ.xlsx" A L'IDENTIQUE du
fichier fait a la main (feuilles Pays + Fichier import, memes colonnes, memes
formules, memes couleurs), en partant du fichier final comme MODELE et en n'y
remplacant que les donnees, tirees de l'export du suivi Delivengo (.xls).

Mapping export -> onglet "Fichier import" :
  A Date remise      <- "Remise en poste"        F Statut       <- "Statut General"
  G Numero de suivi  <- "Numero de suivi"        H Destinataire <- "Destinataire nom"
  X Pays (nom)       <- "Destinataire pays"      P Droits&taxes  = poids export brut (recherchex tracking)
  I E/P              = E si raison sociale, sinon P
  B/C/N/L/U          = constants (DELIVENGO-LPPQ / 1er du mois / avecsuivi / 1 / 1)
  J,K,M,O,W          = FORMULES (recherche table Pays)   Q Assurance = poids export du suivi (colonne Poids) / 1000
  Dans le CLASSEUR (cet onglet), P et Q RESTENT remplies (servent aussi a M = MAX(P,Q)).
  Le fichier import ERP separe (23 colonnes standard), lui, remet DroitsTaxes/Assurance
  a 0 -- cf. index.js. En fin de traitement : vider W (Taxe Gasoil, "Pas de TG") si 0,
  et vider F (Statut) -> apply_final_cleanup().

Necessite : Windows + Excel + pywin32 + pandas + xlrd.
Usage : python finaliser_delivengo.py "<modele.xlsx>" "<sortie.xlsx>" "<export.xls>"
"""
import sys, os, shutil, time, re
from datetime import datetime

# positions des colonnes utiles dans l'export (index 0-based, robuste aux accents)
X_REMISE, X_SUIVI, X_RS, X_NOM, X_PAYS, X_STATUT, X_POIDS = 7, 9, 10, 11, 16, 24, 30

# formules du modele (ligne 2) ; recopiees vers le bas par FillDown. Cle = colonne (1-based).
ROW2_FORMULAS = {
    10: '=IF(COUNTIF(Pays!A:A,X2)=0,"Pays inconnu",XLOOKUP(X2,Pays!A:A,Pays!B:B))',  # J Pays code
    11: '=IF(COUNTIF(Pays!A:A,X2)=0,"Pays inconnu",XLOOKUP(X2,Pays!A:A,Pays!C:C))',  # K Zone
    13: '=ROUND(MAX(P2:Q2),2)',                                                      # M Poids
    15: '=LOOKUP(X2,Pays!A:A,Pays!D:D)',                                             # O TVA
    23: '=IF(COUNTIF(Pays!A:A,X2)=0,"",LOOKUP(X2,Pays!A:A,Pays!E:E))',               # W Taxe Gasoil
}

def to_date(s):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s).strip()[:10], fmt)
        except ValueError:
            pass
    return None

def num(s):
    try:
        return float(str(s).replace(",", ".").strip())
    except (ValueError, AttributeError):
        return 0.0

EXCEL_EPOCH = datetime(1899, 12, 30)
def excel_serial(d):
    """Date -> numero de serie Excel (evite tout souci de fuseau/locale)."""
    return (d - EXCEL_EPOCH).days if d else None

def load_export(path):
    import pandas as pd
    df = pd.read_excel(path, dtype=str, header=0)
    df = df.where(df.notna(), "")
    return [list(r) for r in df.itertuples(index=False, name=None)]

# ---- Poids reel : export WMS « expeditions brut » ----
X_BRUT_TRACK, X_BRUT_POIDS = 41, 34  # PRO_TRACKING (AP) / INFO_POIDSRETENU (AI, deja en kg)

def normalize_track(t):
    """Normalise un n° de suivi pour le matching (upper + retrait espaces/tirets)."""
    return re.sub(r"[^A-Z0-9]", "", str(t).upper()) if t is not None else ""

def load_brut(brut_paths):
    """Map {tracking normalise -> poids kg}. Ordre des fichiers = priorite (mois courant, puis M-1).
    Comme le RECHERCHEX du fichier fait main : la PREMIERE occurrence gagne
    (un tracking duplique dans le brut garde le poids de sa 1ere ligne)."""
    import openpyxl
    m = {}
    for p in [bp for bp in brut_paths if bp]:
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > X_BRUT_TRACK:
                t = normalize_track(row[X_BRUT_TRACK])
                w = row[X_BRUT_POIDS]
                if t and t not in m and isinstance(w, (int, float)):
                    m[t] = w
        wb.close()
    return m

def build_block(rows, brut_map, period=None):
    """Bloc 2D (n x 24). Colonnes-formule laissees VIDES (posees ensuite par FillDown).
    'period' (AAAA_MM) = mois choisi dans l'UI, source de verite prioritaire pour la Date
    validite tarif -- BUG TROUVE 2026-09-01 : elle restait calculee sur la 1ere date de remise
    trouvee dans l'export uploade, jamais sur le choix utilisateur (meme piege deja corrige
    cote UPS, carriers/ups/index.js)."""
    m = re.fullmatch(r"(\d{4})_(\d{2})", period or "")
    if m:
        date_valid = datetime(int(m.group(1)), int(m.group(2)), 1)
    else:
        if period:
            print(f"AVERTISSEMENT: --period {period!r} invalide (attendu AAAA_MM) -- date de remise de l'export utilisee en repli.")
        valid = next((to_date(r[X_REMISE]) for r in rows if to_date(r[X_REMISE])), None)
        date_valid = datetime(valid.year, valid.month, 1) if valid else None
    nb_reel = 0
    manquants = []   # trackings sans poids reel (a verifier/saisir a la main, comme l'operateur)
    block = []
    for r in rows:
        ep = "P"  # constante Delivengo (B2C) : fait-main = P sur 935/935, meme avec raison sociale
        track = str(r[X_SUIVI]).strip()
        poids_reel = brut_map.get(normalize_track(track))  # kg, apparie par tracking (P)
        poids_suivi = round(num(r[X_POIDS]) / 1000.0, 3) if len(r) > X_POIDS and str(r[X_POIDS]).strip() else None  # kg, export du suivi (Q)
        if poids_reel is not None:
            nb_reel += 1
        elif not poids_suivi:  # ni P (brut WMS) ni Q (export suivi) : la un vrai plancher 0,15 est applique
            manquants.append(track)
        line = [None] * 24
        line[0] = str(r[X_REMISE]).strip()           # A Date remise (TEXTE, comme le modele)
        line[1] = "DELIVENGO-LPPQ"                   # B Transporteur
        line[2] = excel_serial(date_valid)           # C Date validite tarif (n. de serie -> date)
        line[5] = str(r[X_STATUT]).strip()           # F Statut General
        line[6] = track                              # G Numero de suivi
        line[7] = str(r[X_NOM]).strip()              # H Destinataire nom
        line[8] = ep                                 # I E / P
        line[11] = 1                                 # L Nbr Colis
        line[13] = "avecsuivi"                       # N mode envoi
        line[15] = poids_reel if poids_reel is not None else 0.15  # P Droits et taxes (scratch) = poids export brut (recherchex) ; plancher 0,15 kg si absent
        line[16] = poids_suivi  # Q Assurance (scratch) = poids export du suivi (colonne Poids) / 1000
        line[20] = 1                                 # U Fret
        line[23] = str(r[X_PAYS]).strip()            # X Pays (nom)
        block.append(line)
    return block, date_valid, nb_reel, manquants

def apply_final_cleanup(ws, newLast):
    """Etape finale p.14 "Dans le fichier import" -- pour le CLASSEUR (onglet "Fichier
    import") : P (Droits et taxes) et Q (Assurance) RESTENT remplies (poids brut WMS /
    poids export du suivi /1000). Seul le fichier import ERP separe (23 colonnes
    standard) force ses colonnes DroitsTaxes/Assurance a 0, independamment (cf.
    index.js/HEADER_TO_KEY, applique a la lecture) :
      - geler M et W (formules) en valeurs (sinon M/W resteraient des formules figees
        sur un classeur qui peut etre rouvert hors de ce script)
      - Taxe gasoil (W) : "Pas de TG" -> vider si 0
      - Statut (F) : toujours vide (info operationnelle interne, pas destinee a l'import)
    """
    if newLast < 2:
        return
    for col in (13, 23):  # M Poids, W Taxe Gasoil : formules -> valeurs figees
        rng = ws.Range(ws.Cells(2, col), ws.Cells(newLast, col))
        retry(lambda rng=rng: setattr(rng, "Value", rng.Value))
    n = newLast - 1
    wVals = ws.Range(ws.Cells(2, 23), ws.Cells(newLast, 23)).Value
    wVals = [[wVals]] if n == 1 else wVals
    for i, rowv in enumerate(wVals):
        v = rowv[0] if isinstance(rowv, (tuple, list)) else rowv
        if v in (0, None, ""):
            retry(lambda i=i: ws.Cells(2 + i, 23).ClearContents())
    retry(lambda: ws.Range(ws.Cells(2, 6), ws.Cells(newLast, 6)).ClearContents())    # F Statut

def retry(fn, tries=8, delay=0.6):
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:  # Excel occupe -> RPC_E_CALL_REJECTED : on reessaie
            last = e
            time.sleep(delay)
    raise last

def parse_args(argv):
    """<modele> <sortie> --export <suivi.xls> --brut <brut_M.xlsx> [<brut_M-1.xlsx> ...]
    [--period AAAA_MM] (retrocompat : le 1er positionnel apres sortie = export)."""
    modele, sortie = argv[1], argv[2]
    export, brut, period, cur = None, [], None, None
    for a in argv[3:]:
        if a == "--export": cur = "e"
        elif a == "--brut": cur = "b"
        elif a == "--period": cur = "p"
        elif cur == "b": brut.append(a)
        elif cur == "p": period = a; cur = None
        elif cur == "e" or export is None: export = a
        else: brut.append(a)
    return modele, sortie, export, brut, period

def main():
    modele, sortie, export, brut_paths, period = parse_args(sys.argv)
    shutil.copyfile(modele, sortie)
    rows = load_export(export)
    brut_map = load_brut(brut_paths)
    block, date_valid, nb_reel, manquants = build_block(rows, brut_map, period)
    n = len(block)
    print(f"Export : {n} lignes | periode {date_valid.strftime('%Y_%m') if date_valid else '?'} "
          f"| poids reel apparie (brut) : {nb_reel}/{n} ({len(brut_map)} trackings dans le brut)")
    # lignes lisibles par le serveur -> remontees dans le dashboard comme infos
    for t in manquants:
        print(f"INFO_POIDS_MANQUANT:{t}")

    import win32com.client as win32
    xlUp = -4162
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier")
        time.sleep(0.4)
        ws = wb.Sheets("Fichier import")
        oldLast = ws.Cells(ws.Rows.Count, 7).End(xlUp).Row  # via col G
        newLast = 1 + n
        maxLast = max(oldLast, newLast)
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(maxLast, 24)).ClearContents())
        # colonne A = TEXTE (sinon Excel auto-convertit "01/06/2026" en date US)
        retry(lambda: setattr(ws.Range(ws.Cells(2, 1), ws.Cells(maxLast, 1)), "NumberFormat", "@"))
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(newLast, 24)).__setattr__("Value", block))
        # formules : pose ligne 2 puis recopie vers le bas (FillDown, natif Excel)
        for col, f in ROW2_FORMULAS.items():
            retry(lambda col=col, f=f: setattr(ws.Cells(2, col), "Formula", f))
            if newLast > 2:
                retry(lambda col=col: ws.Range(ws.Cells(2, col), ws.Cells(newLast, col)).FillDown())
        xl.Calculate()
        apply_final_cleanup(ws, newLast)
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)
        print("OK -> " + sortie)
    finally:
        try:
            xl.Quit()
        except Exception:
            pass

if __name__ == "__main__":
    main()
