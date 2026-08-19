#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR CHRONOPOST — produit "AAAA_MM_Facture Chronopost.xlsx" A
L'IDENTIQUE du fichier fait a la main (14 feuilles), en partant du fichier
existant comme MODELE. Le transporteur le plus complexe du projet : 2
fichiers Excel bruts (1 par sous-compte/contrat) consolides, un POOL gazole
par facture redistribue au prorata du fret, un TCD en cascade (TCD poids +
TCD alimentent "Fichier import" par formule), plusieurs tables de lookup
(Bibliotheque transporteurs, Categories, Pays TVA, Sous-comptes, Zoning
2shop).

Feuilles STATIQUES (jamais touchees, copiees telles quelles avec le modele) :
  Pays TVA, Sous-comptes, Bibliothèque transporteurs, Catégories, TARIFS,
  cap à 5%, Zoning 2shop, Bilan clients, Avoir.

"Facture Chronopost" (A1:AJ) : colonne A ("ID Clients") est une colonne
CALCULEE (PAS dans le brut -- confirme : le brut recu commence directement
par "No Facture" = colonne B du modele, pas colonne A). Donnees brutes des 2
fichiers sources collees a partir de la colonne B (en-tete des 2 fichiers
Chronopost STABLE, confirme identique sur les 2 sous-comptes de juin 2026 --
pas de decalage mois a mois comme Mondial Relay/DPD, donc pas de resolution
par NOM colonne par colonne necessaire pour le COLLAGE, seulement pour les
formules qui les referencent), puis colonnes calculees W->AE reconstruites
par formule :
  W (Code produit modifie 2shop-Kersun) : resynchronisee sur le modele 2026_07
  (a jour, 2026-08-12) -- =IF(S deja code standard connu,S,IF(sous-compte 0/1
  + produit 5X/5Y/6B/6C,S,IF(C<>2,"",IF(S="5X","5XK",...)))). La version de
  juin ne gerait que le dernier cas (Kersun) -- colonne d'affichage/controle
  interne uniquement, n'alimente aucun calcul en aval de ce script.
  X (Zoning 2shop) : =IF(S="6B",XLOOKUP(H,'Zoning 2shop'!A:A,B:B),
     IF(S="6C",XLOOKUP(G,'Zoning 2shop'!C:C,D:D),""))
  Y (Gazole %) : =AB/Z   Z (fret) : =T   AA (surete+eco taux) : =IF(...,0.08,0.5)
  AB (gazole) : =$AG$5/$AG$2*Z  (pool gazole / total fret * fret ligne)
  AC (hors gazole) : =Z+AA   AD (Categories) : =IF(COUNTIF(...)=0,"catégorie
  inconnue",LOOKUP(N,Catégories!A:A,Catégories!B:B))
  AE (Total avec GO) : =SUM(Z:AB)
Zone recap AF1:AI9 (labels/sous-totaux Frêt/eco/sureté/Gazole/Gestion, taux
gasoil routier/aerien) : AG2/AG3/AG4/AG5 (Frêt/eco/sûreté/Gazole) calcules en
PYTHON (somme par prefixe Numero LT : CAP*/ECO*/SUR*/reste) et ecrits EN
VALEURS ici -- PAS par formule Excel (positions/plages variables mois a
mois selon le tri manuel, cf. transcription video, non fiable a automatiser
par formule). Le reste de la zone recap (AF6:AI9, "Gestion"/"Gazole
reel"/taux gasoil routier-aerien) reste fige au modele, jamais recalcule.

"TCD poids"/"TCD"/"Contrôle pdf" : VRAIS TCD Excel, PivotCache redirige vers
la nouvelle plage de "Facture Chronopost" (meme piege que DPD/Geodis/Mondial
Relay : le cache du modele pointe une plage figee/etroite).

"TCD" colonnes de calcul manuel (S/T/U/V, puis A/B/C/W/Y/Z a partir de la
ligne ou "ID client" natif du TCD existe) reconstruites par formule EXACTE
du modele (RECHERCHEX vers Bibliothèque transporteurs/Facture Chronopost).

"Fichier import" : formules FIXES referencant TCD/TCD poids par POSITION
(stables une fois TCD/TCD poids reconstruits) -- MAIS le carrier Node a DEJA
identifie/supprime les lignes forfaitaires CAP/ECO/SUR de l'import (jamais
dans le TCD natif ventile par Numero LT de facon utile pour l'import ERP) --
donc ICI on ecrit "Fichier import" en VALEURS deja calculees par le carrier
Node (fiable, deja valide sur les 7 PDF de juin 2026 a 0,00€ d'ecart), pas
par re-derivation de formules TCD -- meme principe que "Import CSV" de BLS.

Reconciliation PDF (onglet "Contrôle pdf") : Total HT extrait du bloc "TOTAL
FACTURE<HT><TVA><TTC>" de chaque PDF (format texte structure, confirme sur
les 7 PDF de juin 2026), colle en face du No Facture correspondant.

Necessite : Windows + Excel + pywin32 + pypdf.
Usage :
  python finaliser_chronopost.py "<modele.xlsx>" "<sortie.xlsx>" "<facture1.xlsx>" ["<facture2.xlsx>"...] [--pdf <pdf1> [<pdf2>...]]
"""
import sys, os, shutil, re


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def compare_key(h):
    s = normalize_header(h).lower()
    s = s.translate(str.maketrans("éèêëàâäùûüôöîïç", "eeeeaaauuuooiic"))
    s = re.sub(r"[.,]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def is_xlsx(path):
    import zipfile
    try:
        return zipfile.is_zipfile(path)
    except Exception:
        return False


def col_letter(idx0):
    import openpyxl
    return openpyxl.utils.get_column_letter(idx0 + 1)


def col_index(header, name):
    target = compare_key(name)
    for i, h in enumerate(header):
        if compare_key(h) == target:
            return i
    return None


def to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_facture_brute(path):
    """Facture Chronopost brute : feuille unique, en-tete EXACT en ligne 4 (index 3),
    donnees a partir de la ligne 5 -- meme structure confirmee sur les 2 sous-comptes."""
    import openpyxl, tempfile
    load_path = path
    tmp_path = None
    if not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copyfile(path, tmp.name)
        load_path = tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(load_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    header = [normalize_header(h) for h in rows[3]]
    data = [r for r in rows[4:] if any(v not in (None, "") for v in r)]
    return {"file": os.path.basename(path), "header": header, "rows": data}


def excel_serial_from_date(v):
    """v = 'DD/MM/AAAA' str, serial Excel deja numerique, ou datetime.datetime."""
    import datetime as _dt
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    if isinstance(v, _dt.datetime):
        return (v - EXCEL_EPOCH).days
    if isinstance(v, (int, float)):
        return int(v)
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", str(v or "").strip())
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return (_dt.datetime(y, mo, d) - EXCEL_EPOCH).days


def extract_pdf_total(pdf_path):
    """PDF Chronopost : bloc 'Compte<n>Facture<n>...' + 'TOTAL FACTURE<HT><TVA><TTC>'
    colle en fin de document (format texte structure, confirme sur les 7 PDF de juin
    2026) -- meme logique que src/carriers/chronopost/index.js, a garder synchronisee."""
    try:
        import pypdf
    except ImportError:
        return None
    try:
        text = "\n".join((page.extract_text() or "") for page in pypdf.PdfReader(pdf_path).pages)
    except Exception:
        return None
    m_compte = re.search(r"Compte\s*(\d+)\s*Facture\s*(\d+)", text)
    m_total = re.search(r"TOTAL\s*FACTURE\s*([\d\s]+[.,]\d{2})\s*([\d\s]+[.,]\d{2})\s*([\d\s]+[.,]\d{2})", text)
    if not m_compte or not m_total:
        return None
    return {
        "file": os.path.basename(pdf_path),
        "compte": m_compte.group(1),
        "facture": m_compte.group(2),
        "total_ht": round(to_num(m_total.group(1).replace(" ", "")), 2),
    }


def parse_args(argv):
    modele, sortie = argv[1], argv[2]
    factures, pdfs, cur = [], [], "in"
    for a in argv[3:]:
        if a == "--pdf":
            cur = "pdf"
        elif cur == "pdf":
            pdfs.append(a)
        else:
            factures.append(a)
    return modele, sortie, factures, pdfs


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def redirect_pivot_caches(wb, source_ws, sheet_names, source_sheet, last_row, first_col, last_col):
    """Redirige le PivotCache de tous les TCD des feuilles donnees vers la vraie plage de
    donnees (colonnes first_col->last_col de source_ws), SEULEMENT si leur source actuelle
    reference deja source_sheet (evite de casser un TCD imbrique sur un autre TCD, meme piege
    que Mondial Relay/DPD).

    IMPORTANT : la SourceData du MODELE pointe vers une plage MINUSCULE et locale (constate
    via COM sur les 3 TCD Chronopost : 'Facture Chronopost'!C2:C18/C2:C30/C2:C20, PAS
    B1:AD1048576 comme une premiere lecture openpyxl l'avait suggere -- openpyxl et COM
    peuvent lire des metadonnees differentes/desynchronisees) -- meme "vestige de plage
    figee" deja rencontre sur DPD/Geodis/Mondial Relay. On ignore donc DELIBEREMENT cette
    plage figee et on redirige vers TOUTE la largeur utile (first_col->last_col, fixee par
    l'appelant selon les colonnes reelles de 'Facture Chronopost'), PAS une colonne derivee
    de l'ancienne SourceData."""
    xlDatabase = 1
    new_range = source_ws.Range(source_ws.Cells(1, first_col), source_ws.Cells(last_row, last_col))
    for sheet_name in sheet_names:
        ws = wb.Sheets(sheet_name)
        count = ws.PivotTables().Count
        for i in range(1, count + 1):
            pt = ws.PivotTables(i)
            src = str(pt.PivotCache().SourceData)
            if source_sheet not in src:
                print(f"TCD '{pt.Name}' ({sheet_name}) : source '{src}' n'est pas '{source_sheet}' "
                      f"(TCD imbrique ?) -> non redirige, config preservee.")
                continue
            new_cache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=new_range)
            pt.ChangePivotCache(new_cache)
            print(f"TCD '{pt.Name}' ({sheet_name}) : PivotCache redirige vers {new_range.Address} "
                  f"(ancienne source figee '{src}' ignoree).")


def main():
    modele, sortie, facture_paths, pdf_paths = parse_args(sys.argv)
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    factures_brutes = [read_facture_brute(p) for p in facture_paths if is_xlsx(p)]
    if not factures_brutes:
        raise RuntimeError("Aucune facture Chronopost fournie (xlsx, 1 par sous-compte).")

    header = factures_brutes[0]["header"]
    all_rows = []
    for f in factures_brutes:
        i_map = {name: col_index(f["header"], name) for name in header}
        for r in f["rows"]:
            all_rows.append([r[i_map[name]] if i_map[name] is not None and i_map[name] < len(r) else None for name in header])
    ncol = len(header)
    n = len(all_rows)
    print(f"Entrée : {len(factures_brutes)} facture(s), {n} ligne(s)")

    i_facture = col_index(header, "No Facture")
    i_numero_lt = col_index(header, "Numero LT")
    i_produit = col_index(header, "Produit")
    i_date_lt = col_index(header, "Date LT")
    i_sous_compte = col_index(header, "Sous-compte")
    i_pays_arrivee = col_index(header, "Pays arrivee")
    i_pays_depart = col_index(header, "Pays depart")

    # "Sous-compte" est du TEXTE dans le brut recu ('0'/'1'/'2', pas 0/1/2) -- BUG TROUVE
    # 2026-08-14 : la formule W ("Code produit modifie") compare C<>2 (nombre), qui est
    # TOUJOURS VRAI face a du texte en Excel (PAS de coercion implicite dans les
    # comparaisons =/<> d'un IF, contrairement a une addition/soustraction) -- resultat, les
    # 1383 lignes 2SHOP (sous-compte='2', produit 6C/5Y) de juin 2026 tombaient TOUTES dans
    # le cas vide au lieu de devenir 6CK/5YK. Converti en nombre ICI (a la source, avant
    # collage dans "Facture Chronopost") plutot que de reecrire la formule en comparaison
    # texte -- decision utilisateur 2026-08-14 (corriger le type de donnee, pas la formule).
    if i_sous_compte is not None:
        for r in all_rows:
            v = r[i_sous_compte]
            if isinstance(v, str) and v.strip().lstrip("-").isdigit():
                r[i_sous_compte] = int(v.strip())

    # TRI DES LIGNES (2026-08-13, decision utilisateur revisee) : "Facture Chronopost" triee
    # STRICTEMENT A -> Z par "Numero LT" (colonne L), y compris les lignes forfaitaires
    # CAP*/ECO*/SUR* -- elles se retrouvent donc melangees aux vraies lignes de colis (leurs
    # codes "CAPI1"/"ECORI"/"SURTI" etc. sont disperses alphabetiquement entre les trackings
    # reels type "XF..."/"XN..."). Un tri anterieur regroupait ces lignes forfaitaires EN TETE
    # (CAP puis ECO puis SUR, imitant le fichier fait a la main) pour permettre des formules
    # Excel AF2:AF5 en plages contiguees -- devenu inutile depuis que AG2:AG5 utilisent des
    # formules SUMIF sur "Type prestation" (independantes de la position des lignes, cf. plus
    # bas), donc ce tri alphabetique strict n'a aucun impact sur ces formules.
    all_rows.sort(key=lambda row: str(row[i_numero_lt] or "").strip())

    # Date validite = mois MAJORITAIRE sur l'ensemble des lignes (PAS la 1re ligne trouvee) --
    # BUG TROUVE 2026-08-13 (meme fix cote Node, chronopost/index.js) : le brut Chronopost
    # recu pour un mois donne contient parfois quelques lignes residuelles du mois precedent
    # EN TETE de fichier (livraisons tardives facturees avec la coupure du mois suivant,
    # confirme sur facture_chronopost_51291303_202607.xlsx : 3 lignes datees du 30/06/2026 en
    # tete de fichier, sur 1717 lignes totales dont 1714 de juillet). Prendre la 1re ligne
    # donnait "juin" pour un fichier envoye et rempli a 99,8% de juillet.
    import datetime as _dt
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    comptage_mois = {}
    for r in all_rows:
        s = excel_serial_from_date(r[i_date_lt]) if i_date_lt is not None else None
        if s is not None:
            d = EXCEL_EPOCH + _dt.timedelta(days=s)
            cle = (d.year, d.month)
            comptage_mois[cle] = comptage_mois.get(cle, 0) + 1
    date_validite_serial = None
    if comptage_mois:
        annee, mois = max(comptage_mois.items(), key=lambda kv: kv[1])[0]
        date_validite_serial = (_dt.datetime(annee, mois, 1) - EXCEL_EPOCH).days
        if len(comptage_mois) > 1:
            detail = ", ".join(f"{a}-{m:02d}: {n} ligne(s)" for (a, m), n in sorted(comptage_mois.items(), key=lambda kv: -kv[1]))
            print(f"INFO: plusieurs mois detectes dans les fichiers recus ({detail}) — mois retenu : {mois:02d}/{annee} (majoritaire).")

    pdfs = [r for r in (extract_pdf_total(p) for p in pdf_paths) if r]

    import win32com.client as win32
    xlUp = -4162
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) "Facture Chronopost" : purge + collage donnees brutes a partir de la colonne B
        #    (colonne A = "ID Clients", CALCULEE, PAS dans le brut recu -- confirme : le brut
        #    commence directement par "No Facture" = colonne B du modele). En-tete des 2
        #    fichiers Chronopost STABLE (confirme identique sur les 2 sous-comptes de juin
        #    2026) -- pas de decalage mois a mois comme Mondial Relay/DPD, mais l'OFFSET fixe
        #    de +1 colonne (a cause de "ID Clients") doit etre applique partout.
        ws = wb.Sheets("Facture Chronopost")
        # PIEGE MAJEUR (decouvert en debug) : le modele a un AutoFilter actif sur cette feuille
        # -> ClearContents()/Value= sur une plage de lignes echouent SILENCIEUSEMENT (aucune
        # exception levee, aucun avertissement) des que le filtre masque des lignes -- Excel
        # restreint alors les operations de plage aux lignes VISIBLES uniquement, ce qui laisse
        # les anciennes valeurs du modele intactes tout en semblant reussir. DOIT etre desactive
        # AVANT toute manipulation de plage sur cette feuille.
        if ws.AutoFilterMode:
            ws.AutoFilterMode = False
        FIRST_RAW_COL = 2  # colonne B
        LAST_RAW_COL = FIRST_RAW_COL - 1 + ncol  # B -> U (20 colonnes brutes)
        oldLast = ws.Cells(ws.Rows.Count, 2).End(xlUp).Row
        newLast = 1 + n
        maxLast = max(oldLast, newLast, 2)
        LAST_CALC_COL = 31  # AE (colonnes calculees W->AE = 23->31 ; AF/32 = 'Total', reste une formule fixe du modele, jamais reecrite ici)

        # +15 lignes de marge : purge aussi l'ancienne liste des 9 postes ERP (colonne AD,
        # ecrite a oldLast+3..oldLast+11 par un mois precedent) -- ces lignes n'ont rien en
        # colonne B, donc "oldLast" (End(xlUp) sur B) ne les detecte pas si le nouveau mois a
        # moins de lignes de donnees que l'ancien, ce qui laisserait un residu de liste au
        # milieu des donnees.
        #
        # BUG TROUVE 2026-08-14 : la marge "+15" partait de maxLast = max(oldLast, newLast),
        # donc si newLast (nouveau mois) > oldLast (taille du modele au moment de la marge
        # d'origine), la marge ne couvrait plus la VRAIE position de cette liste -- laissee
        # residuelle bien AU-DELA de la plage purgee (constate : liste a la ligne 3717+ alors
        # que maxLast+15 s'arretait avant). Consequence concrete : le PivotCache "TCD" (colField
        # "Categories") scannait cette liste-legende comme si c'etaient de vraies lignes,
        # ajoutant "Assurance"/"Frais facturation" comme items meme sans aucune vraie donnee
        # -- OU, pire, les FAISANT DISPARAITRE du TCD si la plage de redirection PivotCache
        # (calculee sur newLast, PAS oldLast+marge) s'arretait avant cette liste desormais mal
        # placee, alors qu'elles existaient reellement mais ailleurs dans les vraies donnees.
        # Fix : purge sur une marge FIXE et large (200 lignes) au-dela du max des 2 tailles,
        # jamais relative a une seule des deux.
        purgeUntil = maxLast + 200
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(purgeUntil, LAST_CALC_COL)).ClearContents())
        # "No Facture" (et "Numero LT") sont du TEXTE dans le modele ET dans le brut recu
        # ("13655988", pas 13655988) -- ecrire via COM .Value sur une plage dont le format de
        # cellule n'est pas force en Texte laisse Excel AUTO-CONVERTIR certaines valeurs en
        # nombre selon le contexte, ce qui casse le TCD natif "Contrôle pdf"/"TCD" (champ
        # rowField mixte texte/nombre -> agregation aberrante, constate : une concatenation de
        # plusieurs "13655988" lue comme un seul grand nombre). Force NumberFormat="@" (Texte)
        # sur ces 2 colonnes AVANT d'ecrire les valeurs pour eviter toute auto-conversion.
        col_no_facture_letter = col_letter(FIRST_RAW_COL - 1 + i_facture) if i_facture is not None else None
        col_numero_lt_letter = col_letter(FIRST_RAW_COL - 1 + i_numero_lt) if i_numero_lt is not None else None
        for letter in (col_no_facture_letter, col_numero_lt_letter):
            if letter:
                retry(lambda l=letter: ws.Range(f"{l}2:{l}{maxLast}").__setattr__("NumberFormat", "@"))
        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, LAST_RAW_COL)).__setattr__("Value", all_rows))
        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, LAST_CALC_COL)).ClearContents())

        # Colonnes calculees X->AF (formules du modele, colonnes source resolues par nom + offset).
        col_c = col_letter(FIRST_RAW_COL - 1 + i_sous_compte) if i_sous_compte is not None else "D"
        col_s = col_letter(FIRST_RAW_COL - 1 + i_produit) if i_produit is not None else "T"
        col_g = col_letter(FIRST_RAW_COL - 1 + i_pays_depart) if i_pays_depart is not None else "H"
        col_h = col_letter(FIRST_RAW_COL - 1 + i_pays_arrivee) if i_pays_arrivee is not None else "I"
        col_t = col_letter(FIRST_RAW_COL - 1 + col_index(header, "Montant HT")) if col_index(header, "Montant HT") is not None else "U"
        col_n = col_letter(FIRST_RAW_COL - 1 + col_index(header, "Type prestation")) if col_index(header, "Type prestation") is not None else "O"

        # Colonnes/references VERIFIEES cellule par cellule contre le vrai modele (row 2,
        # 2026_06_Facture Chronopost.xlsx) -- un decalage d'1 colonne ici (cle ET references
        # internes AA/AB/AC/Z) avait ete introduit par erreur, laissant W (Code produit
        # modifie 2shop-Kersun) TOUJOURS VIDE malgre la bibliotheque a jour (bug trouve et
        # corrige 2026-08-12, cf. capture ecran utilisateur "Facture Chronopost").
        formulas_w_ae = {
            # W : formule resynchronisee sur le classeur 2026_07 (a jour, verifie 2026-08-12
            # -- celle de juin ne gerait QUE le cas Kersun, juillet ajoute 2 cas manquants :
            # (a) produit deja un code standard connu -> recopie tel quel, (b) sous-compte 0/1
            # (pas Kersun) avec produit 5X/5Y/6B/6C -> recopie tel quel sans le suffixe K).
            23: f'=IF(OR({{col_s}}{{{{row}}}}=2,{{col_s}}{{{{row}}}}=16,{{col_s}}{{{{row}}}}=17,{{col_s}}{{{{row}}}}=44,{{col_s}}{{{{row}}}}=86,{{col_s}}{{{{row}}}}=1,{{col_s}}{{{{row}}}}="1S",{{col_s}}{{{{row}}}}="3Z",{{col_s}}{{{{row}}}}="X"),{{col_s}}{{{{row}}}},IF(AND(OR({{col_c}}{{{{row}}}}=0,{{col_c}}{{{{row}}}}=1),OR({{col_s}}{{{{row}}}}="5X",{{col_s}}{{{{row}}}}="5Y",{{col_s}}{{{{row}}}}="6B",{{col_s}}{{{{row}}}}="6C")),{{col_s}}{{{{row}}}},IF({{col_c}}{{{{row}}}}<>2,"",IF({{col_s}}{{{{row}}}}="5X","5XK",IF({{col_s}}{{{{row}}}}="5Y","5YK",IF({{col_s}}{{{{row}}}}="6B","6BK",IF({{col_s}}{{{{row}}}}="6C","6CK","")))))))'.format(col_c=col_c, col_s=col_s),  # W
            24: f'=IF({{col_s}}{{{{row}}}}="6B",_xlfn.XLOOKUP({{col_h}}{{{{row}}}},\'Zoning 2shop\'!A:A,\'Zoning 2shop\'!B:B),IF({{col_s}}{{{{row}}}}="6C",_xlfn.XLOOKUP({{col_g}}{{{{row}}}},\'Zoning 2shop\'!C:C,\'Zoning 2shop\'!D:D),""))'.format(col_s=col_s, col_g=col_g, col_h=col_h),  # X
            25: "=AB{row}/Z{row}",  # Y (Gazole %) = gazole / fret
            26: f"={{col_t}}{{{{row}}}}".format(col_t=col_t),  # Z (fret)
            27: f'=IF(OR({{col_s}}{{{{row}}}}="6C",{{col_s}}{{{{row}}}}="6B",{{col_s}}{{{{row}}}}="5X",{{col_s}}{{{{row}}}}="5Y"),0.08,0.5)'.format(col_s=col_s),  # AA (taux surete+eco)
            28: "=$AG$5/$AG$2*Z{row}",  # AB (gazole reparti -- pool/total fret * fret ligne)
            29: "=Z{row}+AA{row}",  # AC (hors gazole)
            30: f'=IF(COUNTIF(Catégories!A:A,{{col_n}}{{{{row}}}})=0,"catégorie inconnue",LOOKUP({{col_n}}{{{{row}}}},Catégories!A:A,Catégories!B:B))'.format(col_n=col_n),  # AD (Categories)
            31: "=SUM(Z{row}:AB{row})",  # AE (Total avec GO)
        }
        for col_idx, tmpl in formulas_w_ae.items():
            retry(lambda c=col_idx, t=tmpl: ws.Range(ws.Cells(2, c), ws.Cells(newLast, c))
                  .__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))

        # Liste des 9 postes ERP EN DUR en colonne AD, 2 lignes apres la derniere ligne de
        # donnees (confirme sur 2026_07_Facture Chronopost.xlsx : derniere ligne de donnees
        # 4149, lignes 4150-4151 vides, liste en 4152-4160 -- source de la liste de validation
        # utilisee ailleurs dans le classeur). Reference de table qui NE DOIT JAMAIS rester
        # vide pour les prochaines facturations (decision utilisateur 2026-08-13) -- le modele
        # de juin ne l'avait pas, cette liste etait donc absente de toute generation.
        POSTE_KEYS_LISTE = ["Adresse", "Assurance", "Colis volumineux", "Corse", "Droits et taxes", "Frais facturation", "Frêt", "Gazole", "Zones éloignées"]
        for i, poste in enumerate(POSTE_KEYS_LISTE):
            ws.Cells(newLast + 3 + i, 30).Value = poste

        # Zone recap AG2:AG5 : VRAIES FORMULES Excel SUMIF sur "Type prestation" (colonne N),
        # PAS des valeurs Python (decision utilisateur 2026-08-13 : "pour les colonnes AF et
        # AG faut laisser les formules dans les cellules"). SUMIF sur le LIBELLE plutot que
        # SUM sur une plage fixe recalee a la main (methode du fichier reel, cf. transcription
        # video "AF2:AF5 = SOMME(T<debut>:T<fin>) sur plage FIXE... recalee a la main apres
        # tri") : verifie sur 2026_07_Facture Chronopost.xlsx que la plage fixe AG2=SUM(T31:
        # T5000) EMPIETE en fait sur les lignes SURT* (31-33 sont encore forfaitaires a ce
        # moment-la, pas des lignes Transport) -- erreur de calage manuel qui gonfle le Fret
        # de ~130€. SUMIF(Type prestation) est ROBUSTE a ce risque (independant du tri/de la
        # position des lignes) : AG2=SUMIF(N,"Transport",T), AG3=SUMIF(N,"Participation
        # Eco-Responsable",T), AG4=SUMIF(N,"Sûreté colis",T) (libelles confirmes via la table
        # 'categories' de config.json). AG5 (Gazole, pool CAP*) reste par PREFIXE Numero LT
        # (2 libelles distincts "Surcharge Carburant Aérien"/"Routier" regroupes sous un seul
        # poste "Gazole") -- SUMIF sur les 2 libelles cumules.
        # CONSEQUENCE ATTENDUE (2026-08-13, confirme utilisateur) : la colonne AB "gazole"
        # (=$AG$5/$AG$2*Z{row}) depend directement de ce AG2 -- comme notre AG2 est maintenant
        # STRICTEMENT "Transport" (vs le fichier reel dont l'AG2 mal cale inclut ~130€ de
        # lignes SURT*), le taux AG5/AG2 differe legerement du fichier reel (0,109 ici vs
        # 0,101 dans '2026_07_Facture Chronopost.xlsx'), ce qui cree un ecart cumule de
        # plusieurs centaines d'euros sur la colonne AB par rapport A CE FICHIER REEL
        # SPECIFIQUE -- CE N'EST PAS UN BUG : c'est le fichier reel qui a l'erreur de calage,
        # notre AG2 est correct par construction (SUMIF, insensible au tri/position).
        # AG2 (Frêt) : SUMIF sur PLUSIEURS libelles "Type prestation" (pas seulement
        # "Transport") -- liste confirmee par capture ecran du filtre AutoFilter reel du
        # classeur (2026-08-14, cases cochees/decochees) : toutes les lignes cochees entrent
        # dans le Frêt, INDEPENDAMMENT de leur classement dans la table "Categories" (qui sert
        # au TCD/reclassement final ERP, un usage different -- ex. "Supplement Corse 18h" est
        # classe poste "Corse" dans Categories mais reste inclus dans AG2 ici). Cases
        # DECOCHEES sur la capture (donc PAS dans AG2) : "Participation Eco-Responsable"
        # (deja AG3), "Sûreté colis" (deja AG4), "Surcharge Carburant Aérien"/"Routier" (deja
        # AG5) -- coherent, chacune deja comptee ailleurs dans la zone recap.
        FRET_TYPES_PRESTATION = [
            "Transport", "Correction d'adresse", "Supp Retour Expediteur Europe",
            "Supp Retour Expediteur Inter", "Supp Zone Internationale Eloignee",
            "Supplement Annonce incomplète", "Supplement Corse 18h",
            "Supplement domicile prive", "Supplement Douane Zone C4",
            "Supplement Etiquette Non Conforme", "Supplement Forfait Expedition",
            "Supplement GT", "Supplement Manutention", "Supplement Retour Expediteur",
            "Supplement Retrait Bureau", "Traitement SAV complémentaire",
            "Zones Difficiles d'accès",
        ]
        ws.Cells(2, 33).Formula = "=" + "+".join(
            f'SUMIF({col_n}:{col_n},"{t}",{col_t}:{col_t})' for t in FRET_TYPES_PRESTATION
        )  # AG2 Frêt
        ws.Cells(3, 33).Formula = f'=SUMIF({col_n}:{col_n},"Participation Eco-Responsable",{col_t}:{col_t})'  # AG3 eco
        ws.Cells(4, 33).Formula = f'=SUMIF({col_n}:{col_n},"Sûreté colis",{col_t}:{col_t})'  # AG4 sûreté
        ws.Cells(5, 33).Formula = (
            f'=SUMIF({col_n}:{col_n},"Surcharge Carburant Aérien",{col_t}:{col_t})'
            f'+SUMIF({col_n}:{col_n},"Surcharge Carburant Routier",{col_t}:{col_t})'
        )  # AG5 Gazole
        print("Zone récap : formules SUMIF ecrites en AG2 (Frêt) / AG3 (eco) / AG4 (sûreté) / AG5 (Gazole).")

        # Reactive l'AutoFilter sur toute la largeur/hauteur utile (desactive plus haut, ligne
        # 325, pour eviter le piege ClearContents/lignes masquees -- jamais reactive avant ce
        # fix). BUG TROUVE 2026-08-13 (capture ecran utilisateur : en-tete "Facture Chronopost"
        # retrouve au MILIEU du tableau apres un tri manuel A->Z dans Excel) -- sans AutoFilter
        # actif sur la ligne 1, Excel ne distingue plus cette ligne comme un en-tete special
        # lors d'un tri par selection, et peut la trier comme une ligne de donnee ordinaire.
        # Le modele reel a TOUJOURS cet AutoFilter actif (confirme sur 2026_07_Facture
        # Chronopost.xlsx) -- notre sortie en etait depourvue depuis l'introduction du fix
        # AutoFilterMode=False (necessaire pour ClearContents/Value=, mais jamais restaure).
        ws.Range(ws.Cells(1, 1), ws.Cells(newLast, LAST_CALC_COL)).AutoFilter()

        # 2) "TCD poids"/"Contrôle pdf"/"TCD" : PivotCache redirige vers TOUTE la largeur utile
        #    de "Facture Chronopost" (B->AF), PAS la plage figee/etroite du modele (cf.
        #    docstring de redirect_pivot_caches -- meme piege que DPD/Geodis/Mondial Relay).
        #
        #    "TCD" etait EXCLU de cette redirection jusqu'au 2026-08-14 (limite documentee
        #    2026-08-12 : #VALUE! partout en test a l'epoque) -- RE-TESTE le 2026-08-14 (a la
        #    demande utilisateur, pour changer son 2e RowField 'Produit'->'Code produit
        #    modifie') : la redirection fonctionne desormais SANS #VALUE! (cause de l'ancien
        #    echec non identifiee avec certitude -- possible etat different du classeur a
        #    l'epoque -- mais confirme reproductible sur plusieurs essais). Reintegre a la
        #    redirection standard.
        #
        #    BUG TROUVE 2026-08-14 : la plage redirigee s'arretait a newLast, EXCLUANT la
        #    liste des 9 postes ERP ecrite juste apres (newLast+3..newLast+11, cf. plus haut)
        #    -- consequence, le PivotCache "TCD" (colField "Catégories") ne connaissait QUE
        #    les categories reellement presentes ce mois-ci (souvent 5 a 7 sur 9), les autres
        #    n'apparaissant PAS du tout dans le tableau (contrairement au fichier fait-main de
        #    reference, ou les 9 colonnes sont TOUJOURS affichees, a 0 si aucune donnee).
        #    Fix : la plage redirigee inclut desormais cette liste (jusqu'a newLast+11) --
        #    chaque ligne de la liste n'a NI "No Facture" NI "Numero LT" renseignes, donc reste
        #    invisible pour toute la logique metier en aval (Fichier import, reconciliation
        #    PDF, zone recap AG2:AG5 -- toutes filtrent explicitement sur ces champs ou sur
        #    "Type prestation"='Transport', jamais rempli sur ces lignes-legende).
        redirect_pivot_caches(wb, ws, ["TCD poids", "Contrôle pdf", "TCD"], "Facture Chronopost",
                               max(newLast + 11, 2), FIRST_RAW_COL, LAST_CALC_COL)
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        xl.Calculate()

        # "TCD" 2e RowField : 'Produit' -> 'Code produit modifie' (demande utilisateur
        # 2026-08-14, cf. capture ecran du TCD reel montrant deja cette colonne en usage).
        # Necessite que la redirection ci-dessus ait deja rendu 'Code produit modifie'
        # disponible dans PivotFields() (absent de la plage figee d'origine du modele).
        xlHidden, xlRowField = 0, 1
        wsTcd = wb.Sheets("TCD")
        ptTcd = wsTcd.PivotTables(1)
        try:
            ptTcd.PivotFields("Produit").Orientation = xlHidden
            codeProduitField = ptTcd.PivotFields("Code produit modifié")
            codeProduitField.Orientation = xlRowField
            codeProduitField.Position = 2
            # Desactive les sous-totaux automatiques Excel sur ce champ ("Total 5YK"/"Total
            # 3Z"...) -- sans ca, le pivot intercale 1 ligne de sous-total par groupe et
            # double quasiment le nombre de lignes affichees (3525->7047 lignes constate en
            # test), alors que 'Produit' (champ precedent) n'avait pas ce comportement visible
            # (decision utilisateur 2026-08-14 : revenir a un tableau simple ligne par ligne).
            codeProduitField.Subtotals = tuple([False] * 12)
            wb.RefreshAll()
            try:
                xl.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            xl.Calculate()
            print("'TCD' : 2e RowField changé de 'Produit' vers 'Code produit modifié'.")
        except Exception as e:
            print(f"AVERTISSEMENT: impossible de changer le RowField 'TCD' vers 'Code produit modifié' ({e}) -- reste sur 'Produit'.")

        # "TCD poids" colonne C "Poids arrondi" (=ROUNDUP(B{row},1)) : BUG TROUVE 2026-08-13
        # (capture ecran utilisateur) -- cette formule manuelle juxtaposee au TCD natif
        # restait FIGEE a la taille du modele de juin (2209 lignes) apres redirection du
        # PivotCache, alors que le TCD natif (colonnes A/B) s'etend correctement a la vraie
        # taille du mois traite (3940+ lignes en juillet 2026) -- laissant ~1700 lignes de
        # "Poids arrondi" vides en bas du tableau. Sans impact sur "Fichier import" (qui lit
        # le poids via XLOOKUP sur toute la colonne, pas par position), mais incomplet
        # visuellement. Etiree ici jusqu'a la meme derniere ligne que la colonne B native.
        wsTcdPoids = wb.Sheets("TCD poids")
        lastTcdPoids = wsTcdPoids.Cells(wsTcdPoids.Rows.Count, 2).End(xlUp).Row
        if lastTcdPoids >= 2:
            retry(lambda: wsTcdPoids.Range(wsTcdPoids.Cells(2, 3), wsTcdPoids.Cells(lastTcdPoids, 3))
                  .__setattr__("Formula", [["=ROUNDUP(B{row},1)".format(row=r)] for r in range(2, lastTcdPoids + 1)]))
            print(f"'TCD poids' : colonne C 'Poids arrondi' etiree jusqu'a la ligne {lastTcdPoids}.")

        # "TCD" colonnes A/B/C (Total GO / Total avec CAP+ECO / Total hors GO) et S/T/U/V/W
        # (sureté+eco, mode envoi, zone, Transporteur, Frêt+CAP+ECO hors gazole) : MEME BUG
        # que "TCD poids" ci-dessus -- ces formules manuelles restaient FIGEES a la taille du
        # modele de juin (~2210 lignes) alors que la colonne NATIVE E ("Somme de Montant HT")
        # s'actualise deja correctement a la vraie taille du mois traite (verifie : 3942
        # lignes en juillet 2026, cf. wb.RefreshAll() plus haut qui touche tous les
        # PivotCache du classeur meme sans redirection explicite pour "TCD"). Consequence
        # visible (capture ecran utilisateur 2026-08-13) : colonnes S-W vides + #DIV/0! des
        # que la formule W (qui divise par C) rencontre une ligne sans sa contrepartie A/B/C.
        # Colonne D ("ID client") reste NON etiree : ce sont des VALEURS BRUTES figees (ex.
        # 6235, 8656), pas des formules -- aucune source automatisable identifiee dans le
        # classeur (decision utilisateur 2026-08-13, pas de donnee inventee).
        wsTcd = wb.Sheets("TCD")
        lastTcd = wsTcd.Cells(wsTcd.Rows.Count, 5).End(xlUp).Row  # colonne E native
        # Ligne 1 = titres generaux (Total GO/ID client/...), ligne 2 = EN-TETES du TCD natif
        # (Numero LT/Produit/TVA/...) -- les DONNEES commencent en ligne 3, PAS 2 (piege trouve
        # en test : ecrire la formule sur la ligne 2 ecrase l'en-tete et produit un #VALUE! qui
        # se propage en cascade dans toutes les lignes suivantes via SUM(C:C)).
        if lastTcd >= 3:
            formulas_tcd = {
                1: "=C{row}/SUM(C:C)*(SUM(C:C)+$X$2)",  # A (Total GO)
                2: "=C{row}/SUM(C:C)*(SUM(C:C)+SUM($N$3:$O$7))+S{row}",  # B (Total avec CAP+ECO)
                3: "=SUM(I{row}:Q{row})-N{row}-P{row}",  # C (Total hors GO)
                19: "=_xlfn.XLOOKUP(E{row},'Facture Chronopost'!L:L,'Facture Chronopost'!AA:AA,\"\")",  # S (sûreté+eco)
                20: '=IF(COUNTIF(\'Bibliothèque transporteurs\'!C:C,F{row})=0,"inconnu",_xlfn.XLOOKUP(F{row},\'Bibliothèque transporteurs\'!C:C,\'Bibliothèque transporteurs\'!A:A))',  # T (mode envoi)
                21: '=IF(OR(F{row}="6C",F{row}="6B",F{row}="6BK",F{row}="6CK"),_xlfn.XLOOKUP(E{row},\'Facture Chronopost\'!L:L,\'Facture Chronopost\'!X:X),IF(OR(F{row}=17,F{row}=44),F{row}&"_"&H{row},IF(COUNTIF(\'Bibliothèque transporteurs\'!C:C,F{row})=0,"inconnu",_xlfn.XLOOKUP(F{row},\'Bibliothèque transporteurs\'!C:C,\'Bibliothèque transporteurs\'!B:B))))',  # U (zone)
                22: '=IF(COUNTIF(\'Bibliothèque transporteurs\'!C:C,F{row})=0,"inconnu",_xlfn.XLOOKUP(F{row},\'Bibliothèque transporteurs\'!C:C,\'Bibliothèque transporteurs\'!E:E))',  # V (Transporteur)
                23: '=IF(T{row}="FR_2SHOP",O{row},IF(O{row}=0,0,ROUNDUP(O{row}/C{row}*B{row},2)))',  # W (Frêt+CAP+ECO hors gazole)
            }
            for col_idx, tmpl in formulas_tcd.items():
                retry(lambda c=col_idx, t=tmpl: wsTcd.Range(wsTcd.Cells(3, c), wsTcd.Cells(lastTcd, c))
                      .__setattr__("Formula", [[t.format(row=r)] for r in range(3, lastTcd + 1)]))
            print(f"'TCD' : colonnes A/B/C/S/T/U/V/W etirees jusqu'a la ligne {lastTcd}.")

        # 3) "Fichier import" : formules FIXES du modele fait-main, referencant TCD/TCD poids
        #    PAR POSITION (ligne n de "Fichier import" -> ligne n+1 de "TCD", decalage
        #    constant du a l'entete sur 2 lignes du TCD natif) -- decision utilisateur
        #    2026-08-14 : garder la feuille EXCEL "vivante" (recalcul automatique si le TCD
        #    change, ex. apres saisie manuelle des ID clients), reserver l'ecriture en VALEURS
        #    PURES au CSV livre au client (deja fait cote Node.js, writeImportCsv/excelOut.js,
        #    non touche ici). Formules reprises A L'IDENTIQUE du modele, y compris 2 points
        #    DELIBEREMENT NON corriges malgre un ecart connu avec notre logique Python deja
        #    validee a 0,00€ (choix utilisateur explicite, prevaut sur la logique validee) :
        #      - H (E/P) = IF(RIGHT(J,3)="BTB","E","P") -- le modele utilise le suffixe de
        #        Zone, PAS le mode envoi='1S' confirme par le pole transport (cf. epPourMode
        #        cote index.js/finaliseur precedent, source de 64 avaries evitees en juillet
        #        2026 en corrigeant precisement ce point pour le CSV/carrier Node -- cette
        #        feuille Excel EN REVANCHE reste fidele au modele, sur demande explicite).
        #      - T (Frêt) = TCD!Y{row} -- la colonne Y du TCD actuel n'a AUCUN en-tete (la
        #        vraie donnee "Frêt + CAP + ECO... hors gazole" est en colonne W) -- reference
        #        cassee/obsolete du modele, reproduite TELLE QUELLE (pas corrigee vers W) sur
        #        demande utilisateur explicite -- "Frêt" restera donc vide sur cette feuille
        #        tant que cette reference n'est pas corrigee manuellement dans le modele.
        #    Le TCD ne regroupe pas non plus les lignes multiples par tracking (contrairement
        #    au carrier Node/CSV final) -- une meme consequence du choix "formules du modele
        #    sans exception" : le nombre de lignes de "Fichier import" suit celui du TCD natif
        #    (une ligne par ligne source de "Facture Chronopost", PAS par tracking unique).
        wsImp = wb.Sheets("Fichier import")
        if wsImp.AutoFilterMode:  # meme piege que "Facture Chronopost" ci-dessus
            wsImp.AutoFilterMode = False
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 6).End(xlUp).Row
        LAST_COL_IMPORT = 21  # A->U (21 colonnes du modele -- V/W = gazole/note 2SHOP, jamais remplies)

        newLastImp = max(lastTcd - 1, 2)  # ligne n (import) <-> ligne n+1 (TCD) ; TCD data commence ligne 3
        formulas_import = {
            1: "=TCD!V{tcdrow}",  # A Transporteur
            6: "=TCD!E{tcdrow}",  # F N° Tracking
            8: '=IF(RIGHT(J{row},3)="BTB","E","P")',  # H E/P (fidele au modele, cf. docstring)
            9: "=LOOKUP(F{row},'Facture Chronopost'!L:L,'Facture Chronopost'!H:H)",  # I Pays
            10: "=TCD!U{tcdrow}",  # J Zone
            12: "=_xlfn.XLOOKUP(F{row},'TCD poids'!A:A,'TCD poids'!C:C)",  # L Poids
            13: "=TCD!T{tcdrow}",  # M mode envoi
            14: "=IF(I{row}=\"\",0.2,LOOKUP(I{row},'Pays TVA'!A:A,'Pays TVA'!B:B))",  # N TVA
            15: '=IF(TCD!M{tcdrow}="","",TCD!M{tcdrow})',  # O Droits et taxes
            16: '=IF(TCD!J{tcdrow}="","",TCD!J{tcdrow})',  # P Assurance
            17: '=IF(AND(TCD!L{tcdrow}="",TCD!Q{tcdrow}=""),"",IF(SUM(TCD!L{tcdrow},TCD!Q{tcdrow})>15,29,9.5))',  # Q Zones éloignées
            18: '=_xlfn.XLOOKUP(F{row},TCD!E:E,TCD!K:K,"")',  # R Colis volumineux
            19: '=IF(TCD!I{tcdrow}="","",IF(TCD!I{tcdrow}>8.5,17,8.5))',  # S Adresses
            20: "=TCD!Y{tcdrow}",  # T Frêt (reference telle quelle du modele, cf. docstring)
        }
        # K (Nbr Colis) : valeur fixe 1 en ligne 2, formule "=K2" recopiee ensuite (comme le modele).
        wsImp.Cells(2, 11).Value = 1
        if newLastImp > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(3, 11), wsImp.Cells(newLastImp, 11))
                  .__setattr__("Formula", [["=K{prev}".format(prev=r - 1)] for r in range(3, newLastImp + 1)]))
        # B (Date validité tarif) : valeur fixe en ligne 2 (mise a jour vers le mois traite),
        # formule "=B{n-1}" recopiee ensuite (meme piege deja corrige pour DPD/Mondial Relay/BLS).
        if date_validite_serial is not None:
            wsImp.Cells(2, 2).Value = date_validite_serial
        else:
            print("AVERTISSEMENT: 'Date LT' introuvable -> 'Fichier import'!B2 non mise à jour, reste celle du modèle.")
        if newLastImp > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(3, 2), wsImp.Cells(newLastImp, 2))
                  .__setattr__("Formula", [["=B{prev}".format(prev=r - 1)] for r in range(3, newLastImp + 1)]))
        for col_idx, tmpl in formulas_import.items():
            retry(lambda c=col_idx, t=tmpl: wsImp.Range(wsImp.Cells(2, c), wsImp.Cells(newLastImp, c))
                  .__setattr__("Formula", [[t.format(row=r, tcdrow=r + 1)] for r in range(2, newLastImp + 1)]))
        if newLastImp < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLastImp + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT)).ClearContents())
        print(f"'Fichier import' : formules du modele reconstruites jusqu'a la ligne {newLastImp} (TCD ligne {newLastImp + 1}).")

        # Reactive l'AutoFilter (desactive plus haut, meme piege que "Facture Chronopost").
        wsImp.Range(wsImp.Cells(1, 1), wsImp.Cells(max(newLastImp, 2), LAST_COL_IMPORT)).AutoFilter()

        xl.Calculate()

        # 4) Réconciliation PDF (onglet "Contrôle pdf", colonne C = 'pdf' saisie a la main
        #    dans le modele, en face de No Facture en colonne A du TCD).
        try:
            ws_pdf = wb.Sheets("Contrôle pdf")
            lastA = ws_pdf.Cells(ws_pdf.Rows.Count, 1).End(xlUp).Row
            facture_to_row = {}
            for r in range(1, lastA + 1):
                v = str(ws_pdf.Cells(r, 1).Value or "").strip()
                if v and v not in ("Étiquettes de lignes", "(vide)", "Total général"):
                    facture_to_row[v] = r
            lastC = ws_pdf.Cells(ws_pdf.Rows.Count, 3).End(xlUp).Row
            if lastC >= 4:
                ws_pdf.Range(ws_pdf.Cells(4, 3), ws_pdf.Cells(lastC, 3)).ClearContents()
            matched = 0
            for p in pdfs:
                row = facture_to_row.get(p["facture"])
                if row is None:
                    print(f"AVERTISSEMENT: facture PDF {p['file']} ({p['facture']}) -> introuvable dans Contrôle pdf.")
                    continue
                ws_pdf.Cells(row, 3).Value = p["total_ht"]
                matched += 1
            print(f"Réconciliation PDF : {matched}/{len(pdfs)} facture(s) rapprochée(s) dans Contrôle pdf.")
        except Exception as e:
            print("Réconciliation Contrôle pdf ignorée :", e)

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)
        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
