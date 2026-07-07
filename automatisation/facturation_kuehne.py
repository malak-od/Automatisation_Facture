#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
 AUTOMATISATION FACTURATION KUEHNE+NAGEL
============================================================================
Transforme les CSV bruts K&N en fichier d'import ERP, exactement comme le
classeur Excel manuel (validé 187/187 lignes contre l'export réel juin 2026).

PIPELINE (voir Documentation_Facturation_Kuehne.md) :
   CSV bruts K&N ─► reclassement (T_* → 8 postes) ─► filtrage tracking vide
   ─► calcul zone / champs fixes ─► validation ─► fichier d'import ERP

Toute la logique métier (quel T_* va dans quel poste, champs fixes, TVA…)
est dans le fichier de config JSON — le code, lui, ne change pas d'un
transporteur à l'autre. Pour ajouter un transporteur : nouveau fichier config.

USAGE :
   python facturation_kuehne.py --input ../Kuehne --output import.csv
   python facturation_kuehne.py --input ../Kuehne --check ../Kuehne/2026_06_Kuehne_Import.csv
============================================================================
"""
import csv, glob, os, json, argparse
from collections import defaultdict, Counter

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CONFIG = os.path.join(HERE, "reclassement_config.json")

# En-têtes du fichier d'import ERP (23 colonnes, dans l'ordre attendu)
IMPORT_HEADER = [
    " Transporteur ", "Date validité tarif", "Réf.1", "Réf. 2", "Id client",
    "N° Tracking", "Nom", "E / P", "Pays", " Zone ", " Nbr Colis ", " Poids ",
    "mode envoi", " TVA ", " Droits et taxes ", " Assurance ", " Zones éloignées ",
    " Colis volumineux ", " Adresses ", "    Frêt    ", " plus-value BtoC ",
    "Taxe Gasoil", " Nb colis ",
]

# --------------------------------------------------------------------------
# 1. CONFIG
# --------------------------------------------------------------------------
def load_config(path=DEFAULT_CONFIG):
    with open(path, encoding="utf-8") as f:
        return json.load(f)

# --------------------------------------------------------------------------
# 2. LECTURE DES CSV BRUTS
# --------------------------------------------------------------------------
def read_raw(input_dir, cfg):
    """Concatène tous les CSV 'FcCSV*' du dossier. Renvoie (lignes, index_colonnes)."""
    enc, sep = cfg["csv"]["encoding"], cfg["csv"]["sep"]
    rows, header = [], None
    for path in sorted(glob.glob(os.path.join(input_dir, "FcCSV*.csv"))):
        with open(path, encoding=enc, newline="") as fh:
            data = list(csv.reader(fh, delimiter=sep))
        header = data[0]
        for r in data[1:]:
            if len(r) < len(header):            # ligne tronquée → complète
                r += [""] * (len(header) - len(r))
            rows.append(r)
    # index : nom de colonne → position (on mappe par NOM, tolérant aux espaces)
    idx = {name: i for i, name in enumerate(header)}
    return rows, idx, header

def num(x):
    """Convertit un montant CSV ('1 234,56' ou '') en float."""
    x = (x or "").strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(x)
    except ValueError:
        return 0.0

def col(idx, header, wanted):
    """Retourne l'index d'une colonne par nom (match exact puis 'contient')."""
    if wanted in idx:
        return idx[wanted]
    for i, name in enumerate(header):
        if wanted in name:
            return i
    raise KeyError(f"Colonne introuvable dans le CSV : {wanted!r}")

# --------------------------------------------------------------------------
# 3. RECLASSEMENT (le cœur — piloté par la config)
# --------------------------------------------------------------------------
def reclasser(r, idx, header, cfg, warnings):
    """Somme les colonnes T_* en 8 postes, selon postes_from_columns."""
    postes = {}
    used = set()
    for poste, cols in cfg["postes_from_columns"].items():
        total = 0.0
        for cname in cols:
            try:
                i = col(idx, header, cname); used.add(cname); total += num(r[i])
            except KeyError:
                pass
        postes[poste] = round(total, 2)
    hors_go = round(sum(postes.values()), 2)
    avec = num(r[col(idx, header, cfg["montant_avec_frais"])])
    postes["Gazole"] = round(avec - hors_go, 2)         # résiduel (cf. doc)

    # garde-fou : détecte un code tarifaire non classé (fuiterait dans le gazole)
    for i, name in enumerate(header):
        if name.startswith("T_") and name not in used and name != cfg.get("colonne_gazole"):
            if num(r[i]) != 0:
                warnings.append(f"Code tarifaire NON classé '{name}'={num(r[i])} "
                                f"(fuite dans Gazole) — à ajouter dans la config")
    return postes

# --------------------------------------------------------------------------
# 4. ZONE
# --------------------------------------------------------------------------
def build_dept_lookup(rows, idx, header, cfg):
    """Reproduit le XLOOKUP Excel : dept indexé par 'Ref EDI' uniquement."""
    i_edi = col(idx, header, cfg["tracking"]["primary"])
    i_de = col(idx, header, cfg["identite"]["dept_exp"])
    i_dd = col(idx, header, cfg["identite"]["dept_dest"])
    de, dd = {}, {}
    for r in rows:
        e = r[i_edi].strip()
        if e:
            de.setdefault(e, r[i_de].strip()); dd.setdefault(e, r[i_dd].strip())
    return de, dd

def zone(track, pays, de_map, dd_map):
    de = de_map.get(track, "0") or "0"
    dd = dd_map.get(track, "0") or "0"
    try: de = str(int(de))
    except ValueError: pass
    try: dd = f"{int(dd):02d}"          # seul le dept DESTINATAIRE est paddé
    except ValueError: pass
    return f"{de}{'France' if pays == 'FR' else 'étranger'}{dd}"

# --------------------------------------------------------------------------
# 5. CONSTRUCTION DES LIGNES D'IMPORT
# --------------------------------------------------------------------------
def build_import(rows, idx, header, cfg):
    i_edi = col(idx, header, cfg["tracking"]["primary"])
    i_com = col(idx, header, cfg["tracking"]["fallback"])
    i_dest = col(idx, header, cfg["identite"]["destinataire"])
    i_pays = col(idx, header, cfg["identite"]["pays_dest"])
    i_poids = col(idx, header, cfg["identite"]["poids"])
    i_colis = col(idx, header, cfg["identite"]["colis"])
    de_map, dd_map = build_dept_lookup(rows, idx, header, cfg)
    fx = cfg["champs_fixes"]; hors_ue = set(cfg["tva"]["pays_hors_ue"])

    out, warnings, controle = [], [], defaultdict(float)
    for r in rows:
        postes = reclasser(r, idx, header, cfg, warnings)
        for k, v in postes.items():
            controle[k] += v
        edi = r[i_edi].strip()
        track = edi if edi else r[i_com].strip()
        if not track:                       # RÈGLE : exclusion tracking vide
            continue
        pays = r[i_pays].strip()
        z = zone(track, pays, de_map, dd_map)
        row = {
            " Transporteur ": fx["Transporteur"],
            "Date validité tarif": fx["date_validite"],
            "Réf.1": r[i_com].strip(), "Réf. 2": "", "Id client": "",
            "N° Tracking": track, "Nom": r[i_dest].strip(),
            "E / P": fx["E/P"], "Pays": pays, " Zone ": z,
            " Nbr Colis ": int(round(num(r[i_colis]))),
            " Poids ": round(num(r[i_poids]), 1),
            "mode envoi": fx["mode envoi"],
            " TVA ": cfg["tva"]["hors_ue"] if pays in hors_ue else cfg["tva"]["defaut"],
            " Droits et taxes ": postes["Droits et taxes"],
            " Assurance ": postes["Assurance"],
            " Zones éloignées ": postes["Zones eloignees"],
            " Colis volumineux ": postes["Colis volumineux"],
            " Adresses ": postes["Adresses"],
            "    Frêt    ": postes["Fret"],
            " plus-value BtoC ": postes["Plus value B2C"],
            "Taxe Gasoil": "",              # gazole refacturé hors import
            " Nb colis ": "",
        }
        out.append(row)
    return out, controle, warnings

# --------------------------------------------------------------------------
# 6. VALIDATION (check-list opérateur → automatique)
# --------------------------------------------------------------------------
def validate(rows):
    alerts = []
    for i, o in enumerate(rows, start=2):
        t = o["N° Tracking"]
        if o[" Poids "] in (0, 0.0):        alerts.append(f"L{i} {t}: POIDS = 0")
        if o[" Nbr Colis "] == 0:           alerts.append(f"L{i} {t}: COLIS = 0")
        if o["    Frêt    "] in (0, 0.0) and not any(
            o[k] for k in (" Droits et taxes "," Assurance "," Zones éloignées ",
                           " Colis volumineux "," Adresses "," plus-value BtoC ")):
            alerts.append(f"L{i} {t}: FRÊT = 0 (ligne sans montant)")
        if "0France00" in o[" Zone "] or "0étranger00" in o[" Zone "]:
            alerts.append(f"L{i} {t}: ZONE inconnue ({o[' Zone ']}) — dept non trouvé")
        if "étranger" in o[" Zone "]:       alerts.append(f"L{i} {t}: ZONE étranger ({o[' Zone ']})")
    return alerts

# --------------------------------------------------------------------------
# 7. EXPORT + AUTO-CONTRÔLE
# --------------------------------------------------------------------------
def fmt(v):
    if v == "" or v is None: return ""
    if isinstance(v, float):
        return ("%.2f" % v).rstrip("0").rstrip(".").replace(".", ",") if v else ""
    return str(v)

def export(rows, path, cfg):
    """Export CSV (décimale virgule, latin-1)."""
    with open(path, "w", encoding=cfg["csv"]["encoding"], newline="") as fh:
        w = csv.writer(fh, delimiter=cfg["csv"]["sep"])
        w.writerow(IMPORT_HEADER)
        for o in rows:
            w.writerow([fmt(o[h]) for h in IMPORT_HEADER])

# colonnes numériques : on écrit de vrais nombres dans le .xlsx (pas du texte)
NUM_COLS = {" Nbr Colis ", " Poids ", " TVA ", " Droits et taxes ", " Assurance ",
            " Zones éloignées ", " Colis volumineux ", " Adresses ", "    Frêt    ",
            " plus-value BtoC "}

def export_xlsx(rows, path):
    """Export Excel VALEURS SEULES (aucune formule) = le fichier à déposer dans l'ERP."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kuehne_Import"
    ws.append(IMPORT_HEADER)
    for o in rows:
        line = []
        for h in IMPORT_HEADER:
            v = o[h]
            if h in NUM_COLS:
                line.append(v if isinstance(v, (int, float)) and v != "" else None)
            else:
                line.append(v if v != "" else None)
        ws.append(line)
    wb.save(path)

def self_check(rows, reference_csv, cfg):
    """Compare (en multiset) la sortie générée à un export ERP de référence."""
    with open(reference_csv, encoding=cfg["csv"]["encoding"], newline="") as fh:
        ref = [r for r in list(csv.reader(fh, delimiter=cfg["csv"]["sep"]))[1:] if any(r)]
    def sig_ref(r):
        f = lambda s: round(num(s), 2)
        return (r[5].strip(), r[9].strip(), int(f(r[10])), f(r[11]),
                f(r[14]), f(r[15]), f(r[16]), f(r[17]), f(r[18]), f(r[19]), f(r[20]))
    def sig_gen(o):
        f = lambda v: round(float(v), 2) if isinstance(v, (int, float)) and v != "" else 0.0
        return (o["N° Tracking"], o[" Zone "], int(o[" Nbr Colis "]), round(o[" Poids "], 2),
                f(o[" Droits et taxes "]), f(o[" Assurance "]), f(o[" Zones éloignées "]),
                f(o[" Colis volumineux "]), f(o[" Adresses "]), f(o["    Frêt    "]),
                f(o[" plus-value BtoC "]))
    ce, cg = Counter(map(sig_ref, ref)), Counter(map(sig_gen, rows))
    ok = sum((ce & cg).values())
    return ok, len(ref), list((ce - cg).elements()), list((cg - ce).elements())

# --------------------------------------------------------------------------
# 8. CLASSEUR INTERMÉDIAIRE  2026_06_Facture Kuehne.xlsx  (livrable de contrôle)
# --------------------------------------------------------------------------
POSTE_KEYS = ["Droits et taxes","Assurance","Zones eloignees","Colis volumineux",
              "Adresses","Fret","Plus value B2C","Gazole"]

def process_lines(rows, idx, header, cfg):
    """Un enregistrement par ligne CSV brute (postes + identité + zone)."""
    i_edi=col(idx,header,cfg["tracking"]["primary"]); i_com=col(idx,header,cfg["tracking"]["fallback"])
    i_dest=col(idx,header,cfg["identite"]["destinataire"]); i_pays=col(idx,header,cfg["identite"]["pays_dest"])
    i_poids=col(idx,header,cfg["identite"]["poids"]); i_colis=col(idx,header,cfg["identite"]["colis"])
    de_map,dd_map=build_dept_lookup(rows,idx,header,cfg)
    recs=[]
    for r in rows:
        postes=reclasser(r,idx,header,cfg,[])
        edi=r[i_edi].strip(); track=edi if edi else r[i_com].strip()
        pays=r[i_pays].strip()
        recs.append({
            "tracking":track,"comref":r[i_com].strip(),"dest":r[i_dest].strip(),"pays":pays,
            "poids":num(r[i_poids]),"colis":num(r[i_colis]),"postes":postes,
            "hors_go":round(sum(v for k,v in postes.items() if k!="Gazole"),2),
            "avec_go":round(sum(postes.values()),2),
            "zone":zone(track,pays,de_map,dd_map) if track else "",
            "raw":r,
        })
    return recs

def read_pdf_totals(input_dir):
    """Extrait (taxable, TVA, TTC) de chaque facture PDF, via la relation TTC=taxable*1.2."""
    try:
        import pypdf
    except ImportError:
        return None
    import re
    res=[]
    for path in sorted(glob.glob(os.path.join(input_dir,"*.pdf"))):
        try:
            txt=pypdf.PdfReader(path).pages[0].extract_text() or ""
        except Exception:
            continue
        nums=[float(x.replace(" ","").replace("\xa0","").replace(",",".")) for x in re.findall(r"\d[\d \xa0]*,\d{2}",txt)]
        best=None
        for t in nums:
            if t<=0: continue
            tva,ttc=round(t*0.2,2),round(t*1.2,2)
            if any(abs(v-tva)<0.03 for v in nums) and any(abs(v-ttc)<0.03 for v in nums):
                if best is None or t>best[0]: best=(t,tva,ttc)
        if best:
            res.append({"file":os.path.basename(path),"taxable":best[0],"tva":best[1],"ttc":best[2]})
    return res

def _coerce(s):
    """Cellule brute : convertit les montants (virgule) en nombre, garde le reste en texte."""
    if isinstance(s,str) and "," in s:
        try: return float(s.replace(" ","").replace("\xa0","").replace(",","."))
        except ValueError: return s
    return s if s != "" else None

def build_workbook(rows, idx, header, cfg, imp, controle, warnings, alerts, input_dir, path):
    import openpyxl
    from openpyxl.styles import Font
    bold=Font(bold=True)
    recs=process_lines(rows,idx,header,cfg)
    wb=openpyxl.Workbook(); wb.remove(wb.active)

    # --- Feuille Fichier Kuehne : 8 postes calculés + données brutes ---
    ws=wb.create_sheet("Fichier Kuehne")
    comp=["ID Client","Tracking","Total hors GO","Total + GO","Droits et taxes","Assurance",
          "Zones éloignées","Colis volumineux","Adresses","Frêt","Plus value B2C","Gazole"]
    ws.append(comp+list(header))
    for c in ws[1]: c.font=bold
    for rec in recs:
        p=rec["postes"]
        line=[None,rec["tracking"],rec["hors_go"],rec["avec_go"],
              p["Droits et taxes"],p["Assurance"],p["Zones eloignees"],p["Colis volumineux"],
              p["Adresses"],p["Fret"],p["Plus value B2C"],p["Gazole"]]
        line+=[_coerce(v) for v in rec["raw"]]
        ws.append(line)

    # --- Feuille TCD : consolidation par tracking (contrôle) ---
    ws=wb.create_sheet("TCD")
    ws.append(["Tracking","ComRefExpedition","Destinataire","Pays","Zone",
               "Droits et taxes","Assurance","Zones éloignées","Colis volumineux","Adresses",
               "Frêt","Plus value B2C","Gazole","Total hors GO","Total + GO","Qté colis","Qté poids"])
    for c in ws[1]: c.font=bold
    agg={}
    for rec in recs:
        k=rec["tracking"]
        a=agg.setdefault(k,{"comref":rec["comref"],"dest":rec["dest"],"pays":rec["pays"],"zone":rec["zone"],
                            "postes":defaultdict(float),"hg":0.0,"ag":0.0,"colis":0.0,"poids":0.0})
        for pk,pv in rec["postes"].items(): a["postes"][pk]+=pv
        a["hg"]+=rec["hors_go"]; a["ag"]+=rec["avec_go"]; a["colis"]+=rec["colis"]; a["poids"]+=rec["poids"]
    for k,a in agg.items():
        P=a["postes"]
        ws.append([k,a["comref"],a["dest"],a["pays"],a["zone"],
                   round(P["Droits et taxes"],2),round(P["Assurance"],2),round(P["Zones eloignees"],2),
                   round(P["Colis volumineux"],2),round(P["Adresses"],2),round(P["Fret"],2),
                   round(P["Plus value B2C"],2),round(P["Gazole"],2),round(a["hg"],2),round(a["ag"],2),
                   int(round(a["colis"])),round(a["poids"],1)])

    # --- Feuille Kuehne_Import : VALEURS SEULES (= fichier ERP) ---
    ws=wb.create_sheet("Kuehne_Import")
    ws.append(IMPORT_HEADER)
    for c in ws[1]: c.font=bold
    for o in imp:
        ws.append([(o[h] if isinstance(o[h],(int,float)) and o[h]!="" else (o[h] or None))
                   if h in NUM_COLS else (o[h] if o[h]!="" else None) for h in IMPORT_HEADER])

    # --- Feuille Réconciliation : calcul vs PDF ---
    ws=wb.create_sheet("Réconciliation")
    ws["A1"]="RÉCONCILIATION"; ws["A1"].font=bold
    ws.append([]); ws.append(["Poste calculé","Montant €"])
    for k in POSTE_KEYS: ws.append([k,round(controle[k],2)])
    hg=round(sum(controle[k] for k in POSTE_KEYS if k!="Gazole"),2)
    ws.append(["Total hors gazole",hg]); ws.append(["Gazole",round(controle["Gazole"],2)])
    total_ht=round(sum(controle[k] for k in POSTE_KEYS),2)
    ws.append(["TOTAL HT calculé",total_ht]); ws.append([])
    pdfs=read_pdf_totals(input_dir)
    if pdfs:
        ws.append(["Facture PDF","Taxable","TVA","TTC"])
        s_tax=0.0
        for p in pdfs:
            ws.append([p["file"][:40],p["taxable"],p["tva"],p["ttc"]]); s_tax+=p["taxable"]
        ws.append(["Somme taxable PDF",round(s_tax,2)])
        ws.append(["ECART calcul - PDF",round(total_ht-s_tax,2),
                   "OK" if abs(total_ht-s_tax)<0.05 else "!!! ECART"])
    else:
        ws.append(["(PDF non lus : pypdf indisponible)"])
    ws.append([]); ws.append([f"Alertes de validation : {len(alerts)}"])
    for a in alerts: ws.append([a])
    if warnings:
        ws.append([f"Alertes reclassement : {len(warnings)}"])
        for w in warnings: ws.append([w])

    # --- Feuille Barème gazole : recopiée du modèle original si présent ---
    src=os.path.join(input_dir,"2026_06_Facture Kuehne.xlsx")
    if os.path.exists(src):
        try:
            m=openpyxl.load_workbook(src,data_only=True)
            if "Bareme gazole" in m.sheetnames:
                ms=m["Bareme gazole"]; ws=wb.create_sheet("Bareme gazole")
                for row in ms.iter_rows(values_only=True): ws.append(list(row))
        except Exception:
            pass

    wb.save(path)
    return total_ht, pdfs

def add_live_tcd_com(path):
    """Remplace la feuille TCD statique par un VRAI tableau croisé dynamique.
    Pilote Excel via COM → nécessite Windows + Excel + pywin32."""
    import win32com.client as win32
    xlUp, xlDatabase, xlRowField, xlSum = -4162, 1, 1, -4157
    abspath = os.path.abspath(path)
    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False; xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(abspath)
        src = wb.Sheets("Fichier Kuehne")
        last = src.Cells(src.Rows.Count, 2).End(xlUp).Row          # dernière ligne (col B)
        rng = src.Range(src.Cells(1, 1), src.Cells(last, 12))      # A1:L{last} = colonnes propres
        for s in list(wb.Sheets):                                  # supprime le TCD statique
            if s.Name == "TCD":
                s.Delete()
        tcd = wb.Sheets.Add(After=src); tcd.Name = "TCD"
        pc = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=rng)
        pt = pc.CreatePivotTable(TableDestination=tcd.Range("A3"), TableName="TCD_Kuehne")
        pt.PivotFields("Tracking").Orientation = xlRowField
        for p in ["Total hors GO", "Total + GO", "Droits et taxes", "Assurance",
                  "Zones éloignées", "Colis volumineux", "Adresses", "Frêt",
                  "Plus value B2C", "Gazole"]:
            pt.AddDataField(pt.PivotFields(p), "Somme " + p, xlSum)
        wb.Save(); wb.Close(SaveChanges=True)
    finally:
        xl.Quit()

# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Automatisation facturation Kuehne")
    ap.add_argument("--input", required=True, help="dossier contenant les CSV FcCSV*")
    ap.add_argument("--output", help="fichier CSV d'import à générer")
    ap.add_argument("--xlsx", help="fichier Excel d'import (valeurs seules) à déposer dans l'ERP")
    ap.add_argument("--workbook", help="classeur intermédiaire complet 2026_XX_Facture Kuehne.xlsx (livrable de contrôle)")
    ap.add_argument("--live-tcd", action="store_true", help="convertit la feuille TCD en vrai tableau croisé dynamique (Excel + pywin32, Windows)")
    ap.add_argument("--config", default=DEFAULT_CONFIG)
    ap.add_argument("--check", help="export ERP de référence pour auto-contrôle")
    args = ap.parse_args()

    cfg = load_config(args.config)
    rows, idx, header = read_raw(args.input, cfg)
    imp, controle, warnings = build_import(rows, idx, header, cfg)

    print(f"== {cfg['carrier']} : {len(imp)} lignes d'import générées ==")
    print("Totaux de contrôle (postes, gazole inclus) :")
    for k in ("Droits et taxes","Assurance","Zones eloignees","Colis volumineux",
              "Adresses","Fret","Plus value B2C","Gazole"):
        print(f"   {k:<18}= {round(controle[k],2)}")
    print(f"   {'TOTAL HT':<18}= {round(sum(controle.values()),2)}")

    alerts = validate(imp)
    if warnings:
        print(f"\n[!] {len(warnings)} alerte(s) reclassement :")
        for w in warnings[:10]: print("   ", w)
    print(f"\n[!] {len(alerts)} alerte(s) de validation :")
    for a in alerts[:15]: print("   ", a)

    if args.check:
        ok, tot, missing, extra = self_check(imp, args.check, cfg)
        print(f"\n== AUTO-CONTRÔLE vs référence : {ok}/{tot} lignes identiques "
              f"(manquantes {len(missing)}, surplus {len(extra)}) ==")

    def safe(label, fn):
        """Exécute un export ; un fichier verrouillé (ouvert dans Excel) n'interrompt pas les autres."""
        try:
            fn(); return True
        except PermissionError:
            print(f"[X] {label} : fichier VERROUILLÉ (ferme-le dans Excel puis relance)")
            return False

    if args.workbook:                       # livrable principal → généré en premier
        res = {}
        def _wb():
            res["v"] = build_workbook(rows, idx, header, cfg, imp, controle,
                                      warnings, alerts, args.input, args.workbook)
        if safe(f"Classeur {args.workbook}", _wb):
            total_ht, pdfs = res["v"]
            print(f"\nClasseur intermédiaire écrit : {args.workbook}")
            print(f"   Feuilles : Fichier Kuehne | TCD | Kuehne_Import (valeurs) | Réconciliation | Bareme gazole")
            if pdfs:
                s = round(sum(p['taxable'] for p in pdfs), 2)
                print(f"   Reconciliation : calcul {total_ht} EUR vs PDF {s} EUR -> ecart {round(total_ht-s,2)} EUR")
            if args.live_tcd:
                try:
                    add_live_tcd_com(args.workbook)
                    print("   TCD converti en tableau croise dynamique VIVANT (Excel COM)")
                except Exception as e:
                    print(f"   [X] TCD vivant impossible : {repr(e)[:130]}")
                    print("       (necessite Windows + Excel installe + 'pip install pywin32')")
    if args.xlsx:
        if safe(f"Import XLSX {args.xlsx}", lambda: export_xlsx(imp, args.xlsx)):
            print(f"Fichier d'import XLSX écrit : {args.xlsx}  (valeurs seules, prêt ERP)")
    if args.output:
        if safe(f"Import CSV {args.output}", lambda: export(imp, args.output, cfg)):
            print(f"Fichier d'import CSV écrit  : {args.output}")

if __name__ == "__main__":
    main()
