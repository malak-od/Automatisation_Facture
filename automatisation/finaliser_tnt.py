#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR TNT — produit "AAAA_MM_Facture TNT.xlsx" A L'IDENTIQUE du fichier
fait a la main (8 feuilles : Catégories, Recherche tracking, Nb colis et
poids, Facture TNT, TCD, Bilan clients, Import csv, Demande d'avoir), en
partant du fichier existant comme MODELE.

Classeur reel (facture FedEx Express FR SAS "services TNT") :
  Facture TNT : 1 ligne = 1 EVENEMENT de facturation brut (PAS 1 colis) --
    colonnes D->AD = brut recu TEL QUEL (en-tete stable, meme structure que
    "La Ruche <compte> - <mois>.xlsx"), colonnes A->C CALCULEES par formule
    (A="Probleme" jamais rempli, B="Categorie"=XLOOKUP(code complement,
    Categories!A:A,B:B,"inconnu") ou "Frêt" si vide, C="Tracking"=
    RIGHT(LOOKUP(numero envoi,'Recherche tracking'!A:A,B:B),16)).
  Recherche tracking : VRAI TCD (RowFields=numero envoi+id colis, source=
    'Facture TNT'!J:L) -- sert de table de correspondance envoi->tracking(s).
  Nb colis et poids : 2 VRAIS TCD sur 'Facture TNT' -- #1 (A3:B104) filtre
    PageField "code complement"="TMC", Somme de montant HT par Tracking (sert
    a compter les colis via Import csv!K = LOOKUP(...)+1) ; #2 (E1:F516) Max
    de poids colis par Tracking, sans filtre.
  TCD : VRAI TCD principal (F1:K517, RowField=Tracking, ColField=Categorie,
    DataField=Somme de montant HT -- colonnes G/H/I/J natives = Colis
    volumineux/Frêt/Gazole/Zones eloignees) + colonnes MANUELLES juxtaposees
    A (Total avec GO)/B (Total hors GO)/C (ID client, valeurs figees saisies
    a la main)/D (Tracking, LOOKUP)/E (Poids, XLOOKUP vers 'Nb colis et
    poids'). PIEGE CONNU (comme Chronopost/Mondial Relay) : "Tableau croisé
    dynamique4" (Y1:Z57, Poids->Moyenne de Frêt) est un TCD IMBRIQUE sur la
    feuille TCD elle-meme (source TCD!Q1:V1048576, PAS 'Facture TNT') --
    JAMAIS redirige (meme limite que Mondial Relay 'Bilan clients'/Chronopost
    'TCD' avant reconstruction complete).
  Bilan clients : 1 TCD (RowField=ID client, source='TCD'!A1:C1048576).
  Import csv : formules FIXES referencant TCD par POSITION (ligne n import
    <-> ligne n+1 TCD, decalage constant du a l'entete 2 lignes du TCD natif).

Taxe Gasoil (SURCH_CARB) : montant GLOBAL par compte expediteur, PAS ventile
  par tracking dans le brut (contrairement au pool gazole Chronopost) --
  jamais repercute au client (colonne Gazole/V de "Import csv" toujours vide).
  PDF affiche 2 taux ("officiel" et "reel") -- SEUL le taux officiel fait foi
  (consigne utilisateur). Reconciliation : somme des postes hors Gazole
  (colonnes O->U) + surcharge carburant brute = TOTAL GENERAL PDF.

Tracking (id colis) : TOUJOURS du texte (16 chiffres, prefixe "'" dans le
  brut Excel) -- JAMAIS converti en nombre a aucune etape (sinon le tracking
  ne correspond plus a rien dans l'ERP, consigne utilisateur explicite).

Necessite : Windows + Excel + pywin32 + pdfplumber/pypdf (reconciliation PDF,
optionnelle).
Usage :
  python finaliser_tnt.py "<modele.xlsx>" "<sortie.xlsx>" "<brut.xlsx>" [--pdf <pdf1> [<pdf2>...]]
"""
import sys, os, shutil, re, csv


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def compare_key(h):
    s = normalize_header(h).lower()
    s = s.translate(str.maketrans("éèêëàâäùûüôöîïç", "eeeeaaauuuooiic"))
    s = re.sub(r"[.,]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def col_index(header, name):
    target = compare_key(name)
    for i, h in enumerate(header):
        if compare_key(h) == target:
            return i
    return None


def col_letter(idx0):
    import openpyxl
    return openpyxl.utils.get_column_letter(idx0 + 1)


def is_xlsx(path):
    import zipfile
    try:
        return zipfile.is_zipfile(path)
    except Exception:
        return False


def is_xlsb_content(path):
    """.xlsb ET .xlsx sont TOUS LES DEUX des conteneurs ZIP valides (is_xlsx() seul ne les
    distingue pas) -- BUG TROUVE 2026-08-19 : detecter le xlsb par EXTENSION de fichier
    (`path.lower().endswith(".xlsb")`) fonctionne en CLI (chemin d'origine conserve) mais
    ECHOUE SILENCIEUSEMENT quand le fichier arrive via le serveur web (multer renomme les
    uploads avec un hash SANS EXTENSION, ex. 'a1b2c3d4e5f6') -- le fichier etait alors traite
    comme un .xlsx normal, qu'openpyxl refusait avec InvalidFileException, faisant echouer
    tout le finaliseur (repli casse cote serveur, "Cannot read properties of undefined
    (reading 'map')"). Fix : detection par CONTENU, pas extension -- un .xlsb a
    'xl/workbook.bin' dans son ZIP, un .xlsx a 'xl/workbook.xml' (verifie sur le fichier reel
    'Détail facture Juillet 2026_Acc <compte>.xlsb')."""
    import zipfile
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            return "xl/workbook.bin" in names
    except Exception:
        return False


def convert_xlsb_to_xlsx(path):
    """.xlsb (Excel Binary Workbook) : openpyxl NE LE SUPPORTE PAS DU TOUT (leve
    InvalidFileException, contrairement a .xls/.csv ou un simple renommage suffit -- .xlsb
    est un format BINAIRE, pas XML, meme si son conteneur est aussi un ZIP). Constate sur le
    fichier reel de juillet 2026 ('Détail facture Juillet 2026_Acc <compte>.xlsb') : le
    donneur d'ordre livre parfois ce format au lieu du .xlsx habituel. Convertit via une
    instance Excel COM DEDIEE et EPHEMERE (independante de celle du reste du script, qui
    n'est ouverte que plus tard sur le fichier de SORTIE) -- SaveAs FileFormat=51
    (xlOpenXMLWorkbook) vers un fichier temporaire, nettoye apres lecture."""
    import tempfile
    import win32com.client as win32
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(os.path.abspath(path), UpdateLinks=0, ReadOnly=True)
        wb.SaveAs(tmp.name, FileFormat=51)  # xlOpenXMLWorkbook
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()
    return tmp.name


def read_xlsx_rows(path):
    import openpyxl, tempfile
    load_path = path
    tmp_path = None
    if path.lower().endswith(".xlsb") or is_xlsb_content(path):
        load_path = tmp_path = convert_xlsb_to_xlsx(path)
    elif not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copyfile(path, tmp.name)
        load_path = tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(load_path, data_only=True)
        sheet = "DET_FAC" if "DET_FAC" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def read_csv_rows(path):
    with open(path, encoding="latin-1", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def read_input(paths):
    header = None
    all_rows = []
    for p in paths:
        result = read_xlsx_rows(p) if is_xlsx(p) else read_csv_rows(p)
        if result is None:
            raise RuntimeError(f"Fichier vide ou illisible : {p}")
        h, rows = result
        if header is None:
            header = h
        ncol = len(header)
        for r in rows:
            r = (list(r) + [None] * ncol)[:ncol]
            all_rows.append(r)
    return header, all_rows


def coerce(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.fullmatch(r"-?0\d+", s):
        return v
    if re.fullmatch(r"-?\d+(,\d+)?", s):
        return float(s.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


def extract_pdf_infos(pdf_paths):
    """PDF TNT (facture FedEx "services TNT") : blocs "<compte> : Surcharge Carburant
    (taux officiel X,XX %) ... Y,YY% taux reel  Montant" (1 par compte expediteur), et
    "TOTAL GENERAL<colis><envois><poids><montant> EUR" (Poids et Montant colles, 2 nombres
    a virgule consecutifs -- le Montant est le 2e). SEUL le taux officiel fait foi (consigne
    utilisateur) -- le taux reel n'est jamais utilise pour le calcul."""
    try:
        import pypdf
    except ImportError:
        return []
    results = []
    for p in pdf_paths:
        try:
            text = "\n".join((page.extract_text() or "") for page in pypdf.PdfReader(p).pages)
        except Exception:
            continue

        def num_milliers(s):
            return float(s.replace(".", "").replace(",", "."))

        surcharges = []
        for m in re.finditer(
            r"(\d{8})\s*:\s*Surcharge\s*Carburant\s*\(taux\s*officiel\s*([\d,]+)\s*%\)\s*"
            r"([\d,]+)\s*%\s*taux\s*r[ée]el\s*([\d,]+)", text,
        ):
            surcharges.append({
                "compte": m.group(1),
                "taux_officiel": float(m.group(2).replace(",", ".")) / 100,
                "montant": num_milliers(m.group(4)),
            })
        m_total = re.search(
            r"TOTAL\s*GENERAL[\s\S]{0,40}?\d{1,3}(?:\.\d{3})*,\d{2}(\d{1,3}(?:\.\d{3})*,\d{2})\s*EUR",
            text,
        )
        m_numero = re.search(r"FACTURE\s*N[°\s]*(\d[\d\s]*\d)", text)
        results.append({
            "file": os.path.basename(p),
            "numero_facture": re.sub(r"\s", "", m_numero.group(1)) if m_numero else None,
            "total_general": num_milliers(m_total.group(1)) if m_total else None,
            "surcharges": surcharges,
        })
    return results


def fill_reconciliation(wb, pdfs):
    """Ecrit les infos de reconciliation dans les cellules libres de 'Demande d'avoir'
    (feuille de notes, non structuree en TCD) -- pas de zone de collage native dans le
    modele pour la reconciliation PDF TNT (contrairement a Chronopost/'Contrôle pdf' ou
    Mondial Relay/'Controle xls pdf', TNT n'a pas d'onglet dedie) : imprime seulement en
    console, laisse la verification a la charge du pole transport via les 'infos' deja
    remontees par le carrier Node (somme postes + surcharge = TOTAL GENERAL)."""
    for p in pdfs:
        for s in p["surcharges"]:
            print(f"PDF {p['file']} : compte {s['compte']} — Surcharge Carburant taux officiel "
                  f"{s['taux_officiel']*100:.2f}% ({s['montant']:.2f} EUR).")
        if p["total_general"] is not None:
            print(f"PDF {p['file']}{' (' + p['numero_facture'] + ')' if p['numero_facture'] else ''} : "
                  f"TOTAL GENERAL = {p['total_general']:.2f} EUR (comparer a l'info 'Total attendu' du carrier Node).")


def parse_args(argv):
    modele, sortie = argv[1], argv[2]
    inputs, pdfs, cur = [], [], "in"
    for a in argv[3:]:
        if a == "--pdf":
            cur = "pdf"
        elif cur == "pdf":
            pdfs.append(a)
        else:
            inputs.append(a)
    return modele, sortie, inputs, pdfs


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def main():
    modele, sortie, inputs, pdf_paths = parse_args(sys.argv)
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    header, rows = read_input(inputs)
    n = len(rows)
    ncol = len(header)
    data = [[coerce(v) for v in r] for r in rows]
    print(f"Entrée : {n} lignes x {ncol} colonnes")

    pdfs = extract_pdf_infos(pdf_paths) if pdf_paths else []

    # Date validite = mois MAJORITAIRE (meme piege deja rencontre sur Chronopost : le brut
    # peut contenir quelques lignes residuelles d'un autre mois).
    i_mois = col_index(header, "Mois de Facturation")
    date_validite_serial = None
    if i_mois is not None:
        import datetime as _dt
        from collections import Counter
        compte = Counter()
        for r in rows:
            v = str(r[i_mois] or "").strip()
            if re.match(r"^\d{6}$", v):
                compte[v] += 1
        if compte:
            mois_maj = compte.most_common(1)[0][0]
            annee, mois = int(mois_maj[:4]), int(mois_maj[4:6])
            EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
            date_validite_serial = (_dt.datetime(annee, mois, 1) - EXCEL_EPOCH).days

    import win32com.client as win32
    xlUp = -4162
    xlDatabase = 1
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) "Facture TNT" : purge + collage donnees brutes a partir de la colonne D
        #    (A-C = calculees : Probleme/Categorie/Tracking, jamais dans le brut recu).
        ws = wb.Sheets("Facture TNT")
        if ws.AutoFilterMode:  # meme piege que Chronopost : AutoFilter actif -> ecritures silencieusement ignorees
            ws.AutoFilterMode = False
        FIRST_RAW_COL = 4  # colonne D
        LAST_RAW_COL = FIRST_RAW_COL - 1 + ncol
        LAST_CALC_COL = 3  # A->C (Probleme/Categorie/Tracking)
        oldLast = ws.Cells(ws.Rows.Count, FIRST_RAW_COL).End(xlUp).Row
        newLast = 1 + n
        maxLast = max(oldLast, newLast, 2)

        # Marge fixe et large (pas relative a maxLast seul) -- meme fix que le bug Chronopost
        # 2026-08-14 (une marge relative peut laisser un residu si le nouveau mois est plus
        # grand que l'ancien modele).
        purgeUntil = maxLast + 200
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(purgeUntil, max(LAST_RAW_COL, LAST_CALC_COL))).ClearContents())

        # "id colis"/"numero envoi" (donc "Tracking" derive) sont du TEXTE dans le brut --
        # force NumberFormat="@" AVANT ecriture pour eviter l'auto-conversion Excel qui
        # casserait le TCD "Recherche tracking" (meme piege que "No Facture"/"Numero LT" sur
        # Chronopost) ET violerait la consigne explicite "ne jamais convertir en nombre".
        i_id_colis = col_index(header, "id colis")
        i_numero_envoi = col_index(header, "numero envoi")
        col_id_colis = col_letter(FIRST_RAW_COL - 1 + i_id_colis) if i_id_colis is not None else None
        col_numero_envoi = col_letter(FIRST_RAW_COL - 1 + i_numero_envoi) if i_numero_envoi is not None else None
        for letter in (col_id_colis, col_numero_envoi):
            if letter:
                retry(lambda l=letter: ws.Range(f"{l}2:{l}{maxLast}").__setattr__("NumberFormat", "@"))

        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, LAST_RAW_COL)).__setattr__("Value", data))
        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, max(LAST_RAW_COL, LAST_CALC_COL))).ClearContents())

        # Liste de reference des 4 postes ERP (colonne B, JUSTE APRES la derniere ligne de
        # donnees, fond bleu clair) -- demande utilisateur explicite (consigne texte + capture
        # ecran du fichier fait-main) : "dans la colonne B rajouter ca a la fin de la colonne :
        # Colis Volumineux / Frêt / Gazole / Zones eloignees". "Assurance" testee puis RETIREE
        # (2026-08-19) : le fichier fait-main REEL (capture ecran fournie, colonnes G->K =
        # Colis volumineux/Frêt/Gazole/Zones eloignees/(vide), PAS Assurance) confirme que
        # cette 5e ligne etait une extrapolation non voulue -- meme si "Assurance" existe dans
        # la table Categories (code AS_AD_VAL), le process reel du pole transport ne l'inclut
        # PAS dans cette liste de reference. Fidelite au fichier fait-main > exhaustivite
        # theorique de la table Categories.
        # "Zones éloignées " (AVEC espace de fin) -- orthographe EXACTE de la table 'Catégories'
        # du classeur Excel reel (confirme via COM : 2 occurrences, 'ZEL' et le libelle par
        # defaut, toutes deux avec l'espace) -- PAS celle de config.json (carrier Node.js, sans
        # espace, sans impact sur le CSV livre qui ne depend pas du TCD Excel). Un espace en
        # trop ici creerait un DOUBLON dans le ColField natif du TCD (2 colonnes "Zones
        # eloignees" distinctes, l'une avec et l'une sans espace) -- bug constate en test.
        POSTE_KEYS_LISTE = ["Colis volumineux", "Frêt", "Gazole", "Zones éloignées "]
        refStart = newLast + 1
        for i, poste in enumerate(POSTE_KEYS_LISTE):
            ws.Cells(refStart + i, 2).Value = poste
        refRange = ws.Range(ws.Cells(refStart, 2), ws.Cells(refStart + len(POSTE_KEYS_LISTE) - 1, 2))
        # Fond bleu clair : ThemeColor/TintAndShade exacts LUS via COM sur la cellule de
        # reference DEJA presente dans le modele de juin (B2076, "Colis volumineux") --
        # ThemeColor=5, TintAndShade=0.7999816888943144 -- PAS repris d'une cellule de donnees
        # normale (celles-ci ont un theme DIFFERENT, cf. derniere ligne de vraies donnees
        # B2075, theme distinct du bloc de reference).
        refRange.Interior.ThemeColor = 5
        refRange.Interior.TintAndShade = 0.7999816888943144

        # Colonnes A-C : formules du modele, resolues par nom de colonne source.
        i_code_complement = col_index(header, "code complement")
        col_code_complement = col_letter(FIRST_RAW_COL - 1 + i_code_complement) if i_code_complement is not None else None
        col_numero_envoi_ref = col_numero_envoi  # meme colonne, reference dans C

        if col_code_complement:
            formula_b = f'=IF({{col}}{{{{row}}}}="","Frêt",_xlfn.XLOOKUP({{col}}{{{{row}}}},Catégories!A:A,Catégories!B:B,"inconnu"))'.format(col=col_code_complement)
            retry(lambda t=formula_b: ws.Range(ws.Cells(2, 2), ws.Cells(newLast, 2)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))
        else:
            print("AVERTISSEMENT: 'code complement' introuvable -> colonne B (Catégorie) non reconstruite.")

        if col_numero_envoi_ref:
            formula_c = f"=RIGHT(LOOKUP({{col}}{{{{row}}}},'Recherche tracking'!A:A,'Recherche tracking'!B:B),16)".format(col=col_numero_envoi_ref)
            retry(lambda t=formula_c: ws.Range(ws.Cells(2, 3), ws.Cells(newLast, 3)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))
        else:
            print("AVERTISSEMENT: 'numero envoi' introuvable -> colonne C (Tracking) non reconstruite.")

        ws.Range(ws.Cells(1, 1), ws.Cells(newLast, max(LAST_RAW_COL, LAST_CALC_COL))).AutoFilter()

        # 2) TCD "Recherche tracking"/"Nb colis et poids" (les 2 TCD)/"TCD" (colonnes G-K
        #    natives)/"Bilan clients" : PivotCache redirige vers la vraie plage de donnees.
        #    "Recherche tracking" et "Nb colis et poids" (2 TCD) pointent sur 'Facture TNT' --
        #    redirection standard (meme piege que DPD/Geodis/Mondial Relay/Chronopost : cache
        #    du modele fige sur une plage figee/etroite).
        #
        #    "TCD" (ColField="Catégorie") a besoin d'une plage ETENDUE jusqu'a la liste de
        #    reference des 5 postes (refStart..refStart+4, colonne B) -- meme bug/fix DEJA
        #    trouve sur Chronopost ("bug 8", 2026-08-14) : le PivotCache ne peut afficher QUE
        #    les items de colField qu'il a REELLEMENT rencontres en scannant sa SourceData --
        #    sans cette extension, un poste sans aucune vraie ligne ce mois-ci (souvent
        #    "Assurance") disparaitrait purement et simplement du TCD/ColField, au lieu
        #    d'apparaitre a 0/vide comme attendu. Chaque ligne de la liste n'a NI Tracking NI
        #    aucune autre colonne renseignee -> reste invisible pour toute la logique metier
        #    en aval (RowFields Tracking/numero envoi des 2 autres TCD n'y trouvent rien).
        newRangeFacture = ws.Range(ws.Cells(1, 1), ws.Cells(newLast, max(LAST_RAW_COL, LAST_CALC_COL)))
        newRangeFactureAvecRef = ws.Range(ws.Cells(1, 1), ws.Cells(refStart + len(POSTE_KEYS_LISTE) - 1, max(LAST_RAW_COL, LAST_CALC_COL)))
        pivots_non_redires = []
        for sheet_name in ("Recherche tracking", "Nb colis et poids", "TCD"):
            wsPivot = wb.Sheets(sheet_name)
            rangeCible = newRangeFactureAvecRef if sheet_name == "TCD" else newRangeFacture
            for i in range(1, wsPivot.PivotTables().Count + 1):
                pt = wsPivot.PivotTables(i)
                src = str(pt.PivotCache().SourceData)
                if "Facture TNT" not in src:
                    # "Tableau croisé dynamique4" (TCD!Y1:Z57) : TCD IMBRIQUE sur TCD elle-meme
                    # (source TCD!Q1:V1048576) -- JAMAIS redirige, meme limite connue que
                    # Mondial Relay/'Bilan clients' et l'ancien Chronopost/'TCD'.
                    print(f"TCD '{pt.Name}' ({sheet_name}) : source '{src}' n'est pas 'Facture TNT' "
                          f"(TCD imbrique) -> config preservee, reste sur ses valeurs figees du modele.")
                    pivots_non_redires.append(pt)
                    continue
                newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=rangeCible)
                pt.ChangePivotCache(newCache)
                print(f"TCD '{pt.Name}' ({sheet_name}) : PivotCache redirigé vers {rangeCible.Address}.")
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        xl.Calculate()

        # "Bilan clients" (source='TCD'!A1:C1048576, colonnes MANUELLES) : redirige aussi vers
        # la hauteur reelle du TCD (le TCD natif s'etend deja via RefreshAll ci-dessus, mais les
        # colonnes A/B/C manuelles n'existent que jusqu'a lastTcd -- redirection sur toute la
        # hauteur POSSIBLE, pas juste celle deja peuplee, pour rester correct si le TCD change
        # de taille a la prochaine generation).
        wsTcdTmp = wb.Sheets("TCD")
        lastTcd = wsTcdTmp.Cells(wsTcdTmp.Rows.Count, 6).End(xlUp).Row  # colonne F = Tracking natif
        for i in range(1, wb.Sheets("Bilan clients").PivotTables().Count + 1):
            pt = wb.Sheets("Bilan clients").PivotTables(i)
            newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=wsTcdTmp.Range(wsTcdTmp.Cells(1, 1), wsTcdTmp.Cells(max(lastTcd, 2), 3)))
            pt.ChangePivotCache(newCache)
        wb.RefreshAll()
        xl.Calculate()

        # 3) "TCD" colonnes manuelles A/B/D/E (etendues jusqu'a la vraie derniere ligne du TCD
        #    natif, colonne F) -- MEME BUG que Chronopost/"TCD poids" deja documente : ces
        #    formules manuelles restent FIGEES a la taille du modele si non re-etirees apres
        #    RefreshAll. Colonne C (ID client) reste NON etiree : valeurs FIGEES saisies a la
        #    main (decision utilisateur, aucune source automatisable identifiee -- meme regle
        #    transversale que BLS/Chronopost/Colissimo, cf. memoire id_client_regle_classeurs_
        #    facture.md).
        #
        #    BUG TROUVE 2026-08-18 : le ColField natif du TCD ("Catégorie") n'affiche QUE les
        #    postes reellement presents dans le mois traite -- son ORDRE/POSITION (colonnes
        #    G/H/I/J...) depend des donnees, PAS fixe comme dans le modele juin (qui avait les
        #    4 postes 'Colis volumineux'/'Frêt'/'Gazole'/'Zones éloignées' dans cet ordre
        #    precis). Un mois SANS 'Colis volumineux' (ex. juin 2026 reel, 0 ligne CNM) decale
        #    tout : G devient 'Frêt' au lieu de 'Colis volumineux' -- coder G/H/I/J en dur
        #    cassait silencieusement le mapping poste->colonne (Frêt lu comme
        #    ColisVolumineux, etc.). Fix : resout les 4 lettres par NOM d'en-tete (ligne 2 du
        #    TCD natif), APRES le RefreshAll ci-dessus.
        wsTcd = wb.Sheets("TCD")
        # BUG TROUVE 2026-08-19 (signale utilisateur, "lignes a la fin du TCD sans aucune
        # information") : lastTcd (colonne F, RowField natif) donne bien la vraie derniere
        # ligne de CE mois -- mais les colonnes MANUELLES A/B/D/E gardaient les VALEURS
        # FIGEES d'un mois PRECEDENT plus grand (ex. juin 2026, 531 lignes, vs juillet 2026,
        # 144 lignes) au-dela de lastTcd, jamais purgees avant reecriture -- Excel ne les
        # efface pas tout seul, ce sont des VALEURS residuelles (pas des formules qui se
        # recalculeraient a vide), visibles comme des lignes "fantomes" (Tracking=#N/A,
        # tout a 0) apres la vraie derniere ligne. Purge d'abord TOUTE la plage A:E jusqu'a
        # l'ancienne derniere ligne UTILISEE de la feuille (oldLastTcdUsed, calculee AVANT
        # toute reecriture) -- meme principe que la purge deja faite sur 'Facture TNT'/
        # 'Import csv', simplement absente ici jusqu'a present.
        oldLastTcdUsed = wsTcd.UsedRange.Rows.Count
        lastTcd = wsTcd.Cells(wsTcd.Rows.Count, 6).End(xlUp).Row  # colonne F native
        if oldLastTcdUsed > lastTcd:
            retry(lambda: wsTcd.Range(wsTcd.Cells(lastTcd + 1, 1), wsTcd.Cells(oldLastTcdUsed, 5)).ClearContents())
            print(f"'TCD' : lignes fantômes {lastTcd + 1}..{oldLastTcdUsed} (colonnes A-E, résidus d'un mois précédent) purgées.")
        headerRowTcd = [normalize_header(wsTcd.Cells(2, c).Value) for c in range(7, 20)]

        def find_tcd_col(nom):
            for i, h in enumerate(headerRowTcd):
                if compare_key(h) == compare_key(nom):
                    return col_letter(6 + i)  # colonne 7 = G
            return None

        col_cv = find_tcd_col("Colis volumineux")
        col_fret = find_tcd_col("Frêt")
        col_gazole = find_tcd_col("Gazole")
        col_ze = find_tcd_col("Zones éloignées")
        missing = [n for n, c in (("Colis volumineux", col_cv), ("Frêt", col_fret), ("Gazole", col_gazole), ("Zones éloignées", col_ze)) if c is None]
        if missing:
            print(f"INFO: poste(s) ERP absent(s) du TCD ce mois-ci (aucune ligne) : {', '.join(missing)} -- colonne(s) correspondante(s) non trouvée(s), formules A/B les ignorent.")

        if lastTcd >= 3:
            # A (Total avec GO) : SUM(hors-Gazole) + SUM(tout y compris Gazole) -- formule
            # modele exacte, colonnes resolues dynamiquement (fallback sur une plage vide
            # "Z:Z" si un poste est absent ce mois-ci, pour ne jamais casser la formule).
            cols_hors_gazole = [c for c in (col_cv, col_fret, col_ze) if c]
            cols_tout = [c for c in (col_cv, col_fret, col_gazole, col_ze) if c]
            sum_hors_gazole = "+".join(f"SUM({c}:{c})" for c in cols_hors_gazole) or "0"
            sum_tout_row = "+".join(f"{c}{{row}}" for c in cols_tout) or "0"
            sum_tout_col = "+".join(f"SUM({c}:{c})" for c in cols_tout) or "0"
            formulas_tcd = {
                1: f"=ROUND(B{{row}}/({sum_hors_gazole})*({sum_tout_col}),2)",  # A Total avec GO
                2: f"=({sum_tout_row})",  # B Total hors GO
                4: "=RIGHT(_xlfn.XLOOKUP(F{row},'Recherche tracking'!A:A,'Recherche tracking'!B:B),16)",  # D Tracking
                5: "=_xlfn.XLOOKUP(F{row},'Nb colis et poids'!E:E,'Nb colis et poids'!F:F)",  # E Poids
            }
            for col_idx, tmpl in formulas_tcd.items():
                retry(lambda c=col_idx, t=tmpl: wsTcd.Range(wsTcd.Cells(3, c), wsTcd.Cells(lastTcd, c))
                      .__setattr__("Formula", [[t.format(row=r)] for r in range(3, lastTcd + 1)]))
            print(f"'TCD' : colonnes A/B/D/E étirées jusqu'à la ligne {lastTcd} (Colis volumineux={col_cv}, Frêt={col_fret}, Gazole={col_gazole}, Zones éloignées={col_ze}).")

        # 4) "Import csv" : formules FIXES du modele, referencant TCD par POSITION (ligne n
        #    import <-> ligne n+1 TCD). Nombre de lignes = nombre de trackings UNIQUES du TCD
        #    (lastTcd - 2), PAS le nombre de lignes brutes recues.
        wsImp = wb.Sheets("Import csv")
        if wsImp.AutoFilterMode:
            wsImp.AutoFilterMode = False
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 6).End(xlUp).Row
        LAST_COL_IMPORT = 21

        newLastImp = max(lastTcd - 1, 2)
        # Q/R/T referencent les MEMES colonnes TCD resolues dynamiquement plus haut (col_ze/
        # col_cv/col_fret) -- jamais G/H/I/J en dur (meme bug que les formules A/B du TCD,
        # cf. docstring 2026-08-18). Poste absent ce mois-ci (colonne None) -> formule vide.
        formulas_import = {
            6: "=TCD!F{tcdrow}",  # F N° Tracking
            12: "=TCD!E{tcdrow}",  # L Poids
            11: "=IF(COUNTIF('Nb colis et poids'!A:A,F{row})=0,1,LOOKUP(F{row},'Nb colis et poids'!A:A,'Nb colis et poids'!B:B+1))",  # K Nbr Colis
        }
        # Si un poste est ABSENT ce mois-ci (col_x=None), la colonne Import csv correspondante
        # est explicitement VIDEE (pas juste "non ecrite") -- sinon l'ANCIENNE formule du
        # modele (referencant une lettre TCD qui a change de sens depuis, cf. bug ci-dessus)
        # reste en place et pointe vers un AUTRE poste par accident (constate : R gardait
        # '=TCD!G3' du modele meme quand col_cv=None, G etant devenu 'Frêt' -> ColisVolumineux
        # affichait en fait du Frêt).
        if col_ze:
            formulas_import[17] = f'=IF(TCD!{col_ze}{{tcdrow}}="","",TCD!{col_ze}{{tcdrow}})'  # Q Zones éloignées
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 17), wsImp.Cells(newLastImp, 17)).ClearContents())
        if col_cv:
            formulas_import[18] = f'=IF(TCD!{col_cv}{{tcdrow}}="","",TCD!{col_cv}{{tcdrow}})'  # R Colis volumineux
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 18), wsImp.Cells(newLastImp, 18)).ClearContents())
        if col_fret:
            formulas_import[20] = f"=TCD!{col_fret}{{tcdrow}}"  # T Frêt
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 20), wsImp.Cells(newLastImp, 20)).ClearContents())
        # Colonnes fixes (ligne 2 = valeur, ligne 3+ = formule "=<col><n-1>", comme le modele).
        FIXED_COLS = {1: "TNT", 8: "E", 9: "FR", 10: "France", 13: "ST", 14: 0.2}
        for col_idx, val in FIXED_COLS.items():
            wsImp.Cells(2, col_idx).Value = val
            if newLastImp > 2:
                letter = col_letter(col_idx - 1)
                retry(lambda c=col_idx, l=letter: wsImp.Range(wsImp.Cells(3, c), wsImp.Cells(newLastImp, c))
                      .__setattr__("Formula", [[f"={l}{r - 1}"] for r in range(3, newLastImp + 1)]))
        if date_validite_serial is not None:
            wsImp.Cells(2, 2).Value = date_validite_serial
        else:
            print("AVERTISSEMENT: 'Mois de Facturation' introuvable -> 'Import csv'!B2 non mise à jour, reste celle du modèle.")
        if newLastImp > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(3, 2), wsImp.Cells(newLastImp, 2))
                  .__setattr__("Formula", [["=B{prev}".format(prev=r - 1)] for r in range(3, newLastImp + 1)]))
        for col_idx, tmpl in formulas_import.items():
            retry(lambda c=col_idx, t=tmpl: wsImp.Range(wsImp.Cells(2, c), wsImp.Cells(newLastImp, c))
                  .__setattr__("Formula", [[t.format(row=r, tcdrow=r + 1)] for r in range(2, newLastImp + 1)]))
        if newLastImp < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLastImp + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT)).ClearContents())
        print(f"'Import csv' : formules du modèle reconstruites jusqu'à la ligne {newLastImp} (TCD ligne {newLastImp + 1}).")

        wsImp.Range(wsImp.Cells(1, 1), wsImp.Cells(max(newLastImp, 2), LAST_COL_IMPORT)).AutoFilter()
        xl.Calculate()

        # 5) Réconciliation PDF (info console uniquement, pas d'onglet dédié dans ce modele).
        if pdfs:
            try:
                fill_reconciliation(wb, pdfs)
            except Exception as e:
                print("Réconciliation TNT ignorée :", e)

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)
        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
