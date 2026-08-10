#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR GEODIS — produit "AAAA_MM_Facture Geodis.xlsx" A L'IDENTIQUE du
fichier fait a la main (memes feuilles : Factures Geodis, Bilan Factures,
Import CSV, Bilan Client ; memes formules, memes TCD/pivots, meme mise en
forme), en partant du fichier existant comme MODELE et en n'y remplacant que
les DONNEES BRUTES (colonne N "N pi�ce" et suivantes dans "Factures Geodis").

Le fichier d'entree peut etre le CSV/XLSX brut Geodis (entete en ligne 1) OU
un rapport client type "Facturation_client.xlsx" (memes colonnes mais entete
decalee de quelques lignes de titre) -> l'entete est detectee dynamiquement
(recherche de "N recepisse"), comme cote Node (src/carriers/geodis/index.js).

ATTENTION (cf. FACTURATION EXCEL.pdf p.3) : les colonnes brutes Geodis se
decalent d'un mois a l'autre (colonne ajoutee/retiree/renommee legerement).
Les 8 postes ERP (colonnes F->M de "Factures Geodis") ne sont donc PAS colles
avec des formules a positions fixes (=CS2+CT2+CU2) : on colle l'entete recu
TEL QUEL (avec le style bleu du modele) a partir de la colonne N, PUIS on
reconstruit les formules F->M en cherchant CHAQUE colonne par NOM dans cet
entete (postes_from_columns du config.json Node, source unique de verite,
partagee avec facturation-app/src/carriers/geodis/config.json). Une colonne
de poste introuvable ce mois-ci est ignoree dans la somme (poste calcule sur
les colonnes presentes) + avertissement -- jamais bloquant, coherent avec le
comportement Node (idx() -> -1, colonne juste ignoree).

Necessite : Windows + Excel + pywin32 + openpyxl.
Usage :
  python finaliser_geodis.py "<modele.xlsx>" "<sortie.xlsx>" "<entree1>" [<entree2>...] [--pdf-taxable N] [--frais-gestion N]
"""
import sys, os, shutil, re, csv, json

FIRST_RAW_COL = 14  # colonne N : debut des donnees brutes Geodis dans "Factures Geodis"
LAST_COL_IMPORT_CSV = 23        # Import CSV : colonnes A->W (Gazole/NbColis compris, meme si vides)
HEADER_FILL_ARGB = "FF3200E6"   # bleu du modele (colonnes N+, ligne 1)
HEADER_FONT_ARGB = "FFFFFFFF"   # texte blanc

# Colonnes de "Factures Geodis" -> cle postes_from_columns (config.json Node).
# L/A/B/C/D/E n'ont pas de somme de colonnes nommees (formules fixes du modele,
# non touchees : Recepisse=RIGHT(V,8), Total hors GO=SUM(F:L), Total+GO=D+M).
POSTE_COLS = {
    "F": "DroitsTaxes", "G": "Assurance", "H": "ZonesEloignees", "I": "ColisVolumineux",
    "J": "Adresses", "K": "Fret", "L": "PlusValueB2C", "M": "Gazole",
}


def load_postes_from_columns():
    """Lit postes_from_columns depuis le config.json Node (source unique de
    verite, deja utilise par src/carriers/geodis/index.js -> pas de liste
    dupliquee/desynchronisable entre les deux implementations)."""
    cfg_path = os.path.join(os.path.dirname(__file__), "..", "facturation-app",
                             "src", "carriers", "geodis", "config.json")
    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg["postes_from_columns"]


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def compare_key(h):
    """Cle de comparaison ROBUSTE aux variantes cosmetiques vues entre exports
    Geodis (ex. 'RV tel,' vs 'RV tel.', accents) : on ne veut PAS ignorer un
    vrai renommage/decalage de colonne, juste la ponctuation/accentuation."""
    s = normalize_header(h).lower()
    s = s.translate(str.maketrans("éèêëàâäùûüôöîïç", "eeeeaaauuuooiic"))
    s = re.sub(r"[.,]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def find_header_row(rows):
    """Cherche, dans les 15 premieres lignes, celle qui contient une colonne
    ressemblant a 'N recepisse' (insensible aux accents/casse)."""
    for i, row in enumerate(rows[:15]):
        for v in row:
            s = normalize_header(v).lower()
            if "recepiss" in s or "récépiss" in s:
                return i
    return None


def read_csv_rows(path):
    with open(path, encoding="latin-1", newline="") as f:
        return list(csv.reader(f, delimiter=";"))


def is_xlsx(path):
    """Detection par contenu (magic bytes ZIP), PAS par extension : les fichiers
    uploades (multer) arrivent sans extension du tout."""
    import zipfile
    try:
        return zipfile.is_zipfile(path)
    except Exception:
        return False


def read_xlsx_rows(path):
    """Cherche, sur TOUTES les feuilles, celle qui contient l'entete Geodis
    (ex. Facturation_client.xlsx : feuille 'Detail', entete en ligne 4).
    openpyxl refuse d'ouvrir un fichier sans extension .xlsx (verifiee sur le
    NOM, pas le contenu) -> copie temporaire renommee si besoin."""
    import openpyxl, tempfile
    load_path = path
    tmp_path = None
    if not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copyfile(path, tmp.name)
        load_path = tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(load_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
            hidx = find_header_row(rows)
            if hidx is not None:
                wb.close()
                return rows[hidx:]
        wb.close()
        return None
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def read_input(paths):
    """Concatene toutes les sources -> (header, data_rows). Format detecte par
    CONTENU (is_xlsx), jamais par extension (absente sur les fichiers uploades)."""
    header = None
    all_rows = []
    for p in paths:
        as_xlsx = is_xlsx(p)
        rows = read_xlsx_rows(p) if as_xlsx else read_csv_rows(p)
        if rows is None:
            raise RuntimeError(f"Format non reconnu (aucune feuille avec 'N° récépissé') : {p}")
        hidx = 0  # read_xlsx_rows renvoie deja a partir de l'entete
        if not as_xlsx:
            hidx = find_header_row(rows)
            if hidx is None:
                raise RuntimeError(f"Format non reconnu (aucune ligne d'entete avec 'N° récépissé') : {p}")
            rows = rows[hidx:]
        if header is None:
            header = [normalize_header(h) for h in rows[0]]
        data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
        all_rows.extend(data)
    return header, all_rows


def coerce(v):
    """Nombre si la valeur est un nombre francais valide (virgule ou point),
    sinon la valeur telle quelle (texte/date/None).
    Exception : un entier avec un zero de tete (ex. '000093630935') est une
    reference/code, pas une quantite -- le convertir perdrait les zeros de
    tete de maniere irreversible (float n'a pas de notion de zero non
    significatif conserve). Reste texte, comme dans le classeur fait a la
    main (colonne Ref.1).
    datetime/date -> NOMBRE SERIEL Excel (jours depuis 1899-12-30, meme epoque
    que Excel) : passer un datetime.datetime naif directement a une propriete
    COM Value le fait repasser par le fuseau horaire LOCAL de la machine lors
    de la conversion en PyTime (pywin32), ce qui peut faire glisser la date
    d'un jour (ex. 31/07 00:00 -> 30/07 22:00 constate). Une chaine 'JJ/MM/AAAA'
    est ambigue aussi (Excel COM ne la reconvertit pas forcement en date meme
    si la cellule est deja au format Date) -> le nombre seriel est la seule
    valeur non ambigue, la cellule (deja au format dd/mm/yyyy dans le modele)
    l'affiche nativement comme une date."""
    import datetime as _dt
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    if isinstance(v, _dt.datetime):
        return (v - EXCEL_EPOCH).total_seconds() / 86400
    if isinstance(v, _dt.date):
        return (_dt.datetime(v.year, v.month, v.day) - EXCEL_EPOCH).days
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.fullmatch(r"-?0\d+", s):
        return v
    if re.fullmatch(r"-?\d+(,\d+)?", s):
        return float(s.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


def argb_to_com_bgr(argb):
    """'FF3200E6' (ARGB hex, format openpyxl) -> entier BGR attendu par
    Range.Interior.Color / Font.Color en COM (B*65536 + G*256 + R)."""
    r, g, b = int(argb[2:4], 16), int(argb[4:6], 16), int(argb[6:8], 16)
    return b * 65536 + g * 256 + r


def col_letter(idx0):
    """Index 0-based (0=A) -> lettre(s) de colonne Excel."""
    import openpyxl
    return openpyxl.utils.get_column_letter(idx0 + 1)


def build_poste_formulas(header, first_raw_col):
    """Pour chaque poste ERP (colonnes F->M), cherche ses colonnes nommees
    (postes_from_columns) DANS L'ENTETE REEL recu ce mois-ci -> renvoie
    { "F": "=X2+Y2+...", ... } avec les BONNES lettres de colonnes pour ce
    fichier (peuvent differer du modele si des colonnes ont ete ajoutees/
    retirees/decalees). Une colonne de poste introuvable ce mois-ci est
    ignoree dans la somme (poste = 0 si aucune colonne trouvee) ; les noms
    manquants sont retournes a part pour avertissement, jamais bloquant."""
    postes_from_columns = load_postes_from_columns()
    key_by_pos = {compare_key(h): i for i, h in enumerate(header)}
    formulas, missing = {}, {}
    for col, poste_key in POSTE_COLS.items():
        names = postes_from_columns.get(poste_key, [])
        letters = []
        not_found = []
        for name in names:
            i = key_by_pos.get(compare_key(name))
            if i is None:
                not_found.append(name)
            else:
                letters.append(col_letter(first_raw_col - 1 + i))
        if letters:
            formulas[col] = "=" + "+".join(f"{L}{{row}}" for L in letters)
        if not_found:
            missing[poste_key] = not_found
    return formulas, missing


def extract_pdf_totals(pdf_paths):
    """Facture PDF Geodis : 'Montant Taxable HT' suivi de montants 'X,XX EUR',
    et une ligne dediee 'Frais de gestion de compte X,XX' (cf. adaptateur Node,
    meme logique -> voir src/carriers/geodis/index.js pour le detail du format).
    pypdf garde les espaces INSECABLES (\\xa0) du PDF entre les mots des libelles
    (ex. 'Montant\\xa0Taxable\\xa0HT') alors que pdf-parse (Node) les normalise en
    espaces simples -> on les normalise nous-memes avant toute recherche, sinon
    le libelle ne matche jamais et la reconciliation echoue silencieusement."""
    try:
        import pypdf
    except ImportError:
        return None, None
    taxable, frais = None, None
    for p in pdf_paths:
        try:
            text = "\n".join((page.extract_text() or "") for page in pypdf.PdfReader(p).pages)
        except Exception:
            continue
        text = text.replace("\xa0", " ")
        iTaxable = text.find("Montant Taxable HT")
        if iTaxable != -1:
            amounts = re.findall(r"([\d\s]+,\d{2})\s*EUR", text[iTaxable:])
            if amounts:
                taxable = float(amounts[0].replace(" ", "").replace(",", "."))
        m = re.search(r"Frais de gestion de compte\s*([\d\s]+,\d{2})", text)
        if m:
            frais = float(m.group(1).replace(" ", "").replace(",", "."))
    return taxable, frais


def fill_reconciliation(wb, pdf_taxable, frais_gestion):
    """Onglet Bilan Factures (TCD) : la colonne 'Frais de gestion' (poste sans
    tracking) et 'PDF' (montant taxable HT de la facture) sont saisies a la
    main dans le fichier fait a la main -> on les remplit automatiquement si
    fournies (deja extraites cote Node via pdf-parse).
    Les lignes du pivot demarrent APRES l'entete 'Etiquettes de lignes' (des
    libelles libres au-dessus, ex. 'Frais de tenue de compte 28,5EUR/mois', ne
    font pas partie du tableau -> les compter comme facture fausserait le
    garde-fou 'plusieurs factures' juste en dessous)."""
    ws = wb.Sheets("Bilan Factures")
    last = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
    header_row = None
    for r in range(1, last + 1):
        if str(ws.Cells(r, 1).Value or "").strip() == "Étiquettes de lignes":
            header_row = r
            break
    invoice_rows = []
    blank_row = None
    for r in range((header_row or 0) + 1, last + 1):
        label = str(ws.Cells(r, 1).Value or "").strip()
        if label in ("", "Total général"):
            continue
        if label == "(vide)":
            blank_row = r
        else:
            invoice_rows.append(r)
    if frais_gestion is not None and blank_row:
        ws.Cells(blank_row, 3).Value = frais_gestion  # colonne C = Frais de gestion
    if pdf_taxable is not None:
        if len(invoice_rows) == 1:
            ws.Cells(invoice_rows[0], 4).Value = pdf_taxable  # colonne D = PDF
        else:
            print(f"AVERTISSEMENT: {len(invoice_rows)} ligne(s) facture dans Bilan Factures "
                  f"(pas 1 seule) -> montant PDF ({pdf_taxable}) non affecté automatiquement, à saisir à la main.")


def parse_args(argv):
    """<modele> <sortie> <entree1> [<entree2>...] [--pdf <pdf1> [<pdf2>...]]"""
    modele, sortie = argv[1], argv[2]
    inputs, pdfs, cur = [], [], "in"
    for a in argv[3:]:
        if a == "--pdf":
            cur = "pdf"
        elif cur == "pdf":
            pdfs.append(a)
        else:
            inputs.append(a)
    return modele, sortie, inputs, pdfs


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def main():
    modele, sortie, inputs, pdf_paths = parse_args(sys.argv)
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    header, rows = read_input(inputs)
    n = len(rows)
    ncol = len(header)
    data = [[coerce(v) for v in (list(r) + [None] * ncol)[:ncol]] for r in rows]
    print(f"Entrée : {n} lignes x {ncol} colonnes")

    # Formules F->M reconstruites par NOM de colonne (pas position fixe) : robuste
    # si le fichier recu ce mois-ci a une colonne en plus/moins/decalee vs le
    # modele (cf. FACTURATION EXCEL.pdf p.3, confirme avril vs juin vs juillet).
    poste_formulas, missing_cols = build_poste_formulas(header, FIRST_RAW_COL)
    for poste_key, names in missing_cols.items():
        print(f"AVERTISSEMENT: colonne(s) introuvable(s) pour le poste {poste_key} "
              f"(ignorée(s) dans la somme, calculé sur les colonnes présentes) : {names}")

    pdf_taxable, frais_gestion = extract_pdf_totals(pdf_paths) if pdf_paths else (None, None)

    import win32com.client as win32
    xlUp = -4162
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) Feuille "Factures Geodis" : purge (entete + donnees, colonne N+ jusqu'a la
        #    fin de l'ancien contenu -> l'entete du fichier recu peut avoir moins/plus
        #    de colonnes que le modele), puis collage entete (avec le style bleu du
        #    modele, reapplique explicitement) + donnees brutes telles quelles.
        ws = wb.Sheets("Factures Geodis")
        oldLastCol = ws.Cells(1, ws.Columns.Count).End(-4159).Column  # -4159 = xlToLeft, depuis la derniere colonne
        lastCol = FIRST_RAW_COL + ncol - 1
        oldLast = ws.Cells(ws.Rows.Count, FIRST_RAW_COL).End(xlUp).Row
        newLast = 1 + n
        maxLast = max(oldLast, newLast)
        maxCol = max(oldLastCol, lastCol)
        retry(lambda: ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(maxLast, maxCol)).ClearContents())
        retry(lambda: ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(1, lastCol)).__setattr__("Value", [header]))
        headerRange = ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(1, lastCol))
        headerRange.Interior.Color = argb_to_com_bgr(HEADER_FILL_ARGB)
        headerRange.Font.Color = argb_to_com_bgr(HEADER_FONT_ARGB)
        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, lastCol)).__setattr__("Value", data))
        # 2) Reconstruire les formules des postes calcules (colonnes F->M) ligne a
        #    ligne, avec les lettres de colonnes resolues pour CE fichier -> puis
        #    etendre A/B/C/D/E (formules fixes du modele : Recepisse/Total hors GO/
        #    Total+GO) par recopie normale.
        for col, tmpl in poste_formulas.items():
            colIdx = ord(col) - ord("A") + 1
            retry(lambda c=colIdx, t=tmpl: ws.Range(ws.Cells(2, c), ws.Cells(newLast, c))
                  .__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(newLast, 5)).FillDown())
        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, maxCol)).ClearContents())

        # 3) Onglet "Import CSV" : formules cross-feuille (référencent 'Factures Geodis'
        #    ligne à ligne) -> étendre pareil, sur le même nombre de lignes.
        wsImp = wb.Sheets("Import CSV")
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 6).End(xlUp).Row  # col F = N Tracking
        if newLast > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 1), wsImp.Cells(newLast, LAST_COL_IMPORT_CSV)).FillDown())
        if newLast < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLast + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT_CSV)).ClearContents())

        # 4) Rafraîchir les TCD (Bilan Factures, Bilan Client)
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass

        # 5) Réconciliation (Frais de gestion / PDF), si fournis
        if pdf_taxable is not None or frais_gestion is not None:
            try:
                fill_reconciliation(wb, pdf_taxable, frais_gestion)
            except Exception as e:
                print("Réconciliation Bilan Factures ignorée :", e)

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)
        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
