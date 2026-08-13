#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR MONDIAL RELAY — produit "AAAA_MM_Facture Mondial Relay.xlsx" A
L'IDENTIQUE du fichier fait a la main (7 feuilles : Comptes, Controle xls pdf,
Pays, Facture Mondial Relay, Fichier import, Bilan clients, AVOIR ; memes
formules, memes TCD, meme mise en forme), en partant du fichier existant
comme MODELE.

Meme principe que finaliser_dpd.py/finaliser_geodis.py : les colonnes brutes
CSV ("Annexe_..." Mondial Relay) peuvent se decaler d'un mois a l'autre (cf.
notes terrain video), donc les colonnes calculees de "Facture Mondial Relay"
(A->D en tete, AU->BC en fin) ne sont PAS collees avec des formules a
positions fixes -- on colle l'entete recu TEL QUEL a partir de la colonne E,
PUIS on reconstruit les formules en cherchant CHAQUE colonne source par NOM
dans cet entete.

Feuilles statiques (jamais touchees, copiees telles quelles avec le modele) :
  Comptes, Pays, AVOIR (juste l'entete).

Colonnes calculees de "Facture Mondial Relay" (A->D, AU->BC) :
  A (Clients) : vide dans le modele (jamais rempli).
  B (Taxe gazole) = Indexation gasoil / Total htva (taux, pas montant).
  C (Nb car tracking) = LEN(tracking brut).
  D (Tracking) = tracking brut, complete par des zeros de tete si 5/6/7
    caracteres (padStart a 8), sinon tel quel.
  AU->BC : dimensions/poids convertis (cm, kg) + Max des 3 methodes de poids
    -- resolues par NOM des colonnes source (Longueur/Largeur/Hauteur/Poids
    annonce/Poids mesure), pas par position fixe.

"Fichier import" : formules FIXES referencant les colonnes A->D et E->AS de
"Facture Mondial Relay" (positions stables au sein de CETTE feuille, une fois
reconstruite) -- Zone = XLOOKUP(Mode&"-"&Pays, Pays!C:C, Pays!D:D).

"Bilan clients" (2 TCD) et "Controle xls pdf" (1 TCD) : VRAIS TCD Excel
dynamiques, PivotCache redirige vers la vraie plage de donnees du mois (meme
piege que DPD/Geodis : le cache du modele pointe une plage figee/etroite).

Reconciliation PDF (onglet "Controle xls pdf") : balise <ref>...
numeroFacture:X--montantHT:Y--montantTTC:Z...</ref> en pied de CHAQUE page
PDF Mondial Relay (format deja structure, tres fiable, confirme sur les 12
PDF de juin 2026) -- colle Q (TTC PDF) en face de la ligne 'Num facture'
correspondante ; N/O restent des FORMULES du modele (calcul theorique HT/TTC
depuis le TCD, K+L+M puis *1.2), pas touchees.

Necessite : Windows + Excel + pywin32 + pypdf (reconciliation PDF, optionnelle).
Usage :
  python finaliser_mondial_relay.py "<modele.xlsx>" "<sortie.xlsx>" "<entree1>" [<entree2>...] [--pdf <pdf1> [<pdf2>...]]
"""
import sys, os, shutil, re, csv, json

FIRST_RAW_COL = 5   # colonne E : debut des donnees brutes CSV dans "Facture Mondial Relay" (A-D = calculees)
# colonnes A->U (donnees reelles, 21 colonnes). PAS jusqu'a V : V1:V4 porte une note
# manuelle du modele ('ATTENTION AUX POIDS' / 'KG' / 'arrondi a 0,1'), jamais remplie
# au-dela de la ligne 4 dans le modele lui-meme -- un FillDown jusqu'a V aurait recopie
# cet exemple de juin sur TOUTES les lignes de donnees du mois traite (meme piege deja
# rencontre et corrige sur DPD, colonne X de son propre 'Import ERP').
LAST_COL_IMPORT = 21


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def compare_key(h):
    """Cle de comparaison ROBUSTE aux variantes cosmetiques (accents, points/virgules)
    -- meme logique que DPD/Geodis."""
    s = normalize_header(h).lower()
    s = s.translate(str.maketrans("éèêëàâäùûüôöîïç", "eeeeaaauuuooiic"))
    s = re.sub(r"[.,]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def is_xlsx(path):
    """Detection par contenu (magic bytes ZIP), pas par extension : les fichiers
    uploades (multer) arrivent parfois sans extension."""
    import zipfile
    try:
        return zipfile.is_zipfile(path)
    except Exception:
        return False


def read_xlsx_rows(path):
    import openpyxl, tempfile
    load_path = path
    tmp_path = None
    if not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copyfile(path, tmp.name)
        load_path = tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(load_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def read_csv_rows(path):
    with open(path, encoding="latin-1", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def read_input(paths):
    header = None
    all_rows = []
    for p in paths:
        result = read_xlsx_rows(p) if is_xlsx(p) else read_csv_rows(p)
        if result is None:
            raise RuntimeError(f"Fichier vide ou illisible : {p}")
        h, rows = result
        if header is None:
            header = h
        ncol = len(header)
        for r in rows:
            r = (list(r) + [None] * ncol)[:ncol]
            all_rows.append(r)
    return header, all_rows


def coerce(v):
    """Nombre si la valeur est un nombre francais/anglais valide, sinon telle quelle.
    Un entier a zero de tete reste texte (reference, pas une quantite)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.fullmatch(r"-?0\d+", s):
        return v
    if re.fullmatch(r"-?\d+(,\d+)?", s):
        return float(s.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


def col_letter(idx0):
    import openpyxl
    return openpyxl.utils.get_column_letter(idx0 + 1)


def col_index(header, name):
    target = compare_key(name)
    for i, h in enumerate(header):
        if compare_key(h) == target:
            return i
    return None


def build_facture_formulas(header, first_raw_col):
    """Formules des colonnes calculees de fin (AU->BC de 'Facture Mondial Relay'),
    resolues par NOM de colonne source dans l'entete recu ce mois-ci (pas position
    fixe -- les colonnes CSV Mondial Relay peuvent se decaler)."""
    def L(name):
        i = col_index(header, name)
        if i is None:
            return None
        return col_letter(first_raw_col - 1 + i)

    letters = {
        "longueur": L("Longueur"), "largeur": L("Largeur"), "hauteur": L("Hauteur"),
        "poids_annonce": L("Poids annoncé"), "poids_mesure": L("Poids mesuré"),
    }
    missing = [k for k, v in letters.items() if v is None]
    return letters, missing


def read_header_style(modele_path):
    import openpyxl
    wb = openpyxl.load_workbook(modele_path)
    ws = wb["Facture Mondial Relay"]
    cell = ws.cell(row=1, column=FIRST_RAW_COL)
    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    font = cell.font.color.rgb if cell.font and cell.font.color else None
    wb.close()
    return (fill if isinstance(fill, str) else "FFD9D9D9"), (font if isinstance(font, str) else "FF000000")


def argb_to_com_bgr(argb):
    r, g, b = int(argb[2:4], 16), int(argb[4:6], 16), int(argb[6:8], 16)
    return b * 65536 + g * 256 + r


def first_day_of_month_serial(header, rows, date_col_name="Date"):
    """1er jour du mois de la premiere 'Date' trouvee (format brut Mondial Relay :
    AAAAMMJJ, ex. '20260623'), en NOMBRE SERIEL Excel -- utilise pour 'Fichier import'!B2
    ('Date validite tarif'), une valeur FIXE saisie a la main dans le modele (=B2 recopie
    sur les lignes suivantes) qui reste sinon bloquee sur le mois du modele meme en
    traitant un autre mois (meme piege deja corrige sur DPD)."""
    import datetime as _dt
    i = col_index(header, date_col_name)
    if i is None:
        return None
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    for r in rows:
        v = r[i] if i < len(r) else None
        s = str(v or "").strip()
        m = re.match(r"^(\d{4})(\d{2})(\d{2})$", s)
        if m:
            first = _dt.datetime(int(m.group(1)), int(m.group(2)), 1)
            return (first - EXCEL_EPOCH).days
    return None


def extract_pdf_infos(pdf_paths):
    """PDF Mondial Relay : balise <ref>codeClient:X--libelleClient:Y--numeroFacture:Z--
    montantHT:A--montantTVA:B--montantTTC:C--tauxTVA:D--dateFacture:E</ref> en pied de
    CHAQUE page (format deja structure, confirme identique sur les 12 PDF de juin 2026).
    numeroFacture correspond exactement au nom du CSV Annexe_<numeroFacture>_... (colonne
    'Nr de facture') -> cle de rapprochement directe."""
    try:
        import pypdf
    except ImportError:
        return []
    results = []
    for p in pdf_paths:
        try:
            text = "\n".join((page.extract_text() or "") for page in pypdf.PdfReader(p).pages)
        except Exception:
            continue
        m = re.search(r"<ref>(.+?)</ref>", text)
        if not m:
            continue
        fields = dict(kv.split(":", 1) for kv in m.group(1).split("--"))
        if "numeroFacture" not in fields or "montantHT" not in fields:
            continue

        def to_num(s):
            return float(s.replace(".", "").replace(",", "."))

        results.append({
            "file": os.path.basename(p),
            "numeroFacture": fields["numeroFacture"],
            "montantHt": to_num(fields.get("montantHT", "0")),
            "montantTtc": to_num(fields.get("montantTTC", "0")),
        })
    return results


def fill_reconciliation(wb, pdfs):
    """Onglet 'Controle xls pdf' : colle le TTC extrait du PDF en colonne Q (en face
    de la ligne 'Num facture'/J correspondante -- le TCD source). N/O restent des
    FORMULES du modele (calcul theorique HT puis TTC = *1.2 depuis le TCD K/L/M),
    jamais touchees ici. Purge Q sur toute la hauteur avant de reecrire (meme piege
    deja rencontre sur DPD : une ancienne valeur de mois precedent pourrait rester
    affichee sur une ligne devenue differente si le TCD a change de taille)."""
    xlUp = -4162
    ws = wb.Sheets("Controle xls pdf")
    lastJ = ws.Cells(ws.Rows.Count, 10).End(xlUp).Row  # J = etiquettes de lignes (TCD)
    header_row = None
    for r in range(1, lastJ + 1):
        if str(ws.Cells(r, 10).Value or "").strip() == "Étiquettes de lignes":
            header_row = r
            break
    start = (header_row or 2) + 1

    lastQ = ws.Cells(ws.Rows.Count, 17).End(xlUp).Row  # Q = TTC (PDF)
    if lastQ >= start:
        ws.Range(ws.Cells(start, 17), ws.Cells(lastQ, 17)).ClearContents()

    if not pdfs:
        return

    facture_to_row = {}
    for r in range(start, lastJ + 1):
        v = str(ws.Cells(r, 10).Value or "").strip()
        if v not in ("", "(vide)", "Total général"):
            facture_to_row[v] = r

    matched = 0
    for p in pdfs:
        row = facture_to_row.get(p["numeroFacture"])
        if row is None:
            print(f"AVERTISSEMENT: facture PDF {p['file']} ({p['numeroFacture']}) -> "
                  f"introuvable dans Controle xls pdf, montant non affecté.")
            continue
        ws.Cells(row, 17).Value = p["montantTtc"]
        matched += 1
    print(f"Réconciliation PDF : {matched}/{len(pdfs)} facture(s) rapprochée(s) dans Controle xls pdf.")


def parse_args(argv):
    modele, sortie = argv[1], argv[2]
    inputs, pdfs, cur = [], [], "in"
    for a in argv[3:]:
        if a == "--pdf":
            cur = "pdf"
        elif cur == "pdf":
            pdfs.append(a)
        else:
            inputs.append(a)
    return modele, sortie, inputs, pdfs


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def main():
    modele, sortie, inputs, pdf_paths = parse_args(sys.argv)
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    header, rows = read_input(inputs)
    n = len(rows)
    ncol = len(header)
    data = [[coerce(v) for v in r] for r in rows]
    print(f"Entrée : {n} lignes x {ncol} colonnes")

    letters, missing = build_facture_formulas(header, FIRST_RAW_COL)
    if missing:
        print(f"AVERTISSEMENT: colonne(s) introuvable(s) pour les formules calculées : {missing} "
              f"(colonnes AU->BC laissées vides pour ces cas)")

    date_validite_serial = first_day_of_month_serial(header, rows)
    header_fill, header_font = read_header_style(modele)
    pdfs = extract_pdf_infos(pdf_paths) if pdf_paths else []

    import win32com.client as win32
    xlUp = -4162
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) "Facture Mondial Relay" : purge + collage entete (avec style) + donnees brutes,
        #    puis reconstruction des formules A-D (Clients/Taxe gazole/Nb car/Tracking) et
        #    AU-BC (dimensions/poids convertis), toutes deux etendues par FillDown.
        ws = wb.Sheets("Facture Mondial Relay")
        oldLastCol = ws.Cells(1, ws.Columns.Count).End(-4159).Column  # xlToLeft
        lastCol = FIRST_RAW_COL + ncol - 1  # derniere colonne de donnees brutes recues
        oldLast = ws.Cells(ws.Rows.Count, FIRST_RAW_COL).End(xlUp).Row
        newLast = 1 + n
        maxLast = max(oldLast, newLast)

        # AU->BC (colonnes calculees de fin) : conserver leur POSITION du modele (pas suivre
        # le decalage de largeur des donnees brutes recues) -- elles referencent les colonnes
        # brutes par LETTRE resolue dynamiquement (build_facture_formulas), pas par offset fixe.
        AU_COL = 47  # position fixe dans le modele (colonne AU), quel que soit ncol recu
        BC_COL = 55  # derniere colonne calculee (Max des 3)
        maxCol = max(oldLastCol, lastCol, BC_COL)

        retry(lambda: ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(maxLast, maxCol)).ClearContents())
        retry(lambda: ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(1, lastCol)).__setattr__("Value", [header]))
        headerRange = ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(1, lastCol))
        headerRange.Interior.Color = argb_to_com_bgr(header_fill)
        headerRange.Font.Color = argb_to_com_bgr(header_font)
        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, lastCol)).__setattr__("Value", data))

        # Colonnes A->D : Clients (lookup Comptes), Taxe gazole, Nb car tracking, Tracking (padding).
        i_facture = col_index(header, "Nr de facture")
        i_track = col_index(header, "Nr expédition")
        i_gazole = col_index(header, "Indexation gasoil")
        i_total_htva = col_index(header, "Total htva")
        i_transport = col_index(header, "Mnt.Trans. Poids Mes")
        i_complement = col_index(header, "Complément")
        i_semi = col_index(header, "Semi")
        i_code_client = col_index(header, "Code client chargeur")
        col_track = col_letter(FIRST_RAW_COL - 1 + i_track) if i_track is not None else None
        col_gazole = col_letter(FIRST_RAW_COL - 1 + i_gazole) if i_gazole is not None else None
        col_htva = col_letter(FIRST_RAW_COL - 1 + i_total_htva) if i_total_htva is not None else None
        col_transport = col_letter(FIRST_RAW_COL - 1 + i_transport) if i_transport is not None else None
        col_complement = col_letter(FIRST_RAW_COL - 1 + i_complement) if i_complement is not None else None
        col_semi = col_letter(FIRST_RAW_COL - 1 + i_semi) if i_semi is not None else None
        col_code_client = col_letter(FIRST_RAW_COL - 1 + i_code_client) if i_code_client is not None else None

        formulas_ad = {}
        # 'Clients' (colonne A) = code numerique du client, saisi A LA MAIN par le pole
        # transport dans le fichier fait a la main (PAS une formule -- confirme sur le fichier
        # de juillet 2026 : valeurs litterales '3752', '9294'...), via lookup manuel dans
        # 'Comptes' (colonne A=Code client chargeur texte, colonne C=code numerique). Autant
        # que possible reconstruit ici par XLOOKUP (les clients deja dans 'Comptes' se
        # remplissent automatiquement) -- les clients ABSENTS de 'Comptes' (ex. OCAREO/9243
        # en juillet, jamais ajoute a la table) restent vides, a completer manuellement comme
        # avant (la table 'Comptes' elle-meme n'est jamais modifiee par ce script).
        if col_code_client:
            formulas_ad["A"] = (f'=IF(COUNTIF(Comptes!A:A,{col_code_client}{{row}})=0,"",'
                                 f'_xlfn.XLOOKUP({col_code_client}{{row}},Comptes!A:A,Comptes!C:C))')
        if col_gazole and col_htva:
            formulas_ad["B"] = f'=IF({col_htva}{{row}}=0,0,{col_gazole}{{row}}/{col_htva}{{row}})'
        if col_track:
            formulas_ad["C"] = f"=LEN({col_track}{{row}})"
            formulas_ad["D"] = (f'=IF({col_track}{{row}}=0,0,IF(C{{row}}=5,"000"&{col_track}{{row}},'
                                 f'IF(C{{row}}=6,"00"&{col_track}{{row}},IF(C{{row}}=7,"0"&{col_track}{{row}},{col_track}{{row}}))))')

        for col, tmpl in formulas_ad.items():
            colIdx = {"A": 1, "B": 2, "C": 3, "D": 4}[col]
            retry(lambda c=colIdx, t=tmpl: ws.Range(ws.Cells(2, c), ws.Cells(newLast, c))
                  .__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))

        # Colonnes AU->BC : dimensions/poids convertis, colonnes source resolues par nom.
        if not missing:
            au_formulas = {
                49: f"={letters['longueur']}{{row}}/10",   # AW: Longueur cm
                50: f"={letters['largeur']}{{row}}/10",     # AX: Largeur cm
                51: f"={letters['hauteur']}{{row}}/10",     # AY: Hauteur cm
                52: "=((AW{row}*AX{row}*AY{row})/5000)/5",  # AZ: Poids volumetrique/5
                53: f"={letters['poids_annonce']}{{row}}/1000",  # BA: Poids annonce kg
                54: f"={letters['poids_mesure']}{{row}}/1000",   # BB: Poids mesure kg
                55: "=MAX(AZ{row},BA{row},BB{row})",        # BC: Max des 3
            }
            for colIdx, tmpl in au_formulas.items():
                retry(lambda c=colIdx, t=tmpl: ws.Range(ws.Cells(2, c), ws.Cells(newLast, c))
                      .__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))

        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, maxCol)).ClearContents())

        # 2) "Fichier import" : formules a positions FIXES (referencent Facture Mondial
        #    Relay/Pays, stables une fois "Facture Mondial Relay" reconstruite) -- juste
        #    etendre/reduire par recopie normale depuis la ligne 2 (modele des formules).
        #    EXCEPTION : B2 ("Date validite tarif") est une VALEUR FIXE saisie a la main
        #    dans le modele (B3+ = "=B2" recopie) -> mise a jour vers le 1er du mois
        #    traite AVANT le FillDown, sinon elle reste bloquee sur le mois du modele
        #    (meme piege deja rencontre et corrige sur DPD).
        wsImp = wb.Sheets("Fichier import")
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 6).End(xlUp).Row  # F = Tracking
        if date_validite_serial is not None:
            wsImp.Cells(2, 2).Value = date_validite_serial
        else:
            print("AVERTISSEMENT: 'Date' introuvable dans le fichier reçu -> "
                  "'Date validité tarif' (Fichier import!B2) non mise à jour, reste celle du modèle.")
        # Q2/T2 (modele) referencent 'Facture Mondial Relay'!AH/AI (Montant transport/
        # Complement) par POSITION FIXE avec le seuil 0,03 -- CE SEUIL EST CORRECT (verifie en
        # recalculant les totaux sur 0,03 vs 0,08 et en comparant a 'Controle xls pdf'!K16/M16 :
        # seul 0,03 redonne les valeurs de reference exactes, 1545.20EUR/36385.88EUR -- une
        # premiere lecture avait conclu a tort a 0,08 en se fiant aux formules Q3/T3+ du modele,
        # qui semblent elles-memes non representatives). Reecrites ICI uniquement pour resoudre
        # AH/AI PAR NOM (pas position fixe -- ces lettres ne sont stables que si le nombre de
        # colonnes brutes recues ne change jamais d'un mois a l'autre, meme piege que 'Facture
        # Mondial Relay' elle-meme), avec le MEME seuil 0,03 que le modele.
        if col_transport and col_complement:
            wsImp.Cells(2, 17).Formula = (f"=IF('Facture Mondial Relay'!{col_complement}2=0,\"\","
                                           f"IF('Facture Mondial Relay'!{col_complement}2=0.03,\"\","
                                           f"'Facture Mondial Relay'!{col_complement}2))")
            wsImp.Cells(2, 20).Formula = (f"=ROUND(IF('Facture Mondial Relay'!{col_complement}2=0.03,"
                                           f"'Facture Mondial Relay'!{col_complement}2+'Facture Mondial Relay'!{col_transport}2,"
                                           f"'Facture Mondial Relay'!{col_transport}2),2)")
        else:
            print("AVERTISSEMENT: 'Mnt.Trans. Poids Mes' ou 'Complément' introuvable -> "
                  "formules Q2/T2 (Zones éloignées/Frêt) non corrigées, restent celles du modèle (seuil 0,03).")
        if newLast > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 1), wsImp.Cells(newLast, LAST_COL_IMPORT)).FillDown())
        if newLast < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLast + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT)).ClearContents())

        # 2bis) COLLECTE : 'Bilan clients'!F3 et 'Controle xls pdf'!S30 sont des VALEURS FIXES
        #    saisies a la main dans le modele (480 / 570 respectivement dans le modele de juin
        #    2026 -- deja incoherentes entre elles, signe que ce sont deux saisies manuelles
        #    distinctes jamais synchronisees), PAS des formules -- contrairement aux autres
        #    lignes de 'Bilan clients' (Fret/Plus-value/gazole/total, deja de vraies formules
        #    SUM sur 'Facture Mondial Relay'). La vraie valeur de la collecte = SOMME de la
        #    colonne 'Semi' du fichier recu (confirme par le pole transport : correspond au
        #    montant de collecte du PDF, ex. 570,00€ = 19 x 30,00€ sur la facture LARUCH de
        #    juin) -> recalculee ICI a chaque mois, remplace les 2 valeurs figees.
        if col_semi:
            i_semi0 = i_semi
            total_semi = round(sum((coerce(r[i_semi0]) or 0) for r in rows if isinstance(coerce(r[i_semi0]), (int, float))), 2)
            wb.Sheets("Bilan clients").Cells(3, 6).Value = total_semi  # F3
            wb.Sheets("Controle xls pdf").Cells(30, 19).Value = total_semi  # S30
            print(f"Collecte (colonne 'Semi') recalculée : {total_semi} € (remplace la valeur figée du modèle).")
        else:
            print("AVERTISSEMENT: colonne 'Semi' introuvable -> collecte ('Bilan clients'!F3 et "
                  "'Controle xls pdf'!S30) non recalculée, reste la valeur figée du modèle.")

        # 3) "Bilan clients" (2 TCD) et "Controle xls pdf" (1 TCD) : PivotCache redirige
        #    vers la vraie plage de donnees de ce mois (meme piege que DPD/Geodis : le
        #    cache du modele pointe une plage figee/etroite, RefreshAll() seul ne suffit
        #    pas). ChangePivotCache() juste apres recreation du cache, jamais apres un
        #    retrait+reecriture manuelle du TCD (corrompt le RowField, deja constate).
        #    Plage = A -> 'Total htva' UNIQUEMENT (verifie sur le modele : le PivotCache
        #    reel s'arrete a ce champ, PAS jusqu'aux colonnes calculees de fin AU->BC qui
        #    n'en font pas partie) -- resolue par NOM (pas figee a une colonne fixe, le
        #    nombre de colonnes brutes recues peut varier d'un mois a l'autre).
        #
        #    PIEGE SPECIFIQUE 'Bilan clients' : ce classeur a 2 TCD sur cette feuille. UN SEUL
        #    (Location A1:B26, RowField='Clients') a pour vraie source les donnees brutes de
        #    'Facture Mondial Relay' -- CELUI-LA est redirige/rafraichi correctement ci-dessous.
        #    Le 2e (Location L1:M28, RowField='Etiquettes de lignes', DataField='Somme de Somme
        #    de Total htva') est un TCD IMBRIQUE dont la vraie source est le 1er TCD lui-meme
        #    (ses PivotFields disponibles = exactement les 2 colonnes du 1er TCD) -- son
        #    SourceData pointe deja, meme dans le MODELE ORIGINAL et dans le fichier fait a la
        #    main de juillet 2026, vers 'Bilan clients'!C9:C10 (une plage locale vide, vestige
        #    du process manuel). CE 2e TCD RESTE NON RESOLU (limite connue, 2026-08-10) :
        #    - Le rediriger vers les donnees brutes casse sa config (RowField/DataField perdus).
        #    - RefreshTable() sans toucher son PivotCache ne le met PAS a jour non plus (teste :
        #      reste bloque sur l'etat/mois du MODELE de juin, meme apres RefreshAll()+Calculate).
        #    Il reste donc affiche avec des valeurs figees heritees du modele clone, PAS les
        #    vraies donnees du mois traite -- a rouvrir si une piste plus solide est trouvee
        #    (ex. supprimer/recreer le TCD depuis zero), sinon le TCD principal (A1:B26)
        #    suffit pour le detail par client.
        #
        #    NOTE (colonne 'Clients') : REMPLIE via XLOOKUP vers 'Comptes' (cf. formula_ad['A']
        #    plus haut) -- confirme sur le fichier fait a la main de juillet 2026 : le pole
        #    transport la saisit A LA MAIN chaque mois (valeurs litterales, pas une formule),
        #    avec le code numerique du client -- PAS toujours vide comme suppose initialement
        #    sur la seule base du modele de juin (qui, lui, etait effectivement vide, un cas
        #    particulier pas la regle generale). Les clients absents de 'Comptes' restent
        #    vides (XLOOKUP echoue), a completer manuellement comme le fait deja le pole
        #    transport pour un nouveau client.
        newLastPivot = max(newLast, 2)
        xlDatabase = 1
        i_total_htva_pivot = col_index(header, "Total htva")
        pivotLastCol = (FIRST_RAW_COL - 1 + i_total_htva_pivot + 1) if i_total_htva_pivot is not None else lastCol
        newRange = ws.Range(ws.Cells(1, 1), ws.Cells(newLastPivot, pivotLastCol))
        pivots_non_redired = []
        for sheetName in ("Bilan clients", "Controle xls pdf"):
            wsPivot = wb.Sheets(sheetName)
            for i in range(1, wsPivot.PivotTables().Count + 1):
                pt = wsPivot.PivotTables(i)
                src = str(pt.PivotCache().SourceData)
                if "Facture Mondial Relay" not in src:
                    print(f"TCD '{pt.Name}' ({sheetName}) : source '{src}' n'est pas 'Facture Mondial Relay' "
                          f"(TCD imbrique sur un autre TCD) -> config preservee mais reste affiche avec les "
                          f"valeurs figees du modele (limite connue, ni ChangePivotCache ni RefreshTable ne "
                          f"le mettent a jour -- cf. docstring ci-dessus).")
                    pivots_non_redired.append(pt)
                    continue
                newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=newRange)
                pt.ChangePivotCache(newCache)
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        xl.Calculate()

        # 4) Réconciliation PDF (colonne Q de Controle xls pdf). Appelee MEME SANS PDF
        #    fourni : purge quand meme les anciennes valeurs laissees par le modele clone.
        try:
            fill_reconciliation(wb, pdfs)
        except Exception as e:
            print("Réconciliation Controle xls pdf ignorée :", e)

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)
        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
