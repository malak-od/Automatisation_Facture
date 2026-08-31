#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR UPS — produit "AAAA_MM_Facture UPS.xlsx" A L'IDENTIQUE du fichier
fait a la main (16 feuilles : Gazole, CODES SVCE LEVEL, CODIFICATION CODE
EXCEPTION, Charge.CHG_CODE, ST SV, Zone, Clients log, zone colis poids
assurance, Bilan factures, Facture UPS, Comptes UPS, TCD, Fichier import,
Demande avoir, Bilan clients, Adresse), en partant du fichier existant
comme MODELE.

Classeur reel :
  Facture UPS : 1 ligne = 1 evenement de facturation UPS Billing (PAS 1
    colis), donnees brutes CSV UPS (SANS EN-TETE, 250 colonnes, colonnes
    resolues PAR POSITION -- decalage CONFIRME -4 : position CSV brut N =
    colonne Facture UPS N+4), colonnes A->D CALCULEES : A="Clients"=vide
    (manuel), B="Montant assurance"=IF(codeDescription="EVS",valeurBase,0),
    C="Mode envoi"=IF(codeClasse="FRT",XLOOKUP(description,'ST SV'!A:A,B:B),0),
    D="Categorie"=cascade NB.SI('ST SV'!Q:Q/D:D)->Adresse/plus-value BtoC,
    puis codeClasse="FRT"->Fret, "TAX"->TVA, sinon XLOOKUP(codeDescription,
    'Charge.CHG_CODE'!A:C) -- "CODE INCONNU" est une VRAIE categorie du
    modele reel (239/472 codes), jamais reclassee.
  5 TCD (tous PivotCache source ETROITE/FIGEE sur le modele, meme piege
    que DPD/Geodis/Mondial Relay/Chronopost/TNT/FedEx) :
    - "Bilan factures" (2 TCD) : par compte, et par date+numero facture.
    - "zone colis poids assurance" : Max Nombre de colis/Zone/Assurance/
      Poids par tracking.
    - "ST SV" : Nombre de Categorie par tracking x Mode envoi (ST/SV/vide).
    - "TCD" : Somme Montant net par tracking x Categorie.
    - "Bilan clients" (source='TCD', PAS 'Facture UPS') : par client.
  Colonnes MANUELLES juxtaposees (etirees apres redirection PivotCache) :
    - "zone colis poids assurance"!B (Logistique)=COUNTIF('Clients log'!
      A:A,A), C (Colis)=IF(E=0,1,E), I (Poids UPS)=ROUNDUP(H,0), J (Poids
      UPS_COD)=IF(F>3,ROUNDUP(H,0),IF(H<10,ROUNDUP(H,1),ROUNDUP(H,0))).
    - "ST SV"!N = IF(K<>"","SV","ST").
    - "TCD"!B (Cout)=SUM(F:K)+O+M (controle), D (Poids)=XLOOKUP vers
      'zone colis poids assurance'!I. A (Logistique)/C (Client) manuels,
      jamais reconstruits (Client = regle transversale ID client, saisie
      humaine post-generation).
    - "Fichier import" : formules PAR LIGNE completes (cf. carrier Node
      index.js pour le detail complet des 24 colonnes, notamment
      Transporteur=compte extrait du tracking via 'Comptes UPS', Zone=
      FOURNIE par UPS (native XLOOKUP) SAUF regles CDC 2026-08-27
      ci-dessous, E/P=cascade ERP/plus-value BtoC, TVA=IF(TCD!N="",0,0.2)
      (poste TVA reel, PAS liste de pays -- deja naturellement 0 hors UE,
      valide a posteriori en Python, jamais recalcule), Colis
      volumineux=bareme par palier sur TCD!H (poste ERP MONTANT, PAS le
      poids -- piege deja documente).
      Regles Zone (CDC pole transport 2026-08-27, colonne M) : 1) Zone=0
      interdite (garantie une fois le repli M-1 applique) ; 2) Zone=0 +
      Fret 3-8EUR + Pays vide/FR -> "France" ; 3) Zone=0 sans fret ->
      repli Python sur le fichier import CSV du mois precedent
      (load_import_m1, champ upload dedie --import-m1) ; 4) plus-value
      BtoC<2EUR + Zone="France" (hors WV5788) -> marqueur "A VERIFIER" ;
      5) compte WV5788 (Verde Trad) -> Zone="France" + Mode envoi="ST"
      forces (override absolu, colonnes M et P). Regles 6/7 (validation
      croisee zoning UPS Pays->SV/ST, table PAYS_SV_ST) : alerte console
      uniquement, ne modifie jamais Zone/Mode envoi.

Colis 1Z79 (regle FACTURATION EXCEL.docx) : colis viticulteur retourne
  chez La Ruche -- EXCLUS de Facture UPS (jamais colles), a signaler pour
  demande d'avoir (infos console).

Trackings SANS AUCUNE charge facturable (tous postes ERP a 0, ligne
  "INF"/retour indelivrable isolee sans ligne FRT) : EXCLUS de "Fichier
  import" (confirme par comparaison au fichier reel de juin 2026 :
  8736/8736 lignes ont au moins un poste non nul, 0 ligne totalement
  vide) -- mais RESTENT dans "Facture UPS" (donnees brutes completes,
  la colonne "Facture UPS" du classeur modele n'est jamais filtree).

Necessite : Windows + Excel + pywin32.
Usage :
  python finaliser_ups.py "<modele.xlsx>" "<sortie.xlsx>" --csv <csv1> [<csv2>...] [--brut <brutM.xlsx> [<brutM-1.xlsx>]]
"""
import sys, os, shutil, re, csv, time


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def coerce(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.fullmatch(r"-?0\d+", s):
        return v
    if re.fullmatch(r"-?\d+(,\d+)?", s):
        return float(s.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


def num(x):
    try:
        return float(str(x).replace(",", ".").strip())
    except (ValueError, AttributeError):
        return 0.0


# Positions CSV brut UPS (0-based) -- decalage CONFIRME -4 vs colonnes "Facture UPS" (1-based) :
# position CSV brut N (1-based) = colonne Facture UPS (N+4). Formule uniforme : <col FU> - 4 - 1.
COL = {
    "numero_compte": 7 - 4 - 1,
    "date_facture": 9 - 4 - 1,
    "numero_facture": 10 - 4 - 1,
    "ref1": 20 - 4 - 1,
    "numero_suivi": 25 - 4 - 1,
    "nombre_colis": 23 - 4 - 1,
    "poids_facture": 33 - 4 - 1,
    "zone": 38 - 4 - 1,
    "code_classe": 48 - 4 - 1,
    "code_description": 49 - 4 - 1,
    "description": 50 - 4 - 1,
    "valeur_base": 53 - 4 - 1,
    "montant_net": 57 - 4 - 1,
    "pays": 86 - 4 - 1,
    # "Poids facture"=0 sur les lignes d'AJUSTEMENT/correction (codeClasse=ACC, ex. "Frais de
    # correction d'expedition") -- confirme sur reel juillet 2026 (tracking 1ZA1912WD998484245,
    # verifie contre le site UPS). Le vrai poids est ecrit en texte libre dans cette colonne,
    # format "AUDITED WEIGHT: 43.5 KGS" -- meme position que carriers/ups/index.js (COL.auditedWeight).
    "audited_weight": 180 - 4 - 1,
}

# Table Pays -> Zone SV/ST fournie par le pole transport (2026-08-27), a comparer avec le
# zoning 2026 UPS (Guide des Services Viticolis) -- USAGE VALIDATION UNIQUEMENT (cf. C.4
# ci-dessous) : ne modifie JAMAIS Zone/Mode envoi calcules, sert seulement a alerter en
# console si le calcul existant (XLOOKUP vers l'onglet "Zone" natif UPS) est incoherent avec
# cette table. Forme : code -> (zones_sv possibles, zones_st possibles), chaque zone en
# CHAINE (ex. US="9"/"10" -- garde eclate en valeurs separees, pas la chaine litterale "9-10",
# pour accepter Zone=9 OU Zone=10 comme valides). Tuple vide = pas de valeur fournie ("vide").
PAYS_SV_ST = {
    "AT": (("6",), ("6",)),
    "AU": ((), ()),
    "BE": (("3",), ("4", "5")),
    "BG": (("52",), ("71",)),
    "CH": (("6",), ("8",)),
    "CN": (("11",), ()),
    "CZ": (("51",), ("61",)),
    "DE": (("3",), ("4", "5", "6", "7")),
    "DK": (("4",), ("6", "7")),
    "ES": (("3",), ("4", "5", "6", "7")),
    "FI": (("5",), ("7",)),
    "GB": (("703",), ("704", "705", "706", "707")),
    "GR": (("4",), ("71",)),
    "HK": (("11",), ()),
    "HR": (("51",), ("61",)),
    "HU": (("51",), ("61",)),
    "IE": (("4",), ("6", "7")),
    "IT": (("3",), ("4", "5", "6", "7")),
    "JP": (("11",), ()),
    "KR": (("11",), ()),
    "LT": (("52",), ("71",)),
    "LU": (("3",), ("4", "5")),
    "LV": (("52",), ("71",)),
    "NL": (("3",), ("5", "6")),
    "PL": (("51",), ("61",)),
    "PT": (("5",), ("6", "7", "999")),
    "RO": (("52",), ("71",)),
    "SE": (("5",), ("7",)),
    "SI": (("51",), ("61",)),
    "SK": (("51",), ("61",)),
    "TW": (("11",), ()),
    "US": (("9", "10"), ()),
    "VI": (("12",), ()),
}

# Sous-ensemble UE de PAYS_SV_ST (+FR, absent de la table car jamais recherche via XLOOKUP
# 'Zone' -- deja code en dur "France" dans la formule Zone) -- utilise par la validation TVA
# hors UE (point B) : TVA=0.2 calculee pour un pays absent de cet ensemble = suspect.
PAYS_UE = {
    "AT", "BE", "BG", "CZ", "DE", "DK", "ES", "FI", "GR", "HR", "HU", "IE", "IT",
    "LT", "LU", "LV", "NL", "PL", "PT", "RO", "SE", "SI", "SK", "FR",
}


def read_ups_csv(path):
    """CSV UPS Billing brut, SANS EN-TETE, separateur virgule, latin-1."""
    with open(path, encoding="latin-1", newline="") as f:
        rows = list(csv.reader(f))
    return [r for r in rows if any(v != "" for v in r)]


def load_config(carrier_dir):
    import json
    with open(os.path.join(carrier_dir, "config.json"), encoding="utf-8") as f:
        return json.load(f)


# VESTIGE (2026-08-27) : aucun appel a cette fonction dans ce fichier -- le calcul REEL du
# poids arrondi passe par les formules Excel natives 'zone colis poids assurance'!I/J (cf.
# formulas_zcp, ~ligne 707). Conservee telle quelle (pas supprimee, risque de casser un appel
# externe non detecte), mais ne pas s'y fier pour comprendre le comportement livre au client.
def poids_arrondi(transporteur, nb_colis, poids):
    import math
    if transporteur == "UPS_COD":
        if nb_colis > 3:
            return math.ceil(poids)
        if poids < 10:
            return math.ceil(poids * 10) / 10
        return math.ceil(poids)
    return math.ceil(poids)


def colis_volumineux_montant(montant_reel):
    import math
    if not montant_reel:
        return 0
    if montant_reel < 3:
        return 3
    if montant_reel < 15:
        return 15
    if montant_reel < 50:
        return 35
    if montant_reel < 100:
        return 59
    if montant_reel < 150:
        return 177
    return math.ceil(montant_reel / 59) * 59


def parse_args(argv):
    """--csv <csv...> [--brut <xlsx...>] [--import-m1 <csv>] [--period AAAA_MM]"""
    modele, sortie = argv[1], argv[2]
    csvs, brut, cur = [], [], None
    period = None
    import_m1 = None
    for a in argv[3:]:
        if a == "--csv":
            cur = "c"
        elif a == "--brut":
            cur = "b"
        elif a == "--import-m1":
            cur = "m1"
        elif a == "--period":
            cur = "p"
        elif cur == "c":
            csvs.append(a)
        elif cur == "b":
            brut.append(a)
        elif cur == "m1":
            import_m1 = a
            cur = None
        elif cur == "p":
            period = a
            cur = None
    return modele, sortie, csvs, brut, period, import_m1


def load_brut_ep(brut_paths):
    """Map {tracking -> 'entreprise'/'particulier'} depuis l'export WMS partage (meme
    mecanisme que Delivengo/Geodis/DPD/FedEx) -- colonnes AP=PRO_TRACKING(41)/Q=
    DES_PARTICULIER(16), 0-based. Point relais -> entreprise."""
    import io
    import openpyxl
    m = {}
    for p in [bp for bp in brut_paths if bp]:
        # BUG TROUVE 2026-08-26 : openpyxl refuse un chemin SANS EXTENSION (InvalidFileException
        # "openpyxl does not support  file format") -- les fichiers uploades via multer sont
        # renommes en identifiant hexadecimal sans extension (piege deja connu, cf. BLS). Fix :
        # ouvrir en binaire et passer un buffer memoire (openpyxl accepte un objet fichier-like,
        # detecte le format par SIGNATURE ZIP interne, pas par le nom de fichier).
        with open(p, "rb") as f:
            buf = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 41:
                t = str(row[41] or "").strip()
                v = str(row[16] or "").strip().lower()
                ep = "particulier" if v == "particulier" else ("entreprise" if v in ("entreprise", "point relais") else None)
                if t and ep and t not in m:
                    m[t] = ep
        wb.close()
    return m


def load_brut_poids_colis_pays(brut_paths):
    """Maps {tracking -> poids}, {tracking -> nb_colis} et {tracking -> pays destinataire}
    depuis l'export WMS partage -- meme mecanisme que load_brut_ep, colonnes
    AI=INFO_POIDSRETENU(34), AH=INFO_NBCOLIS(33), AB=DES_PAYS(27), 0-based. Premier fichier
    de brut_paths ayant une valeur pour ce tracking gagne -- brut_paths doit deja etre trie
    mois courant -> plus ancien (fait cote carrier Node avant l'appel CLI, cf.
    computeFinalizerArgs). Lecture fusionnee (poids+colis+pays en un seul passage par
    fichier)."""
    import io
    import openpyxl
    m_poids, m_colis, m_pays = {}, {}, {}
    for p in [bp for bp in brut_paths if bp]:
        with open(p, "rb") as f:
            buf = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 41:
                t = str(row[41] or "").strip()
                if not t:
                    continue
                poids = num(row[34]) if row[34] not in (None, "") else 0
                if poids > 0 and t not in m_poids:
                    m_poids[t] = poids
                colis = num(row[33]) if row[33] not in (None, "") else 0
                if colis > 0 and t not in m_colis:
                    m_colis[t] = colis
                pays = str(row[27] or "").strip().upper() if len(row) > 27 else ""
                if pays and t not in m_pays:
                    m_pays[t] = pays
        wb.close()
    return m_poids, m_colis, m_pays


def load_import_m1(path):
    """Map {tracking -> zone} depuis le fichier import CSV FINAL du mois precedent (deja
    livre a l'ERP -- format ecrit par writeImportCsv() cote Node : latin1, ';', decimale
    virgule) -- meme mecanisme que le repli 'zone colis poids assurance' vu dans la video
    process (RECHERCHEX vers 2026_04_UPS_Import.csv). Repli de dernier recours pour Zone=0
    SANS fret (demande utilisateur 2026-08-27)."""
    if not path:
        return {}
    m = {}
    try:
        with open(path, encoding="latin-1", newline="") as f:
            rows = list(csv.reader(f, delimiter=";"))
    except Exception as e:
        print(f"AVERTISSEMENT: fichier import M-1 illisible ({e}) -- repli Zone (mois precedent) desactive.")
        return {}
    if not rows:
        return {}
    header = [normalize_header(h) for h in rows[0]]
    try:
        i_tracking = header.index("N° Tracking")
        i_zone = header.index("Zone")
    except ValueError:
        print("AVERTISSEMENT: en-tetes 'N° Tracking'/'Zone' introuvables dans le fichier import M-1 -- repli desactive.")
        return {}
    for r in rows[1:]:
        if len(r) <= max(i_tracking, i_zone):
            continue
        t = str(r[i_tracking] or "").strip()
        zone = str(r[i_zone] or "").strip()
        if t and zone and t not in m:
            m[t] = zone
    return m


def poids_audite_de_ligne(r):
    """Extrait le poids AUDITE UPS en texte libre ("AUDITED WEIGHT: X KGS") d'une ligne CSV
    brute, ou None si absent -- meme regex que carriers/ups/index.js (COL.auditedWeight)."""
    import re
    raw = str(r[COL["audited_weight"]] if len(r) > COL["audited_weight"] else "") or ""
    m = re.search(r"AUDITED WEIGHT:\s*([\d.,]+)\s*KGS", raw, re.IGNORECASE)
    return num(m.group(1)) if m else None


def poids_api_ups(tracking, cache):
    """Poids (KGS) via l'API UPS Tracking -- garde de compatibilite, delegue a
    donnees_api_ups() (cf. ci-dessous) qui recupere aussi le nombre de colis."""
    return donnees_api_ups(tracking, cache)[0]


def donnees_api_ups(tracking, cache):
    """(poids KGS, nombre de colis) via l'API UPS Tracking (repli DERNIER RECOURS), avec cache
    memoire pour eviter de rappeler l'API 2x pour le meme tracking. Chaque valeur est None si
    absente/non trouvee. Ne leve jamais pour un cas normal (tracking introuvable/pas de donnee)
    -- seules les erreurs reseau/auth sont juste averties en console (meme principe que
    core/upsApi.js cote Node). AJOUT 2026-08-31 (packageCount) : confirme sur reponse API
    reelle (tracking 1ZA1912WDK91200736, expedition '1 of 3 Piece Shipment' sur le site UPS,
    packageCount=3 dans la reponse JSON) -- champ trackResponse.shipment[].package[].
    packageCount, au niveau du PACKAGE (pas du shipment), ancien commentaire ('pas d'API
    equivalente pour le nombre de colis') etait errone, jamais verifie empiriquement avant.

    ATTENTION POIDS (verifie 2026-08-31, echantillon de 8 trackings multi-colis reels,
    compare export brut WMS vs API) : sur un envoi groupe (plusieurs colis sous le meme
    tracking), la reponse API ne contient QU'UN SEUL package (pas un par colis malgre
    packageCount>1) et son "weight" est le poids d'UN SEUL colis du groupage, PAS le poids
    total de l'expedition -- ex. tracking a 6 colis/105kg reel (WMS) -> API renvoie
    packageCount=6 (correct) mais weight=19.00 (poids d'1 seul colis, PAS 105). Sur 8
    trackings testes, seuls 2/8 (envois a 1 seul colis) avaient un poids API concordant
    avec le WMS ; les 6 envois multi-colis divergeaient tous, l'API sous-evaluant
    systematiquement. NE JAMAIS mettre l'API en PRIORITE sur le Poids (repli export brut en
    premier reste correct, cf. cascade dans main()) -- packageCount (colis) est en revanche
    fiable (8/8 concordant avec le WMS sur le meme echantillon), sans reserve connue a ce
    jour."""
    if tracking in cache:
        return cache[tracking]
    client_id = os.environ.get("UPS_CLIENT_ID")
    client_secret = os.environ.get("UPS_CLIENT_SECRET")
    if not client_id or not client_secret:
        cache[tracking] = (None, None)
        return (None, None)
    import base64
    import json as _json
    import urllib.request
    import urllib.parse
    import urllib.error
    try:
        token = donnees_api_ups._token
    except AttributeError:
        token = None
    if not token:
        try:
            basic = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
            req = urllib.request.Request(
                "https://onlinetools.ups.com/security/v1/oauth/token",
                data=b"grant_type=client_credentials",
                headers={"Authorization": f"Basic {basic}", "Content-Type": "application/x-www-form-urlencoded"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=8) as res:
                token = _json.loads(res.read())["access_token"]
            donnees_api_ups._token = token
        except Exception as e:
            print(f"AVERTISSEMENT: authentification API UPS echouee ({e}) -- repli API desactive pour ce lot.")
            donnees_api_ups._token = ""
            cache[tracking] = (None, None)
            return (None, None)
    try:
        url = f"https://onlinetools.ups.com/api/track/v1/details/{urllib.parse.quote(tracking)}"
        req = urllib.request.Request(url, headers={
            "Authorization": f"Bearer {token}",
            "transId": str(int(time.time() * 1000)),
            "transactionSrc": "facturation-transporteurs",
        })
        with urllib.request.urlopen(req, timeout=8) as res:
            data = _json.loads(res.read())
        for shipment in (data.get("trackResponse", {}) or {}).get("shipment", []):
            for pkg in shipment.get("package", []):
                w = pkg.get("weight") or {}
                poids = num(w.get("weight"))
                if poids and poids > 0:
                    # unitOfMeasurement observe en pratique comme string simple ("KGS"), pas
                    # dict {code: "KGS"} -- BUG TROUVE 2026-08-27 (reponse API reelle testee),
                    # tolerant aux deux formes par securite.
                    raw_unit = w.get("unitOfMeasurement")
                    unit = str((raw_unit.get("code") if isinstance(raw_unit, dict) else raw_unit) or "KGS").upper()
                    poids = round(poids * 0.45359237, 2) if unit == "LBS" else poids
                else:
                    poids = None
                colis = pkg.get("packageCount")
                colis = int(colis) if isinstance(colis, (int, float, str)) and str(colis).strip().isdigit() else None
                if poids or colis:
                    cache[tracking] = (poids, colis)
                    return (poids, colis)
        cache[tracking] = (None, None)
        return (None, None)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            cache[tracking] = (None, None)
            return (None, None)
        print(f"AVERTISSEMENT: API UPS Tracking erreur {e.code} pour {tracking} -- POIDS/COLIS restent a 0.")
        cache[tracking] = (None, None)
        return (None, None)
    except Exception as e:
        print(f"AVERTISSEMENT: appel API UPS echoue pour {tracking} ({e}) -- POIDS/COLIS restent a 0.")
        cache[tracking] = (None, None)
        return (None, None)


def derniere_ligne_reelle(ws, last_col_check, last_row_end_xlup):
    """Rogne 'last_row_end_xlup' (resultat de Cells(Rows.Count,c).End(xlUp).Row) en remontant
    tant que la cellule de la colonne 'last_col_check' est vide/valeur d'erreur (#N/A, #REF!...).
    BUG TROUVE 2026-08-31 (onglet TCD, tracking A1912WTZX8M) : End(xlUp) peut s'arreter a tort
    sur une ligne RESIDUELLE du modele/PivotTable natif (colonne cible vide mais Excel la juge
    "non-vide" pour une raison de mise en forme/etat transitoire) -- ces lignes fantomes
    etirent ensuite les formules bien au-dela des vraies donnees, et se propagent en cascade
    (ex. 'Fichier import', qui s'etend sur la taille du TCD). Lecture en BLOC (1 aller-retour
    COM), pas cellule par cellule."""
    if last_row_end_xlup < 2:
        return last_row_end_xlup
    values = ws.Range(ws.Cells(2, last_col_check), ws.Cells(last_row_end_xlup, last_col_check)).Value
    if not isinstance(values, tuple):
        values = ((values,),)
    row = last_row_end_xlup
    while row >= 2:
        v = values[row - 2][0]
        if v is not None and str(v).strip() and not str(v).strip().startswith("#"):
            break
        row -= 1
    return row


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def load_dotenv_ups():
    """Charge facturation-app/.env dans os.environ (UPS_CLIENT_ID/SECRET) -- parseur minimal
    plutot qu'une dependance python-dotenv non garantie sur le poste d'execution. No-op
    silencieux si le fichier est absent (repli API simplement desactive, cf. poids_api_ups)."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "facturation-app", ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k, v = k.strip(), v.strip()
            if k and k not in os.environ:
                os.environ[k] = v


def main():
    load_dotenv_ups()
    modele, sortie, csv_paths, brut_paths, period, import_m1_path = parse_args(sys.argv)
    if not csv_paths:
        raise RuntimeError("Aucun CSV fourni (--csv <facture1.csv> [...]).")
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    carrier_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "facturation-app", "src", "carriers", "ups")
    cfg = load_config(carrier_dir)

    all_rows = []
    for p in csv_paths:
        all_rows.extend(read_ups_csv(p))
    if not all_rows:
        raise RuntimeError("Fichier(s) UPS vide(s) ou illisible(s).")
    print(f"Entrée : {len(all_rows)} ligne(s) brute(s), {len(csv_paths)} fichier(s).")

    # Repli Zone (regle 3, CDC 2026-08-27) : zone=0 SANS fret -> chercher le tracking dans le
    # fichier import CSV du mois precedent (deja livre a l'ERP). Charge tot avec les autres
    # sources externes ; consomme plus loin, apres le calcul Excel de "Fichier import".
    zone_m1_map = load_import_m1(import_m1_path)
    if zone_m1_map:
        print(f"Fichier import du mois précédent : {len(zone_m1_map)} tracking(s)/zone(s) chargé(s) (repli Zone=0 sans frêt).")
    elif import_m1_path:
        print("AVERTISSEMENT: fichier import M-1 fourni mais aucune zone exploitable n'en a été extraite.")

    # Mois cible = choix utilisateur (--period, source de verite, decision 2026-08-20 --
    # BUG TROUVE 2026-08-26 : --period n'etait jamais transmis au finaliseur, "Date validite
    # tarif" restait calculee par majorite sur "Date de la facture" du CSV brut, peu fiable,
    # cf. carrier Node index.js meme fix) -- repli sur le mois majoritaire si absent.
    from collections import Counter
    import datetime as _dt
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    date_validite_serial = None
    mois_cible = None
    if period:
        m = re.fullmatch(r"(\d{4})_(\d{2})", period)
        if m:
            mois_cible = f"{m.group(1)}{m.group(2)}"
            date_validite_serial = (_dt.datetime(int(m.group(1)), int(m.group(2)), 1) - EXCEL_EPOCH).days
    if mois_cible is None:
        compte_mois = Counter()
        for r in all_rows:
            d = str(r[COL["date_facture"]] if len(r) > COL["date_facture"] else "").strip()
            m = re.match(r"^(\d{4})-(\d{2})-\d{2}$", d)
            if m:
                compte_mois[f"{m.group(1)}{m.group(2)}"] += 1
        if compte_mois:
            mois_cible = compte_mois.most_common(1)[0][0]
            annee, mois = int(mois_cible[:4]), int(mois_cible[4:6])
            date_validite_serial = (_dt.datetime(annee, mois, 1) - EXCEL_EPOCH).days

    # Colis 1Z79 (regle FACTURATION EXCEL.docx) : colis viticulteur retourne chez La Ruche --
    # reste dans "Facture UPS" (donnees brutes, decision utilisateur 2026-08-25 -- confirme
    # contre le fichier reel fait-main de juillet 2026), EXCLU uniquement du fichier import,
    # et agrege dans l'onglet "Demande avoir" (Tracking/Nb colis/Montant/Cause). BUG TROUVE
    # 2026-08-25 (test sur 14 trackings reels de juillet 2026, comparaison au fichier
    # fait-main) : "code_classe=FRT strict" etait FAUX -- Montant = somme de TOUTES les
    # lignes SAUF TVA (code_classe=TAX) et Taxe gazole (code_classe=FSC), confirme exact sur
    # 2/3 cas d'ecart reexamines (le 3e reste un ecart residuel inexplique, probable
    # ajustement manuel isole). Nb colis = FIGE A 1 par tracking (valeur du fait-main sur
    # 13/14 trackings testes, aucune formule de somme sur les colonnes brutes ne colle).
    # Lignes SANS Numero de suivi NI Numero de reference 1 (Ref1, colonne T) : demande
    # utilisateur 2026-08-25 -- supprimees de "Facture UPS" (jamais collees, meme regle que
    # le carrier Node index.js), aucune cle d'identification exploitable. Ref1 vide, "." OU
    # "null" en texte (valeurs non-renseignees frequentes cote UPS) compte comme "pas de reference".
    def ref_vide(v):
        v = (v or "").strip()
        return (not v) or set(v) == {"."} or v.lower() == "null"

    CAUSE_1Z79 = "Nouveau compte pour le renvoi des colis depuis LR  l'avoir est censé être remis automatiquement sans avoir besoin de le demander"
    demandes_avoir_1z79 = {}  # tracking -> montant
    n_sans_identification = 0
    n_montant_net_zero = 0
    lignes_retenues = []
    for r in all_rows:
        t = str(r[COL["numero_suivi"]] if len(r) > COL["numero_suivi"] else "").strip()
        montant_net = coerce(r[COL["montant_net"]] if len(r) > COL["montant_net"] else "")
        montant_net = montant_net if isinstance(montant_net, (int, float)) else 0
        if t.upper().startswith("1Z79"):
            code_classe = str(r[COL["code_classe"]] if len(r) > COL["code_classe"] else "").strip().upper()
            if code_classe not in ("TAX", "FSC"):
                demandes_avoir_1z79[t] = round(demandes_avoir_1z79.get(t, 0.0) + montant_net, 2)
            else:
                demandes_avoir_1z79.setdefault(t, 0.0)
            # Lignes a Montant net = 0 : demande utilisateur 2026-08-25 -- supprimees de
            # "Facture UPS" MEME pour les 1Z79 (BUG TROUVE 2026-08-25, cf. carrier Node
            # index.js pour le detail complet : le fait-main a 0/49 lignes 1Z79 a montant=0,
            # alors que le CSV brut en a 9/58 -- toutes supprimees). L'agregation "Demande
            # avoir" ci-dessus reste faite AVANT ce filtre.
            if montant_net == 0:
                n_montant_net_zero += 1
                continue
            lignes_retenues.append(r)
            continue
        ref1 = str(r[COL["ref1"]] if len(r) > COL["ref1"] else "").strip()
        if not t and ref_vide(ref1):
            n_sans_identification += 1
            continue
        # Lignes a Montant net = 0 (hors 1Z79, deja traite ci-dessus) : demande utilisateur
        # 2026-08-25 -- supprimees de "Facture UPS" (pas seulement du fichier import),
        # confirme sur reel (facture 202600782885 : 13513 lignes a montant=0 AVEC tracking
        # rempli, distinct du filtre "sans identification" ci-dessus).
        if montant_net == 0:
            n_montant_net_zero += 1
            continue
        lignes_retenues.append(r)
    if n_montant_net_zero:
        print(f"{n_montant_net_zero} ligne(s) à Montant net = 0 supprimée(s) de 'Facture UPS'.")
    if demandes_avoir_1z79:
        print(f"{len(demandes_avoir_1z79)} colis en 1Z79 (retour viticulteur chez La Ruche) exclus de l'import — reportés dans l'onglet 'Demande avoir' : {sorted(demandes_avoir_1z79.keys())}.")
    if n_sans_identification:
        print(f"{n_sans_identification} ligne(s) sans Numéro de suivi ni Numéro de référence 1 supprimée(s) (aucune clé d'identification exploitable).")

    # Repli POIDS en cascade (priorite 1: poids audite UPS en texte -- priorite 2: export WMS
    # brut m/m-1/m-2 -- priorite 3: API UPS Tracking, dernier recours) -- BUG TROUVE 2026-08-27 :
    # ce repli existait deja cote carrier Node (carriers/ups/index.js) mais n'avait AUCUN effet
    # sur le fichier reellement livre, car UPS utilise importFromWorkbook=true (le CSV final vient
    # de CE classeur/finaliseur, pas du calcul JS). Porte ici, en amont de data_brut, pour que le
    # poids injecte soit repris par TOUTES les formules Excel en aval (zone colis poids assurance,
    # TCD, Fichier import) exactement comme un vrai poids UPS.
    # BUG TROUVE 2026-08-31 (test reel, tracking A1912WTZX8M) : ces maps etaient calculees sur
    # lignes_retenues (DEJA filtree "Montant net = 0" quelques lignes plus haut) -- or le VRAI
    # poids/colis d'un tracking en lignes multiples (FRT + BRK/GOV/EXM/TAX) est souvent porte par
    # la ligne FRT, qui peut avoir Montant net=0.00 (le vrai cout est sur les lignes BRK/GOV
    # associees) et se faisait donc supprimer AVANT que ce calcul ne la voie -- le tracking
    # n'avait alors plus aucune ligne avec un vrai poids/colis, repli impossible (aucun export
    # WMS/API ne peut "trouver" une donnee qui existait deja dans le CSV brut). Fix : calculer
    # ces maps sur all_rows (TOUTES les lignes brutes, avant filtrage montant=0), meme source que
    # l'agregation 1Z79 plus haut -- l'injection finale reste sur lignes_retenues (seules lignes
    # encore presentes en sortie).
    max_poids_par_tracking = {}
    poids_audite_par_tracking = {}
    for r in all_rows:
        t = str(r[COL["numero_suivi"]] if len(r) > COL["numero_suivi"] else "").strip()
        if not t:
            continue
        # BUG TROUVE 2026-08-27 : "if p > dict.get(t, 0)" ne cree JAMAIS l'entree pour un
        # tracking dont TOUTES les lignes ont poids=0 (0 > 0 est faux) -- le tracking restait
        # absent de max_poids_par_tracking, donc invisible pour trackings_a_zero ci-dessous.
        # setdefault garantit une entree meme a 0.
        max_poids_par_tracking.setdefault(t, 0.0)
        p = num(r[COL["poids_facture"]] if len(r) > COL["poids_facture"] else "")
        if p > max_poids_par_tracking[t]:
            max_poids_par_tracking[t] = p
        pa = poids_audite_de_ligne(r)
        if pa and pa > poids_audite_par_tracking.get(t, 0):
            poids_audite_par_tracking[t] = pa

    # Repli NBRE COLIS -- meme famille de probleme que le poids (lignes d'AJUSTEMENT/correction
    # UPS, ex. codeClasse=ACC "Frais de correction d'expedition", ou "Nombre de colis"=0 dans le
    # CSV brut). CONFIRME 2026-08-27 (tracking 1ZA1912WD990370248) : aucun texte cache
    # equivalent a "AUDITED WEIGHT" pour le nombre de colis -- UPS ne le fournit pas du tout sur
    # ces lignes. Seul repli disponible : export WMS brut (colonne INFO_NBCOLIS), pas d'API
    # UPS equivalente pour le nombre de colis -- contrairement au poids, pas de priorite 1/3.
    # Calcule sur all_rows -- meme raison que le poids ci-dessus (BUG TROUVE 2026-08-31).
    max_colis_par_tracking = {}
    for r in all_rows:
        t = str(r[COL["numero_suivi"]] if len(r) > COL["numero_suivi"] else "").strip()
        if not t:
            continue
        max_colis_par_tracking.setdefault(t, 0.0)
        c = num(r[COL["nombre_colis"]] if len(r) > COL["nombre_colis"] else "")
        if c > max_colis_par_tracking[t]:
            max_colis_par_tracking[t] = c

    # Repli PAYS -- meme famille de probleme que Poids/Colis (lignes d'AJUSTEMENT/correction
    # UPS, colonne Pays vide dans le CSV brut). CONFIRME 2026-08-27 (tracking
    # 1ZA1912WD990370248) : pays destinataire trouve dans l'export WMS brut (colonne
    # DES_PAYS), meme mecanisme que Colis -- pas d'audite texte ni d'API pour le pays.
    # Calcule sur all_rows -- meme raison que le poids/colis ci-dessus (BUG TROUVE 2026-08-31).
    max_pays_par_tracking = {}
    for r in all_rows:
        t = str(r[COL["numero_suivi"]] if len(r) > COL["numero_suivi"] else "").strip()
        if not t:
            continue
        pays_ligne = str(r[COL["pays"]] if len(r) > COL["pays"] else "").strip()
        if pays_ligne and t not in max_pays_par_tracking:
            max_pays_par_tracking[t] = pays_ligne

    trackings_poids_a_zero = [t for t, p in max_poids_par_tracking.items() if not p]
    trackings_colis_a_zero = [t for t, c in max_colis_par_tracking.items() if not c]
    trackings_pays_vide = [t for t in max_poids_par_tracking if not max_pays_par_tracking.get(t)]
    besoin_export = bool(brut_paths and (trackings_poids_a_zero or trackings_colis_a_zero or trackings_pays_vide))
    poids_export_map, colis_export_map, pays_export_map = load_brut_poids_colis_pays(brut_paths) if besoin_export else ({}, {}, {})

    # Cache PARTAGE entre Poids et Colis (meme appel API renvoie les deux, cf.
    # donnees_api_ups) -- evite un 2e appel reseau pour un tracking a la fois Poids=0 ET
    # Colis=0 (deja interroge lors du repli Poids, reponse encore en cache pour Colis).
    # ORDRE VOLONTAIRE (ne pas inverser) : export brut D'ABORD, API en DERNIER recours --
    # verifie empiriquement 2026-08-31 que l'API sous-evalue le Poids sur les envois
    # multi-colis (ne renvoie qu'1 seul package, pas le total du groupage), cf. avertissement
    # complet dans donnees_api_ups() ci-dessus. Le Colis (packageCount) est fiable cote API,
    # mais l'ordre reste le meme pour tout le monde par coherence et simplicite du cache
    # partage.
    api_cache = {}
    n_repli_audite = n_repli_export = n_repli_api = 0
    poids_repli_par_tracking = {}
    for t in trackings_poids_a_zero:
        if poids_audite_par_tracking.get(t):
            poids_repli_par_tracking[t] = poids_audite_par_tracking[t]
            n_repli_audite += 1
        elif poids_export_map.get(t):
            poids_repli_par_tracking[t] = poids_export_map[t]
            n_repli_export += 1
        else:
            p_api, _c_api = donnees_api_ups(t, api_cache)
            if p_api:
                poids_repli_par_tracking[t] = p_api
                n_repli_api += 1

    # AJOUT 2026-08-31 : repli Colis via API UPS Tracking (packageCount), dernier recours --
    # confirme disponible dans la reponse API (cf. donnees_api_ups), contrairement a ce qui
    # etait suppose avant (aucune verification empirique n'avait ete faite). Meme cascade que
    # Poids : export brut d'abord (deja majoritaire), API en dernier recours seulement.
    n_repli_colis_export = n_repli_colis_api = 0
    colis_repli_par_tracking = {}
    for t in trackings_colis_a_zero:
        if colis_export_map.get(t):
            colis_repli_par_tracking[t] = colis_export_map[t]
            n_repli_colis_export += 1
        else:
            _p_api, c_api = donnees_api_ups(t, api_cache)
            if c_api:
                colis_repli_par_tracking[t] = c_api
                n_repli_colis_api += 1

    n_repli_pays_export = 0
    pays_repli_par_tracking = {}
    for t in trackings_pays_vide:
        if pays_export_map.get(t):
            pays_repli_par_tracking[t] = pays_export_map[t]
            n_repli_pays_export += 1

    # BUG TROUVE 2026-08-31 (tracking A1912WTZX8M) : max_poids_par_tracking/max_colis_par_tracking
    # sont calcules sur all_rows (toutes les lignes brutes) mais l'injection ci-dessous ne
    # portait QUE sur poids_repli_par_tracking/colis_repli_par_tracking (repli audite/export/
    # API, cas ou le max sur all_rows est LUI-MEME 0). Si le vrai poids/colis existe deja dans
    # all_rows (ex. ligne FRT a Montant net=0, supprimee de lignes_retenues AVANT ce calcul,
    # mais le tracking garde d'autres lignes a montant non-nul type BRK/GOV/EXM) le tracking
    # ne rentre plus dans trackings_poids_a_zero/trackings_colis_a_zero (max non-nul) -- mais
    # RIEN n'ecrivait alors ce max (deja connu) sur les lignes restantes, qui gardaient leur
    # propre poids/colis individuel a 0. Fix : injecter aussi le max connu (all_rows) sur toute
    # ligne dont le poids/colis INDIVIDUEL est a 0, independamment du mecanisme de repli
    # audite/export/API (qui ne concerne que les trackings dont le max GLOBAL est encore 0).
    if poids_repli_par_tracking or colis_repli_par_tracking or pays_repli_par_tracking or max_poids_par_tracking or max_colis_par_tracking:
        for r in lignes_retenues:
            t = str(r[COL["numero_suivi"]] if len(r) > COL["numero_suivi"] else "").strip()
            if len(r) > COL["poids_facture"]:
                poids_repli = poids_repli_par_tracking.get(t) or max_poids_par_tracking.get(t)
                if poids_repli and num(r[COL["poids_facture"]]) == 0:
                    r[COL["poids_facture"]] = poids_repli
            if len(r) > COL["nombre_colis"]:
                colis_repli = colis_repli_par_tracking.get(t) or max_colis_par_tracking.get(t)
                if colis_repli and num(r[COL["nombre_colis"]]) == 0:
                    r[COL["nombre_colis"]] = colis_repli
            pays_repli = pays_repli_par_tracking.get(t)
            if pays_repli and len(r) > COL["pays"]:
                r[COL["pays"]] = pays_repli
    if n_repli_audite:
        print(f"{n_repli_audite} tracking(s) à Poids = 0 complété(s) avec le poids audité UPS (AUDITED WEIGHT en texte dans la facture).")
    if n_repli_export:
        print(f"{n_repli_export} tracking(s) à Poids = 0 complété(s) via l'export WMS 'expéditions_brut'.")
    if n_repli_api:
        print(f"{n_repli_api} tracking(s) à Poids = 0 complété(s) via l'API UPS Tracking (dernier recours).")
    if n_repli_colis_export:
        print(f"{n_repli_colis_export} tracking(s) à Nombre de colis = 0 complété(s) via l'export WMS 'expéditions_brut'.")
    if n_repli_colis_api:
        print(f"{n_repli_colis_api} tracking(s) à Nombre de colis = 0 complété(s) via l'API UPS Tracking (dernier recours).")
    if n_repli_pays_export:
        print(f"{n_repli_pays_export} tracking(s) à Pays manquant complété(s) via l'export WMS 'expéditions_brut'.")

    ncol = max((len(r) for r in lignes_retenues), default=0)
    data_brut = [[coerce(v) for v in (r + [None] * ncol)[:ncol]] for r in lignes_retenues]

    ep_map = load_brut_ep(brut_paths) if brut_paths else {}
    if not brut_paths:
        print("AVERTISSEMENT: aucun export 'expéditions_brut' fourni -> E/P classé 'P' par défaut sauf plus-value BtoC détectée.")

    import win32com.client as win32
    xlUp = -4162
    xlDatabase = 1
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) "Facture UPS" : purge + collage donnees brutes a partir de la colonne E (A-D =
        #    calculees : Clients/Montant assurance/Mode envoi/Categorie).
        ws = wb.Sheets("Facture UPS")
        if ws.AutoFilterMode:
            ws.AutoFilterMode = False
        FIRST_RAW_COL = 5  # colonne E
        LAST_RAW_COL = FIRST_RAW_COL - 1 + ncol
        LAST_CALC_COL = 4  # A->D
        oldLast = ws.Cells(ws.Rows.Count, FIRST_RAW_COL).End(xlUp).Row
        n = len(data_brut)
        newLast = 1 + n
        maxLast = max(oldLast, newLast, 2)
        purgeUntil = maxLast + 200
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(purgeUntil, max(LAST_RAW_COL, LAST_CALC_COL))).ClearContents())

        # "Numero de suivi" est TOUJOURS numerique/alphanumerique sans zero de tete dans le
        # brut reel UPS -- pas de force NumberFormat texte necessaire (contrairement a TNT).
        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, LAST_RAW_COL)).__setattr__("Value", data_brut))
        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, max(LAST_RAW_COL, LAST_CALC_COL))).ClearContents())

        # Colonnes A-D : formules du modele, resolues par POSITION FIXE (decalage +4 confirme).
        def col_letter(idx0):
            import openpyxl
            return openpyxl.utils.get_column_letter(idx0 + 1)

        col_code_classe = col_letter(FIRST_RAW_COL - 1 + COL["code_classe"])
        col_code_description = col_letter(FIRST_RAW_COL - 1 + COL["code_description"])
        col_description = col_letter(FIRST_RAW_COL - 1 + COL["description"])
        col_valeur_base = col_letter(FIRST_RAW_COL - 1 + COL["valeur_base"])

        # B (Montant assurance) : =IF(codeDescription="EVS",valeurBase,0)
        formula_b = f'=IF({col_code_description}{{row}}="EVS",{col_valeur_base}{{row}},0)'
        retry(lambda t=formula_b: ws.Range(ws.Cells(2, 2), ws.Cells(newLast, 2)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))

        # C (Mode envoi) : =IF(codeClasse="FRT",XLOOKUP(description,'ST SV'!A:A,B:B),0)
        formula_c = f'=IF({col_code_classe}{{row}}="FRT",_xlfn.XLOOKUP({col_description}{{row}},\'ST SV\'!A:A,\'ST SV\'!B:B),0)'
        retry(lambda t=formula_c: ws.Range(ws.Cells(2, 3), ws.Cells(newLast, 3)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))

        # D (Categorie) : cascade exacte du modele.
        formula_d = (
            f'=IF(COUNTIF(\'ST SV\'!Q:Q,{col_description}{{row}})<>0,"Adresse",'
            f'IF(COUNTIF(\'ST SV\'!D:D,{col_description}{{row}})<>0,"plus-value BtoC",'
            f'IF({col_code_classe}{{row}}="FRT","Frêt",'
            f'IF({col_code_classe}{{row}}="TAX","TVA",'
            f'IF(COUNTIF(\'Charge.CHG_CODE\'!A:A,{col_code_description}{{row}})=0,"code inconnu",'
            f'_xlfn.XLOOKUP({col_code_description}{{row}},\'Charge.CHG_CODE\'!A:A,\'Charge.CHG_CODE\'!C:C))))))'
        )
        retry(lambda t=formula_d: ws.Range(ws.Cells(2, 4), ws.Cells(newLast, 4)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))

        ws.Range(ws.Cells(1, 1), ws.Cells(newLast, max(LAST_RAW_COL, LAST_CALC_COL))).AutoFilter()

        # 2) 5 TCD : PivotCache redirige vers la vraie plage large de 'Facture UPS' (cache du
        #    modele fige sur une plage etroite -- meme piege que DPD/Geodis/Mondial Relay/
        #    Chronopost/TNT/FedEx). "Bilan clients" pointe sur 'TCD' (pas 'Facture UPS'),
        #    redirige separement APRES redimensionnement du TCD "TCD" (etape 3).
        newRangeFacture = ws.Range(ws.Cells(1, 1), ws.Cells(newLast, max(LAST_RAW_COL, LAST_CALC_COL)))
        for sheet_name in ("Bilan factures", "zone colis poids assurance", "ST SV", "TCD"):
            wsPivot = wb.Sheets(sheet_name)
            for i in range(1, wsPivot.PivotTables().Count + 1):
                pt = wsPivot.PivotTables(i)
                src = str(pt.PivotCache().SourceData)
                if "Facture UPS" not in src:
                    print(f"TCD '{pt.Name}' ({sheet_name}) : source '{src}' n'est pas 'Facture UPS' -> config préservée.")
                    continue
                newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=newRangeFacture)
                pt.ChangePivotCache(newCache)
                print(f"TCD '{pt.Name}' ({sheet_name}) : PivotCache redirigé vers {newRangeFacture.Address}.")
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        xl.Calculate()

        # 3) "zone colis poids assurance" colonnes manuelles B/C/I/J (etendues jusqu'a la
        #    vraie derniere ligne native, colonne D = Numero de suivi).
        wsZcp = wb.Sheets("zone colis poids assurance")
        lastZcp = wsZcp.Cells(wsZcp.Rows.Count, 4).End(xlUp).Row
        if lastZcp >= 2:
            # CDC pole transport 2026-08-27 ("poids jamais a 0, arrondi sup 0 decimale, seules
            # decimales autorisees 0,1 et 0,5") : colonne I couvre deja "arrondi sup 0 decimale"
            # pour tous les comptes non-COD (ROUNDUP(H,0) = toujours un entier). Colonne J
            # (UPS_COD, <=3 colis et poids<10kg) arrondit techniquement a 1 decimale quelconque
            # (x.1 a x.9), mais aucune occurrence hors {0,1 ; 0,5} n'a ete observee sur les
            # fichiers reels livres -- NON MODIFIEE : la demande utilisateur est traitee comme
            # portant sur l'interdiction de poids=0 (deja couverte par la cascade de repli
            # Python audite/export WMS/API, lignes ~500-611, appliquee AVANT ces formules). Si
            # un poids reel x.3/x.7 apparait un jour, revoir avec MROUND(H,0.5) (+ ROUNDUP
            # minimum pour eviter un retour a 0).
            formulas_zcp = {
                2: "=COUNTIF('Clients log'!A:A,A{row})",          # B Logistique
                3: "=IF(E{row}=0,1,E{row})",                        # C Colis
                9: "=ROUNDUP(H{row},0)",                            # I Poids UPS
                10: '=IF(F{row}>3,ROUNDUP(H{row},0),IF(H{row}<10,ROUNDUP(H{row},1),ROUNDUP(H{row},0)))',  # J Poids UPS_COD
            }
            for col_idx, tmpl in formulas_zcp.items():
                retry(lambda c=col_idx, t=tmpl: wsZcp.Range(wsZcp.Cells(2, c), wsZcp.Cells(lastZcp, c))
                      .__setattr__("Formula", [[t.format(row=r)] for r in range(2, lastZcp + 1)]))
            print(f"'zone colis poids assurance' : colonnes B/C/I/J étirées jusqu'à la ligne {lastZcp}.")

        # 4) "ST SV" colonne manuelle N (etendue jusqu'a la derniere ligne native, colonne H).
        wsStSv = wb.Sheets("ST SV")
        lastStSv = wsStSv.Cells(wsStSv.Rows.Count, 8).End(xlUp).Row
        if lastStSv >= 3:  # entete sur 2 lignes (constate sur le modele)
            retry(lambda: wsStSv.Range(wsStSv.Cells(3, 14), wsStSv.Cells(lastStSv, 14))
                  .__setattr__("Formula", [[f'=IF(K{r}<>"","SV","ST")'] for r in range(3, lastStSv + 1)]))
            print(f"'ST SV' : colonne N étirée jusqu'à la ligne {lastStSv}.")

        # 5) "TCD" colonne manuelle B/D (etendues jusqu'a la vraie derniere ligne native,
        #    colonne E = Numero de suivi). Colonnes A (Logistique)/C (Client) NON reconstruites
        #    (manuelles -- Client = regle transversale ID client, saisie humaine).
        wsTcd = wb.Sheets("TCD")
        # BUG TROUVE 2026-08-31 : End(xlUp) seul peut s'arreter sur une ligne residuelle du
        # modele (colonne E vide, cf. derniere_ligne_reelle) -- rogne AVANT d'etirer B/D pour
        # eviter que ces formules elles-memes ne creent les residus vus plus loin (etape 2,
        # masquage 1Z79) sur des lignes qui n'auraient jamais du exister.
        lastTcd = derniere_ligne_reelle(wsTcd, 5, wsTcd.Cells(wsTcd.Rows.Count, 5).End(xlUp).Row)
        if lastTcd >= 3:  # entete sur 2 lignes
            formulas_tcd = {
                2: "=SUM(F{row}:K{row})+O{row}+M{row}",  # B Cout (controle)
                4: "=_xlfn.XLOOKUP(E{row},'zone colis poids assurance'!D:D,'zone colis poids assurance'!I:I)",  # D Poids
            }
            for col_idx, tmpl in formulas_tcd.items():
                retry(lambda c=col_idx, t=tmpl: wsTcd.Range(wsTcd.Cells(3, c), wsTcd.Cells(lastTcd, c))
                      .__setattr__("Formula", [[t.format(row=r)] for r in range(3, lastTcd + 1)]))
            print(f"'TCD' : colonnes B/D étirées jusqu'à la ligne {lastTcd}.")

        # 1Z79 EXCLUS du TCD "TCD" (donc de "Fichier import", qui s'etend sur lastTcd) --
        # decision utilisateur 2026-08-25 : reste dans "Facture UPS" (donnees brutes) mais pas
        # dans le fichier import. Le TCD etant un vrai PivotTable (source='Facture UPS'
        # redirigee ci-dessus), les 1Z79 y remonteraient automatiquement -- masques ici via
        # PivotItems du champ "Numéro de suivi" (RowField), pas via suppression de lignes
        # (casserait la structure du pivot). Volume negligeable (15 trackings distincts/mois
        # observe sur juin 2026), boucle simple sans impact perf.
        try:
            tcd_pt = wsTcd.PivotTables(1)
            for pf in tcd_pt.PivotFields():
                if pf.Orientation == 1 and "suivi" in str(pf.Name).lower():  # xlRowField
                    n_masques = 0
                    for pi in pf.PivotItems():
                        if str(pi.Name).upper().startswith("1Z79"):
                            try:
                                pi.Visible = False
                                n_masques += 1
                            except Exception as e:
                                print(f"TCD : impossible de masquer le tracking 1Z79 {pi.Name!r} :", e)
                    if n_masques:
                        print(f"'TCD' : {n_masques} tracking(s) 1Z79 masqué(s) (exclus de Fichier import).")
                    break
            wb.RefreshAll()
            xl.Calculate()
            # BUG TROUVE 2026-08-31 (tracking A1912WTZX8M, ligne 6912 "DATE manquante" + lignes
            # fantomes juste apres, ex. "AB") : End(xlUp) peut s'arreter a tort sur des lignes
            # RESIDUELLES du PivotTable natif (cf. derniere_ligne_reelle) -- consequence en
            # cascade : "Fichier import" (qui s'etend sur lastTcd) herite de ces lignes fantomes,
            # et la formule Date validite en chaine (=E{prev}) CASSE en #REF! des qu'une ligne
            # intermediaire est supprimee plus loin (filtre "lignes sans charge"), affectant
            # meme la DERNIERE ligne REELLE juste avant les fantomes.
            lastTcdAvant = wsTcd.Cells(wsTcd.Rows.Count, 5).End(xlUp).Row
            lastTcd = derniere_ligne_reelle(wsTcd, 5, lastTcdAvant)
            if lastTcd < lastTcdAvant:
                print(f"'TCD' : {lastTcdAvant - lastTcd} ligne(s) résiduelle(s) (tracking vide/erreur, artefact PivotTable) exclue(s) en fin de tableau.")
        except Exception as e:
            print("Avertissement : masquage 1Z79 sur le TCD a échoué :", e)

        # "Bilan clients" (source='TCD', PAS 'Facture UPS') : redirige APRES redimensionnement
        # de "TCD" ci-dessus, sur toute la hauteur POSSIBLE (pas juste celle deja peuplee).
        for i in range(1, wb.Sheets("Bilan clients").PivotTables().Count + 1):
            pt = wb.Sheets("Bilan clients").PivotTables(i)
            src = str(pt.PivotCache().SourceData)
            if "TCD" not in src:
                continue
            newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=wsTcd.Range(wsTcd.Cells(1, 1), wsTcd.Cells(max(lastTcd, 2), 16)))
            pt.ChangePivotCache(newCache)
        wb.RefreshAll()
        xl.Calculate()

        # 6) "Fichier import" : formules PAR LIGNE completes (24 colonnes, cf. docstring en
        #    tete + carrier Node index.js pour le detail complet). Nombre de lignes = nombre
        #    de trackings du TCD (lastTcd-1, entete 2 lignes).
        #
        # BUG TROUVE 2026-08-20 (meme piege deja documente sur TNT/Chronopost) : le ColField
        # natif du TCD ("Categorie") n'affiche QUE les postes REELLEMENT presents dans le mois
        # traite -- son ORDRE/POSITION (colonnes F/G/H...) depend des donnees, PAS fixe comme
        # dans le modele de juin. Un mois avec un poste supplementaire (ex. "code inconnu",
        # absent du modele de juin) DECALE tout ce qui suit -- coder TCD!I/N/O/H/F/J/K en dur
        # (comme le finaliseur de premier jet) casse silencieusement le mapping poste->colonne.
        # Fix : resout chaque lettre par NOM d'en-tete (ligne 2 du TCD natif), APRES le
        # RefreshAll deja fait plus haut.
        headerRowTcd = [normalize_header(wsTcd.Cells(2, c).Value) for c in range(6, 17)]  # F..P

        def find_tcd_col(nom):
            for i, h in enumerate(headerRowTcd):
                if h.lower() == nom.lower():
                    return col_letter(5 + i)  # colonne 6 = F
            return None

        col_adresse = find_tcd_col("Adresse")
        col_assurance = find_tcd_col("Assurance")
        col_colis_vol = find_tcd_col("Colis volumineux")
        col_droits_taxes = find_tcd_col("Droits et taxes")
        col_fret = find_tcd_col("Frêt")
        col_plus_value = find_tcd_col("plus-value BtoC")
        col_tva = find_tcd_col("TVA")
        col_zones_eloignees = find_tcd_col("Zones éloignées")
        missing = [n for n, c in (("Adresse", col_adresse), ("Assurance", col_assurance), ("Colis volumineux", col_colis_vol),
                                    ("Droits et taxes", col_droits_taxes), ("Frêt", col_fret), ("plus-value BtoC", col_plus_value),
                                    ("TVA", col_tva), ("Zones éloignées", col_zones_eloignees)) if c is None]
        if missing:
            print(f"INFO: poste(s) ERP absent(s) du TCD ce mois-ci (aucune ligne) : {', '.join(missing)} -- colonne(s) correspondante(s) non trouvée(s), formules correspondantes videes.")

        wsImp = wb.Sheets("Fichier import")
        if wsImp.AutoFilterMode:
            wsImp.AutoFilterMode = False
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 9).End(xlUp).Row
        LAST_COL_IMPORT = 24
        newLastImp = max(lastTcd - 1, 2)

        # BUG TROUVE 2026-08-25 (signale par l'utilisateur : "Fichier import" n'a pas de
        # formules comme le fichier fait-main) : les formules ci-dessous referencaient le TCD
        # par LIGNE FIXE (TCD!col{tcdrow}) -- fige en valeurs plus bas AVANT suppression des
        # lignes vides pour eviter un decalage (une ligne supprimee aurait laisse les lignes
        # suivantes pointer vers le mauvais tracking du TCD). Fix : remplace par un XLOOKUP
        # dynamique sur le tracking (colonne I, deja utilise par les autres formules
        # ci-dessus) -- exactement le meme principe que les formules "zone colis poids
        # assurance"/"Facture UPS" deja en XLOOKUP par tracking, jamais par position. Une
        # formule qui cherche par tracking reste valide meme apres suppression de lignes ->
        # permet de GARDER des formules vivantes dans "Fichier import" (fidele au fait-main)
        # au lieu de figer en valeurs. Definie ICI (avant formulas_import) car la formule Zone
        # (regles 2/4, CDC 2026-08-27) a aussi besoin de tcd_lookup(col_fret)/(col_plus_value).
        def tcd_lookup(col):
            return f"_xlfn.XLOOKUP(I{{row}},TCD!E:E,TCD!{col}:{col})"

        # CDC pole transport 2026-08-27 -- regles Zone (priorite dans cet ordre) :
        #   5) compte WV5788 (Verde Trad, colonne D deja resolue en "Verde Trad") -> Zone=
        #      "France" FORCE (override absolu, prioritaire a tout le reste) + Mode envoi="ST"
        #      force (cf. formulas_import[16] plus bas).
        #   2) Zone brute=0/vide + Fret entre 3 et 8EUR (fourchette large validee, "fret ~5EUR")
        #      + Pays vide ou FR -> Zone="France".
        #   3) Zone brute=0/vide + PAS de fret -> laisser 0 ICI ; corrige ensuite en PYTHON via
        #      le fichier import du mois precedent (zone_m1_map, cf. bloc post-Calculate plus
        #      bas) -- ne peut pas etre une formule Excel (donnee externe).
        #   sinon : zone_brute (formule d'origine, inchangee).
        #   4) (applique sur le RESULTAT ci-dessus, PAS avant) : plus-value BtoC<2EUR ET
        #      Zone="France" ET pas WV5788 -> marqueur "A VERIFIER" (signale au lieu de modifier
        #      silencieusement -- WV5788 exempte car regle 5 est un override volontaire, pas une
        #      zone "a verifier").
        # Regle 1 (interdit Zone=0) : satisfaite une fois le repli Python M-1 applique -- la
        # formule Excel seule peut encore produire 0 de facon transitoire (regle 3), corrige
        # dans le bloc post-Calculate plus bas.
        # zone_brute reference directement C{row} (colonne interne de controle, formula index 2
        # ci-dessous) plutot que de reecrire l'expression XLOOKUP -- C{row} EST deja "IF(LEN>2,
        # zone,'inconnu')" via son propre COUNTIF, donc equivalent a l'ancienne logique inline
        # (LEN(C)>2 -> C ; sinon Pays=FR -> France ; sinon C tel quel, qui vaut alors "inconnu"
        # ou un vrai zonage court) SANS reevaluer XLOOKUP plusieurs fois dans la meme formule
        # (evite un formule M demesuree sur 6000-8700 lignes).
        zone_brute = 'IF(LEN(C{row})>2,C{row},IF(L{row}="FR","France",C{row}))'
        fret_expr = tcd_lookup(col_fret) if col_fret else '""'
        plus_value_expr = tcd_lookup(col_plus_value) if col_plus_value else '""'
        zone_calculee_sans_wv = (
            f'IF(AND(OR({zone_brute}=0,{zone_brute}=""),{fret_expr}<>"",{fret_expr}>=3,{fret_expr}<=8,OR(L{{row}}="",L{{row}}="FR")),"France",'
            f'IF(AND(OR({zone_brute}=0,{zone_brute}=""),OR({fret_expr}="",{fret_expr}=0)),0,'
            f'{zone_brute}))'
        )
        formula_zone = (
            f'=IF(D{{row}}="Verde Trad","France",'
            f'IF(AND({plus_value_expr}<>"",{plus_value_expr}<2,({zone_calculee_sans_wv})="France"),'
            f'"A VERIFIER",{zone_calculee_sans_wv}))'
        )

        # BUG TROUVE 2026-08-31 (tracking A1912WTZX8M, frais de douane GB) : RIGHT(LEFT(I,8),6)
        # suppose un format "1Z"+compte(6) -- faux pour les trackings de frais de dedouanement/
        # ajustement (format compte(6) SANS prefixe "1Z", ex. A1912WTZX8M, A1912WVN938,
        # A1912WSRQJ7 -- meme famille que celles deja vues dans le modele fait-main de juin,
        # A1912WTRHVP/A1912WVMB7Q). Sur ce format, RIGHT(LEFT(I,8),6) extrait "912WTZ" (faux)
        # au lieu de "A1912W" (vrai compte, verifie present dans 'Comptes UPS' -> "UPS").
        # Fix : si l'extraction standard ne matche aucun compte, retente avec LEFT(I,6) (compte
        # en tete du tracking, sans prefixe "1Z") avant de retomber sur "inconnu" -- retrouve le
        # VRAI nom de compte (peut differer de "UPS" simple, ex. UPS_COD) plutot que de forcer
        # une valeur en dur. Corrige aussi Date validite/Ref.1/Ref.2 en aval (deja des formules
        # generiques par tracking, rien a changer la -- elles dependaient seulement de ce que
        # Transporteur ne soit plus "inconnu" pour se comporter comme toute autre ligne).
        compte_standard = "RIGHT(LEFT(I{row},8),6)"
        compte_sans_1z = "LEFT(I{row},6)"
        formulas_import = {
            2: '=IF(COUNTIF(\'zone colis poids assurance\'!D:D,I{row})=0,"inconnu",_xlfn.XLOOKUP(I{row},\'zone colis poids assurance\'!D:D,\'zone colis poids assurance\'!F:F))',  # C Zone (colonne interne de controle)
            4: (f'=IF(COUNTIF(\'Comptes UPS\'!A:A,{compte_standard})<>0,'
                f'_xlfn.XLOOKUP({compte_standard},\'Comptes UPS\'!A:A,\'Comptes UPS\'!B:B),'
                f'IF(COUNTIF(\'Comptes UPS\'!A:A,{compte_sans_1z})<>0,'
                f'_xlfn.XLOOKUP({compte_sans_1z},\'Comptes UPS\'!A:A,\'Comptes UPS\'!B:B),'
                f'"inconnu"))'),  # D Transport
            6: "=_xlfn.XLOOKUP(I{row},'Facture UPS'!Y:Y,'Facture UPS'!T:T)",  # F Ref.1 (simplifie : pas de IF("","") ici, valeur brute)
            7: "=_xlfn.XLOOKUP(I{row},'Facture UPS'!Y:Y,'Facture UPS'!U:U)",  # G Ref.2
            9: "=TCD!E{tcdrow}",  # I N° Tracking
            11: '=IF(B{row}="particulier","P",IF(X{row}="","E","P"))',  # K E/P
            12: "=_xlfn.XLOOKUP(I{row},'Facture UPS'!Y:Y,'Facture UPS'!CH:CH)",  # L Pays
            13: formula_zone,  # M Zone (regles 1/2/4/5 CDC 2026-08-27, cf. ci-dessus)
            14: "=MAX(_xlfn.XLOOKUP(I{row},'zone colis poids assurance'!D:D,'zone colis poids assurance'!E:E),A{row})",  # N Nbr Colis
            15: '=IF(D{row}="UPS_COD",_xlfn.XLOOKUP(I{row},\'zone colis poids assurance\'!D:D,\'zone colis poids assurance\'!J:J),_xlfn.XLOOKUP(I{row},\'zone colis poids assurance\'!D:D,\'zone colis poids assurance\'!I:I))',  # O Poids
            16: '=IF(D{row}="Verde Trad","ST",IF(COUNTIF(\'ST SV\'!H:H,I{row})=0,"inconnu",_xlfn.XLOOKUP(I{row},\'ST SV\'!H:H,\'ST SV\'!N:N)))',  # P mode envoi (Verde Trad force "ST", regle 5)
        }

        if col_tva:
            formulas_import[17] = f'=IF({tcd_lookup(col_tva)}="",0,0.2)'  # Q TVA
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 17), wsImp.Cells(newLastImp, 17)).__setattr__("Value", 0))
        if col_droits_taxes:
            formulas_import[18] = f'=IF({tcd_lookup(col_droits_taxes)}=0,"",{tcd_lookup(col_droits_taxes)})'  # R Droits et taxes
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 18), wsImp.Cells(newLastImp, 18)).ClearContents())
        formulas_import[19] = "=IF(_xlfn.XLOOKUP(I{row},'zone colis poids assurance'!D:D,'zone colis poids assurance'!G:G)=0,\"\",MAX(10,ROUNDUP(0.02*_xlfn.XLOOKUP(I{row},'zone colis poids assurance'!D:D,'zone colis poids assurance'!G:G),2)))"  # S Assurance
        if col_zones_eloignees:
            # BUG TROUVE 2026-08-20 : =IF(TCD!col="","",40) (formule modele litterale) declenche
            # a tort le forfait 40EUR quand la seule ligne "Zones eloignees" du tracking a un
            # montant reel de 0EUR (ex. code ESP "Supplt zone enlev. etendue" facture a 0,00EUR
            # -- observe sur 763/8719 trackings reels de juin 2026) : le TCD affiche alors 0
            # (pas vide), donc TCD!col<>"" est vrai -> formule litterale donne 40 alors que le
            # fichier reel livre montre 0. Fix : traiter aussi 0 comme "absence de charge",
            # coherent avec le fait qu'un forfait de 40EUR n'a pas de sens pour un poste facture
            # 0EUR.
            formulas_import[20] = f'=IF(OR({tcd_lookup(col_zones_eloignees)}="",{tcd_lookup(col_zones_eloignees)}=0),"",40)'  # T Zones éloignées
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 20), wsImp.Cells(newLastImp, 20)).ClearContents())
        if col_colis_vol:
            cv = tcd_lookup(col_colis_vol)
            formulas_import[21] = (
                f'=IF({cv}=0,"",IF({cv}<3,3,'
                f'IF({cv}<15,15,IF({cv}<50,35,'
                f'IF({cv}<100,59,IF({cv}<150,177,'
                f'ROUNDUP({cv}/59,0)*59))))))'
            )  # U Colis volumineux
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 21), wsImp.Cells(newLastImp, 21)).ClearContents())
        if col_adresse:
            formulas_import[22] = f'=IF({tcd_lookup(col_adresse)}="","",11.5)'  # V Adresses
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 22), wsImp.Cells(newLastImp, 22)).ClearContents())
        if col_fret:
            formulas_import[23] = f'=IF({tcd_lookup(col_fret)}="","",ROUND({tcd_lookup(col_fret)},2))'  # W Frêt
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 23), wsImp.Cells(newLastImp, 23)).ClearContents())
        if col_plus_value:
            formulas_import[24] = f'=IF({tcd_lookup(col_plus_value)}="","",{tcd_lookup(col_plus_value)})'  # X plus-value BtoC
        else:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 24), wsImp.Cells(newLastImp, 24)).ClearContents())

        for col_idx, tmpl in formulas_import.items():
            retry(lambda c=col_idx, t=tmpl: wsImp.Range(wsImp.Cells(2, c), wsImp.Cells(newLastImp, c))
                  .__setattr__("Formula", [[t.format(row=r, tcdrow=r + 2) for _ in [0]][0] for r in range(2, newLastImp + 1)]))

        if date_validite_serial is not None:
            wsImp.Cells(2, 5).Value = date_validite_serial
        else:
            print("AVERTISSEMENT: mois de facturation introuvable -> 'Fichier import'!E2 non mise à jour, reste celle du modèle.")
        if newLastImp > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(3, 5), wsImp.Cells(newLastImp, 5))
                  .__setattr__("Formula", [["=E{prev}".format(prev=r - 1)] for r in range(3, newLastImp + 1)]))

        if newLastImp < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLastImp + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT)).ClearContents())
        print(f"'Fichier import' : formules reconstruites jusqu'à la ligne {newLastImp}.")

        # Colonne B "E/P ERP" : ECRITE EN VALEUR (declaration ERP, export WMS m/m-1, meme
        # mecanisme que Delivengo/Geodis/DPD/FedEx -- PAS une formule Excel, la donnee vient
        # d'un fichier externe). Lue via la colonne I (N° Tracking, deja calculee ci-dessus)
        # apres Calculate() pour resoudre les formules TCD!E.
        xl.Calculate()
        if newLastImp >= 2 and ep_map:
            ep_values = []
            for r in range(2, newLastImp + 1):
                tracking = str(wsImp.Cells(r, 9).Value or "").strip()
                ep_values.append([ep_map.get(tracking, "")])
            retry(lambda: wsImp.Range(wsImp.Cells(2, 2), wsImp.Cells(newLastImp, 2)).__setattr__("Value", ep_values))
            print(f"'Fichier import' : colonne B (E/P ERP) renseignée depuis l'export WMS pour {sum(1 for v in ep_values if v[0])}/{newLastImp - 1} ligne(s).")
        else:
            print("AVERTISSEMENT: pas d'export WMS fourni -> colonne B (E/P ERP) laissée vide, E/P final se rabattra sur 'P' par défaut (sauf plus-value BtoC détectée).")

        xl.Calculate()

        # Trackings SANS AUCUNE charge facturable (tous postes ERP a 0 -- ligne "INF"/retour
        # indelivrable isolee sans ligne FRT) : EXCLUS de "Fichier import" (confirme par
        # comparaison au fichier reel de juin 2026, 0 ligne totalement vide sur 8736 -- decision
        # utilisateur 2026-08-20 : garder TOUTES les lignes dans "Facture UPS", filtrer
        # UNIQUEMENT "Fichier import"). BUG TROUVE 2026-08-25 (signale par l'utilisateur :
        # "Fichier import" n'a pas de formules comme le fichier fait-main) -- les formules des
        # postes ERP ci-dessus etaient figees en VALEURS ici avant suppression des lignes vides
        # (necessaire tant qu'elles referencaient le TCD par ligne fixe). Elles referencent
        # desormais le tracking par XLOOKUP (cf. plus haut) -- restent valides meme apres
        # suppression de lignes, donc PLUS BESOIN de figer : "Fichier import" garde des
        # formules vivantes, fidele au fichier fait-main.
        if newLastImp >= 2:
            # Lecture en BLOC (1 seul aller-retour COM au lieu de milliers) -- la boucle
            # cellule-par-cellule/Delete()-par-ligne precedente timeoutait (>590s sur 8700+
            # lignes, BUG PERF TROUVE 2026-08-20).
            rangeImp = wsImp.Range(wsImp.Cells(2, 1), wsImp.Cells(newLastImp, LAST_COL_IMPORT))
            valuesImp = rangeImp.Value  # tuple de tuples (lecture seule, formules NON figees)
            # Colonnes montant (postes ERP, 1-based -> index 0-based dans valuesImp) : R Droits
            # et taxes(18), S Assurance(19), T Zones eloignees(20), U Colis volumineux(21),
            # V Adresses(22), W Fret(23), X plus-value BtoC(24).
            POSTE_COLS_0BASED = [17, 18, 19, 20, 21, 22, 23]
            lignes_vides = [i for i, row in enumerate(valuesImp) if all((row[c] in (None, "", 0)) for c in POSTE_COLS_0BASED)]
            if lignes_vides:
                # Regroupe les indices contigus en plages pour minimiser le nombre d'appels
                # Delete() (une plage multi-lignes se supprime en 1 seul appel COM).
                plages = []
                debut = lignes_vides[0]
                prec = lignes_vides[0]
                for idx in lignes_vides[1:]:
                    if idx != prec + 1:
                        plages.append((debut, prec))
                        debut = idx
                    prec = idx
                plages.append((debut, prec))
                # Suppression de la DERNIERE plage vers la PREMIERE (evite tout decalage des
                # indices de lignes non encore traitees).
                for i0, i1 in reversed(plages):
                    r0, r1 = i0 + 2, i1 + 2  # index 0-based -> ligne Excel 1-based (+entete)
                    retry(lambda a=r0, b=r1: wsImp.Range(wsImp.Cells(a, 1), wsImp.Cells(b, LAST_COL_IMPORT)).EntireRow.Delete())
                newLastImp -= len(lignes_vides)
                print(f"'Fichier import' : {len(lignes_vides)} ligne(s) sans aucune charge facturable supprimée(s) (tous postes ERP à 0).")

        wsImp.Range(wsImp.Cells(1, 1), wsImp.Cells(max(newLastImp, 2), LAST_COL_IMPORT)).AutoFilter()
        xl.Calculate()

        # BUG TROUVE 2026-08-26 (signale par l'utilisateur : "la feuille ne doit pas trainer
        # apres la derniere ligne") : ClearContents() (etape precedente) vide le CONTENU des
        # lignes residuelles du modele clone (mise en forme/bordures gardees), donc Excel
        # continue de considerer ces lignes comme faisant partie de la feuille (UsedRange
        # gonfle bien au-dela de newLastImp -- constate : max_row=7261 pour 6914 lignes de
        # donnees reelles, 347 lignes de residu "vide" mais toujours dans la zone utilisee).
        # Fix : supprimer PHYSIQUEMENT les lignes au-dela de newLastImp (EntireRow.Delete(),
        # pas juste ClearContents) -- force Excel a reduire UsedRange a la vraie derniere ligne.
        wsImpUsedRange = wsImp.UsedRange
        wsImpUsedLastRow = wsImpUsedRange.Row + wsImpUsedRange.Rows.Count - 1
        if wsImpUsedLastRow > newLastImp:
            retry(lambda: wsImp.Range(
                wsImp.Cells(newLastImp + 1, 1), wsImp.Cells(wsImpUsedLastRow, LAST_COL_IMPORT)
            ).EntireRow.Delete())
            print(f"'Fichier import' : {wsImpUsedLastRow - newLastImp} ligne(s) residuelle(s) (mise en forme sans donnees) supprimee(s) au-dela de la ligne {newLastImp}.")
        xl.Calculate()

        # CDC pole transport 2026-08-27 : repli Zone M-1 (regle 3) + validations console
        # (TVA hors UE, point B ; SV/ST, regles 6/7) -- APRES la suppression des lignes vides
        # et residuelles ci-dessus (numeros de ligne definitifs, plus aucun decalage a venir),
        # AVANT le Save()/Close() plus bas. Une seule lecture COM groupee (colonnes I->Q) pour
        # limiter les allers-retours sur potentiellement 6000-8700 lignes (meme principe deja
        # applique ligne ~1059, "lecture en BLOC").
        if newLastImp >= 2:
            rangeZoneVal = wsImp.Range(wsImp.Cells(2, 9), wsImp.Cells(newLastImp, 17))  # I..Q
            valuesZoneVal = rangeZoneVal.Value
            # Offsets 0-based DANS cette plage I..Q (I=0) -- PAS les offsets 1-based habituels
            # de "Fichier import" ni ceux de la plage A..X plus haut (piege deja documente,
            # BUG TROUVE 2026-08-27 sur POSTE_COLS_VALUES) : I=0,...,L(Pays)=3,M(Zone)=4,
            # N=5,O=6,P(Mode envoi)=7,Q(TVA)=8.
            IDX_TRACKING, IDX_PAYS, IDX_ZONE, IDX_MODE, IDX_TVA = 0, 3, 4, 7, 8

            # Repli Zone M-1 (regle 3) : Zone encore 0/vide apres le calcul Excel -> tracking
            # cherche dans le fichier import du mois precedent (zone_m1_map).
            corrections_m1 = []  # (index 0-based dans valuesZoneVal, zone)
            for i, row in enumerate(valuesZoneVal):
                zone_v = row[IDX_ZONE]
                if zone_v in (0, "0", "", None):
                    t = str(row[IDX_TRACKING] or "").strip()
                    zone_m1 = zone_m1_map.get(t)
                    if zone_m1:
                        corrections_m1.append((i, zone_m1))
            if corrections_m1:
                for i, zone_m1 in corrections_m1:
                    r = i + 2
                    wsImp.Cells(r, 13).Value = zone_m1  # M Zone, EN VALEUR (donnee externe)
                print(f"'Fichier import' : {len(corrections_m1)} tracking(s) à Zone=0 sans frêt complété(s) via le fichier import du mois précédent.")
                xl.Calculate()
                valuesZoneVal = rangeZoneVal.Value  # relecture : les validations ci-dessous doivent voir les zones corrigees

            # Validation TVA=0 hors UE (point B) -- ALERTE CONSOLE UNIQUEMENT, aucune ecriture.
            suspects_tva = []
            for i, row in enumerate(valuesZoneVal):
                tva_v = row[IDX_TVA]
                if tva_v and num(tva_v) > 0:
                    pays_v = str(row[IDX_PAYS] or "").strip().upper()
                    if pays_v and pays_v not in PAYS_UE:
                        suspects_tva.append((i + 2, pays_v))
            if suspects_tva:
                echantillon = ", ".join(f"L{r}:{p}" for r, p in suspects_tva[:20])
                print(f"ALERTE: {len(suspects_tva)} ligne(s) à TVA calculée (20%) pour un pays hors UE (a verifier) : {echantillon}{' ...' if len(suspects_tva) > 20 else ''}.")

            # Validation croisee SV/ST (regles 6/7) -- ALERTE CONSOLE UNIQUEMENT, table
            # PAYS_SV_ST = zoning fourni par le pole transport, sert a detecter une
            # incoherence avec le calcul existant, jamais a le modifier.
            incoherences_sv_st = []
            for i, row in enumerate(valuesZoneVal):
                pays_v = str(row[IDX_PAYS] or "").strip().upper()
                ref = PAYS_SV_ST.get(pays_v)
                if not ref:
                    continue
                zones_sv, zones_st = ref
                zone_v = str(row[IDX_ZONE] or "").strip()
                mode_v = str(row[IDX_MODE] or "").strip().upper()
                if mode_v == "SV" and zones_sv and zone_v not in zones_sv:
                    incoherences_sv_st.append((i + 2, pays_v, zone_v, mode_v, "/".join(zones_sv)))
                elif mode_v == "ST" and zones_st and zone_v not in zones_st:
                    incoherences_sv_st.append((i + 2, pays_v, zone_v, mode_v, "/".join(zones_st)))
            if incoherences_sv_st:
                echantillon = ", ".join(f"L{r}:{p} zone={z} mode={m} (attendu {att})" for r, p, z, m, att in incoherences_sv_st[:20])
                print(f"ALERTE: {len(incoherences_sv_st)} ligne(s) Zone/Mode envoi incohérente(s) avec le zoning UPS fourni (aucune valeur modifiée) : {echantillon}{' ...' if len(incoherences_sv_st) > 20 else ''}.")

        # 7) "Demande avoir" (1Z79) : Tracking/Nb colis/Montant/Cause -- Factures/Poids/Mode
        #    livraison laisses vides (saisie manuelle du pole transport, jamais remplis meme
        #    dans le fichier fait-main). Purge d'abord les lignes du modele clone (mois
        #    precedent), puis ecrit les trackings 1Z79 du mois traite (demandes_avoir_1z79,
        #    calcule plus haut, filtre code_classe=FRT uniquement).
        try:
            wsAvoir = wb.Sheets("Demande avoir")
            oldLastAvoir = wsAvoir.Cells(wsAvoir.Rows.Count, 1).End(xlUp).Row
            if oldLastAvoir >= 2:
                retry(lambda: wsAvoir.Range(wsAvoir.Cells(2, 1), wsAvoir.Cells(oldLastAvoir, 7)).ClearContents())
            if demandes_avoir_1z79:
                rows_avoir = [
                    [tracking, None, 1, None, None, montant, CAUSE_1Z79]
                    for tracking, montant in sorted(demandes_avoir_1z79.items())
                ]
                retry(lambda: wsAvoir.Range(wsAvoir.Cells(2, 1), wsAvoir.Cells(1 + len(rows_avoir), 7))
                      .__setattr__("Value", rows_avoir))
                print(f"'Demande avoir' : {len(rows_avoir)} colis 1Z79 reportés.")
        except Exception as e:
            print("Avertissement : remplissage 'Demande avoir' a échoué :", e)

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)

        # 8) Export "Fichier import" en VALEURS (pas en JS recalcule a part) : remontee
        # utilisateur 2026-08-26 -- les alertes affichees (POIDS=0 etc.) et le CSV livre
        # doivent refleter le VRAI classeur (formules XLOOKUP resolues), pas le calcul JS
        # separe du carrier Node (qui peut diverger, ex. Réf.1/Réf.2 jamais remplis cote JS).
        # Meme mecanisme que finaliser_colissimo.py/finaliser_fedex.py (2026-08-24/25) :
        # rouvrir le classeur SAUVEGARDE en lecture (etat stable, pas de valeur d'erreur COM
        # transitoire), lire les colonnes D->X (4->24 = les 21 colonnes standard ERP presentes
        # dans "Fichier import", en commencant par "Transporteur" -- BUG TROUVE 2026-08-26 :
        # demarrer a la colonne E (DateValidite) au lieu de D decalait TOUT l'export d'1
        # colonne vers la gauche par rapport a IMPORT_COLUMNS. TaxeGasoil/NbColis n'existent
        # pas dans cet onglet, resteront vides cote CSV final -- cf. IMPORT_COLUMNS, meme
        # convention que Colissimo).
        wbRead = None
        try:
            wbRead = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=True))
            wsImpRead = wbRead.Sheets("Fichier import")
            impLast = wsImpRead.Cells(wsImpRead.Rows.Count, 9).End(xlUp).Row  # colonne I = N Tracking
            if impLast >= 2:
                values = retry(lambda: wsImpRead.Range(wsImpRead.Cells(2, 4), wsImpRead.Cells(impLast, 24)).Value)
                # BUG TROUVE 2026-08-26 (constate en test reel) : une ligne fantome (Transporteur
                # ="inconnu", Tracking="0.0") apparaissait au tout debut de l'export -- meme
                # famille de bug que Colissimo/FedEx (cellule encore en etat transitoire au
                # moment de la lecture COM, malgre la reouverture apres sauvegarde). Filtre sur
                # le tracking (index 5 dans row = colonne I=9 moins le decalage colonne D=4).
                # BUG TROUVE 2026-08-27 : le tracking fantome n'est pas toujours "0.0" -- observe
                # une fois avec Tracking="AB" (valeur imprevisible selon l'etat transitoire COM),
                # qui echappait donc au filtre. Signal plus fiable : Transporteur="inconnu" (index
                # 0 = colonne D) COMBINE a tous les postes ERP vides -- c'est la vraie signature
                # de la ligne fantome, le tracking associe n'etant qu'un symptome variable.
                # BUG TROUVE 2026-08-27 (test reel, "tuple index out of range") : cette plage
                # "values" va de D a X (24-4=21 colonnes, index 0-20), PAS la meme etendue que
                # POSTE_COLS_0BASED plus haut (calcule sur une plage A->X = 24 colonnes). Postes
                # ERP (Droits et taxes -> plus-value BtoC) dans CETTE plage D->X : index 14-20
                # (R=DroitsTaxes(14), S=Assurance(15), T=ZonesEloignees(16), U=ColisVolumineux(17),
                # V=Adresses(18), W=Fret(19), X=PlusValueB2C(20)) -- PAS 17-23 (hors plage).
                POSTE_COLS_VALUES = [14, 15, 16, 17, 18, 19, 20]
                # BUG TROUVE 2026-08-28 (test reel, tracking "A1912WTZX8M") : la ligne fantome
                # n'a PAS TOUJOURS Transporteur="inconnu" -- observe une fois avec Transporteur
                # ="inconnu" MAIS un poste ERP non vide (Droits et taxes=67,66), qui echappait
                # donc au filtre AND ci-dessus (conjonction trop stricte). Chaque nouvelle
                # variante de cette famille de bug (valeur COM transitoire au moment de la
                # lecture post-sauvegarde) a une signature differente et imprevisible ("0.0",
                # "AB", "A1912WTZX8M"...) -- deviner un nouveau motif a chaque fois n'est pas
                # fiable. Fix definitif : comparer contre la LISTE REELLE des trackings connus
                # (lignes_retenues, deja en memoire Python, source de verite construite AVANT
                # toute manipulation Excel) -- un tracking absent de cet ensemble est PAR
                # DEFINITION une ligne fantome, quelle que soit sa valeur exacte.
                trackings_reels = {
                    str(r[COL["numero_suivi"]] if len(r) > COL["numero_suivi"] else "").strip()
                    for r in lignes_retenues
                }
                trackings_reels.discard("")
                def est_fantome(row):
                    tracking_str = str(row[5] or "").strip()
                    if tracking_str not in trackings_reels:
                        return True
                    if tracking_str in ("", "0.0", "(vide)", "inconnu"):
                        return True
                    transporteur_str = str(row[0] or "").strip().lower()
                    if transporteur_str == "inconnu" and all((row[c] if c < len(row) else None) in (None, "", 0) for c in POSTE_COLS_VALUES):
                        return True
                    return False
                n_fantomes = sum(1 for row in values if est_fantome(row))
                values = [row for row in values if not est_fantome(row)]
                if n_fantomes:
                    print(f"'Fichier import' : {n_fantomes} ligne(s) fantome(s) (tracking absent des donnees sources, artefact COM transitoire) exclue(s) de l'export final.")
                n_err = 0
                def clean(v):
                    nonlocal n_err
                    if isinstance(v, int) and not isinstance(v, bool) and v < -1000000:
                        n_err += 1
                        return ""
                    return "" if v is None else v
                export_path = os.path.splitext(sortie)[0] + "_import_valeurs.csv"
                with open(export_path, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f, delimiter=";")
                    for row in values:
                        w.writerow([clean(v) for v in row])
                if n_err:
                    print(f"AVERTISSEMENT: {n_err} cellule(s) en erreur COM lors de l'export Fichier import (valeurs) -- laissees vides, a verifier.")
                print(f"EXPORT_IMPORT_VALEURS:{export_path}")
        except Exception as e:
            print("Export Fichier import (valeurs) ignore :", e)
        finally:
            if wbRead is not None:
                try:
                    wbRead.Close(SaveChanges=False)
                except Exception:
                    pass

        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
