#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR FEDEX — produit "AAAA_MM_Facture Fedex.xlsx" A L'IDENTIQUE du
fichier fait a la main (6 feuilles : Zoning, Bilan factures, Shipment
Detail, TCD, Bilan clients, Import ERP), en partant du fichier existant
comme MODELE. Entree = CSV BRUT FedEx (export portail, en-tete FR, meme
fichier que celui lu par le carrier Node index.js).

Classeur reel :
  Shipment Detail : 1 ligne = 1 colis/envoi, donnees brutes CSV FedEx
    (colonnes F->BS, en-tete FR, resolues PAR NOM), colonnes A->E
    CALCULEES (A="Montant"=BK+BN, B="Clients"=ID client MANUEL (laisse
    vide a la generation, saisi apres coup, cf. regle transversale ID
    client), C="Categorie"=toujours vide (confirme juin 2026), D="E/P"
    =VALEUR LITERALE ("entreprise"/"particulier", calculee ICI via
    l'export WMS "expeditions_brut" m/m-1 -- meme mecanisme que Delivengo/
    Geodis/DPD, PAS une formule Excel : la colonne D du fichier reel de
    juin 2026 est 100% litterale sur 3607/3607 lignes), E="Mode d'envoi"
    =formule =IF(Service="FedEx Priority","ST","FICP").
  TCD : VRAI TCD natif (RowFields=Shipment Tracking Number+Recipient
    Country/Territory, DataFields=Somme de Montant+Somme de Shipment
    Rated Weight (Pounds)) + colonnes MANUELLES juxtaposees E (Poids kg,
    ROUNDUP*0.453592) / F (E/P, XLOOKUP vers Shipment Detail!D).
  Bilan factures : VRAI TCD (RowField=Invoice Number, DataFields=Fret
    HT+Total TTC) + colonne D "PDF" MANUELLE (collee/saisie), E
    "Ecart"=C-D.
  Bilan clients : VRAI TCD (RowField=Clients, DataField=Somme de
    Montant), source='Shipment Detail' (PAS 'TCD').
  Import ERP : formules PAR LIGNE (pas de decalage fixe comme TNT),
    referencant TCD par XLOOKUP(tracking) et Shipment Detail par
    XLOOKUP(tracking), PAS par position.

PIEGE CONNU (meme famille que DPD/Geodis/Mondial Relay/Chronopost/TNT) :
  les 3 PivotTables (TCD/Bilan factures/Bilan clients) du modele ont un
  PivotCache SOURCE ETROITE/FIGEE ('Shipment Detail'!C1:C70 ou C1:C2,
  constate via COM sur le modele de juin 2026) -- DOIT etre redirige
  vers la vraie plage large de 'Shipment Detail' a chaque generation,
  sinon le TCD reste fige sur les anciennes donnees du modele.

Detection "facture d'un autre mois" (RESOLUE 2026-08-20, cf. commentaire en
tete de facturation-app/src/carriers/fedex/index.js) : une facture du CSV
brut sans PDF correspondant parmi les PDF portail fournis est exclue
automatiquement (le pole transport ne recoit le PDF qu'une fois la facture
reellement due ce mois-ci) -- validee exacte sur juin 2026 (exclut les 2
memes factures que le fichier reel livre, Fret recalcule = 45515,83EUR
exactement). Ne s'applique que si au moins 1 PDF est fourni.

Necessite : Windows + Excel + pywin32.
Usage :
  python finaliser_fedex.py "<modele.xlsx>" "<sortie.xlsx>" --csv <csv1> [<csv2>...] [--brut <brutM.xlsx> [<brutM-1.xlsx>]] [--pdf <pdf1> [<pdf2>...]]
"""
import sys, os, shutil, re, csv


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def compare_key(h):
    s = normalize_header(h).lower()
    s = s.translate(str.maketrans("éèêëàâäùûüôöîïç’‘", "eeeeaaauuuooiic''"))
    s = re.sub(r"[.,]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def col_index(header, name):
    target = compare_key(name)
    for i, h in enumerate(header):
        if compare_key(h) == target:
            return i
    return None


def col_letter(idx0):
    import openpyxl
    return openpyxl.utils.get_column_letter(idx0 + 1)


def read_fedex_csv(path):
    """CSV brut FedEx (export portail, en-tete FR, separateur virgule, UTF-8 --
    confirme sur les CSV bruts reels de juin 2026, cf. carrier Node index.js)."""
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def coerce(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.fullmatch(r"-?0\d+", s):
        return v
    if re.fullmatch(r"-?\d+(,\d+)?", s):
        return float(s.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


# ---- E/P : export WMS "expeditions_brut" (meme mecanisme que Delivengo/Geodis/DPD) ----
X_BRUT_TRACK, X_BRUT_DES_PARTICULIER = 41, 16  # PRO_TRACKING (AP) / DES_PARTICULIER (Q)


def load_brut_ep(brut_paths):
    """Map {tracking (str) -> 'entreprise'/'particulier'}. Ordre des fichiers = priorite
    (mois courant, puis M-1) -- la PREMIERE occurrence gagne, meme regle que Delivengo."""
    import openpyxl
    m = {}
    for p in [bp for bp in brut_paths if bp]:
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > X_BRUT_TRACK:
                t = str(row[X_BRUT_TRACK] or "").strip()
                v = str(row[X_BRUT_DES_PARTICULIER] or "").strip().lower()
                ep = "particulier" if v == "particulier" else ("entreprise" if v in ("entreprise", "point relais") else None)
                if t and ep and t not in m:
                    m[t] = ep
        wb.close()
    return m


def extract_pdf_infos(pdf_paths):
    """PDF FedEx ("Facture de frais de transport") : bloc labels/valeurs colle par
    pypdf ("No de Client:\\nNo de Facture:\\n...\\nMontant du\\n*****1234\\n634323394\\n...")
    + ligne "Total dû EUR <montant>" -- confirme exact sur les 9 PDF reels de juin 2026
    (1 a 195 pages)."""
    try:
        import pypdf
    except ImportError:
        return []
    results = []
    pat_inv = re.compile(
        r"No de Client\s*:\s*\nNo de Facture\s*:\s*\nDate de la facture\s*:\s*\n"
        r"Date d.[eé]ch[eé]ance\s*\nMontant d[uû]\s*\n\*+\d+\s*\n(\d+)\s*\n"
    )
    pat_total = re.compile(r"Total d[uû]\s*EUR\s*([\d.,]+)")
    # Bordereaux (1 par envoi/tracking) : pypdf aplatit tout sur 1 seule ligne texte (contrairement
    # a pdf-parse cote Node qui insere un retour a la ligne par cellule) -- pattern SANS \n entre
    # les groupes, cf. carrier Node index.js pour le meme controle avec un pattern adapte.
    pat_bordereau = re.compile(r"([\d.,]+)(\d{2}/\d{2}/\d{4})(\d{9,15})\s+([\d.,]+)\s+([\d.,]+)")
    # Supplements "surplus" (Zones eloignees/Colis volumineux PDF/Plus-value BtoC) : jamais
    # presents dans le CSV Shipment Detail (mail pole transport "Oubli facturation surplus
    # Fedex janvier a mai 2026", 2026-08-20, 7919,64EUR de surplus non factures sur 5 mois
    # faute d'automatisation) -- extraits par mots-cles, 1 bloc par bordereau (du debut du
    # tracking jusqu'au bordereau suivant), montant precede son libelle sur la meme ligne
    # texte (pypdf aplatit, contrairement a pdf-parse cote Node qui insere 1 retour a la ligne
    # par cellule -- meme logique de decoupage par bloc, pattern adapte). Colis volumineux PDF
    # s'ADDITIONNE a la regle Excel existante (longueur>60cm=10EUR), confirme pole transport
    # 2026-08-20 -- 2 mecanismes independants, pas un remplacement.
    pat_track_start = re.compile(r"\d{2}/\d{2}/\d{4}(\d{9,15})")
    kw_pats = {
        # Zones eloignees : mot-cle generique "zone" (insensible a la casse, pole transport
        # 2026-08-24) + libelle precis "Frais de traitement des importations aux Etats-Unis"
        # (precise par le pole transport 2026-08-25 -- ne contient pas "zone", pattern
        # complementaire), cf. carrier Node index.js (meme changement). BUG TROUVE 2026-08-25 :
        # le PDF tronque le libelle a "...aux Etats-Uni" (SANS le "s" final) -- "s" optionnel.
        "ZonesEloignees": [
            re.compile(r"([\d.,]+)[^\n]{0,60}zone", re.IGNORECASE),
            re.compile(r"([\d.,]+)[^\n]{0,60}Frais de traitement des importations aux [ÉE]tats-Unis?", re.IGNORECASE),
        ],
        # Colis volumineux (PDF) : "Charge pour dépassement de dimension" + "Supplément pour
        # manutention supplémentaire :poids" (precise par le pole transport 2026-08-25, cf.
        # carrier Node index.js, meme bordereau reel 873870075714, 2 lignes distinctes).
        "ColisVolumineuxPdf": [
            re.compile(r"([\d.,]+)[^\n]{0,60}Charge pour d[ée]passement de dimension"),
            re.compile(r"([\d.,]+)[^\n]{0,60}Suppl[ée]ment pour manutention suppl[ée]mentaire\s*:?\s*poids", re.IGNORECASE),
        ],
        "PlusValueB2C": [
            re.compile(r"([\d.,]+)[^\n]{0,60}Plusieurs pi[èe]ces"),
            re.compile(r"([\d.,]+)[^\n]{0,60}Demand Surcharge"),
        ],
    }
    for p in pdf_paths:
        try:
            text = "\n".join((page.extract_text() or "") for page in pypdf.PdfReader(p).pages)
        except Exception:
            continue
        m_inv = pat_inv.search(text)
        m_total = pat_total.search(text)
        bordereaux = [
            {"tracking": m.group(3), "date_envoi": m.group(2),
             "sous_total": float(m.group(5).replace(".", "").replace(",", "."))}
            for m in pat_bordereau.finditer(text)
        ]
        starts = [(m.start(), m.group(1)) for m in pat_track_start.finditer(text)]
        supplements = []
        for i, (idx, tracking) in enumerate(starts):
            next_idx = starts[i + 1][0] if i + 1 < len(starts) else len(text)
            block = text[idx:next_idx]
            for poste, pats in kw_pats.items():
                for pat in pats:
                    m = pat.search(block)
                    if m:
                        supplements.append({"tracking": tracking, "poste": poste,
                                             "montant": float(m.group(1).replace(".", "").replace(",", "."))})
        results.append({
            "file": os.path.basename(p),
            "numero_facture": m_inv.group(1) if m_inv else None,
            "total_du": float(m_total.group(1).replace(".", "").replace(",", ".")) if m_total else None,
            "bordereaux": bordereaux,
            "supplements": supplements,
        })
    return results


def parse_args(argv):
    modele, sortie = argv[1], argv[2]
    csvs, brut, pdfs, cur = [], [], [], None
    for a in argv[3:]:
        if a == "--csv":
            cur = "c"
        elif a == "--brut":
            cur = "b"
        elif a == "--pdf":
            cur = "p"
        elif cur == "c":
            csvs.append(a)
        elif cur == "b":
            brut.append(a)
        elif cur == "p":
            pdfs.append(a)
    return modele, sortie, csvs, brut, pdfs


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def main():
    modele, sortie, csv_paths, brut_paths, pdf_paths = parse_args(sys.argv)
    if not csv_paths:
        raise RuntimeError("Aucun CSV fourni (--csv <shipment_detail.csv> [...]).")
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    # PDF lus TOT (contenu, PAS le nom de fichier -- server.js/multer renomme les fichiers
    # uploades en identifiants aleatoires sans extension/nom d'origine avant de les passer ici
    # -- bug constate 2026-08-20 : le nom "FEDEX_<invoice>_<compte>.pdf" n'existe QUE sur le
    # disque de l'utilisateur avant upload). Reutilise plus bas pour le "Total dû" (pas de 2e
    # lecture).
    pdf_infos_precoce = extract_pdf_infos(pdf_paths) if pdf_paths else []
    pdf_invoice_nums = {p["numero_facture"] for p in pdf_infos_precoce if p["numero_facture"]}

    # Supplements "surplus" par tracking, cumules sur tous les PDF (cf. extract_pdf_infos) --
    # ecrits en VALEUR (pas de formule Excel existante pour ces colonnes dans le modele, cf.
    # meme logique que index.js/carrier Node).
    supplements_par_tracking = {}
    for p in pdf_infos_precoce:
        for s in p.get("supplements", []):
            acc = supplements_par_tracking.setdefault(s["tracking"], {"ZonesEloignees": 0.0, "ColisVolumineuxPdf": 0.0, "PlusValueB2C": 0.0})
            acc[s["poste"]] = round(acc[s["poste"]] + s["montant"], 2)

    header = None
    rows = []
    for p in csv_paths:
        result = read_fedex_csv(p)
        if result is None:
            raise RuntimeError(f"Fichier vide ou illisible : {p}")
        h, d = result
        if header is None:
            header = h
        ncol = len(header)
        for r in d:
            rows.append((list(r) + [None] * ncol)[:ncol])
    n = len(rows)
    print(f"Entrée : {n} lignes x {len(header)} colonnes")

    i_tracking = col_index(header, "Numéro de suivi de l'envoi")
    i_service = col_index(header, "Description du service")
    # BUG TROUVE 2026-08-19 : les colonnes "... en USD" du CSV brut NE SONT PAS celles
    # utilisees par le classeur reel -- le vrai "Montant" (Shipment Detail!A=BK+BN) reference
    # les colonnes EUR "Devise de facturation des frais de transport/de la remise" (noms
    # trompeurs : ce sont des MONTANTS EUR, pas des codes devise -- le code devise est dans
    # la colonne suivante "Code de la devise de facturation"). Reconfirme sur un cas reel
    # (tracking 381547998902) : USD 539.82-458.09=81.73 vs EUR 465.03-394.62=70.41 = valeur
    # reelle du classeur modele. Meme correction que index.js (carrier Node).
    i_fret = col_index(header, "Devise de facturation des frais de transport de l'envoi")
    i_remise = col_index(header, "Devise de facturation de la remise de l'envoi")
    i_taxe = col_index(header, "Devise de facturation des droits de douane et taxes de l'envoi")
    i_invoice_num = col_index(header, "Numéro de facture")
    i_mois = col_index(header, "Mois de facturation (aaaamm)")
    i_invoice_date = col_index(header, "Date de facturation (mm/jj/aaaa)")

    # Controle "bordereau facture par FedEx mais absent du CSV" (2026-08-20, cas reel juillet :
    # facture 634393398, ecart 83,81EUR -- un envoi de NOVEMBRE 2025 apparaissait dans le PDF
    # mais n'a jamais ete reporte dans le CSV Shipment Detail du mois). Meme logique que
    # index.js (carrier Node) -- compare TOUS les trackings du PDF a TOUS les trackings du CSV
    # brut (avant tout filtrage Droits&Taxes/sans PDF).
    if i_tracking is not None:
        csv_trackings = {str(r[i_tracking] or "").strip() for r in rows if r[i_tracking]}
        for p in pdf_infos_precoce:
            manquants = [b for b in p.get("bordereaux", []) if b["tracking"] not in csv_trackings]
            if manquants:
                total_manquant = round(sum(b["sous_total"] for b in manquants), 2)
                detail = ", ".join(f"{b['tracking']} (envoi du {b['date_envoi']}, {b['sous_total']:.2f} EUR)" for b in manquants)
                # BUG TROUVE 2026-08-20 : le caractere '~' (approximation) etait a l'origine
                # U+2248 ("≈") -- absent du charset cp1252 utilise par la console Windows lors
                # de l'appel subprocess depuis server.js, provoquant un UnicodeEncodeError SUR
                # CE print() precis (les accents francais eux passent, deja presents dans cp1252)
                # -- crash masque en cascade cote Node (writeWorkbook, "Cannot read properties
                # of undefined (reading 'Droits et taxes')"), exactement le meme piege que le
                # bug PermissionError/EBUSY deja documente sur TNT.
                print(f"AVERTISSEMENT: PDF {p['file']}" + (f" (n° {p['numero_facture']})" if p['numero_facture'] else "")
                      + f" : {len(manquants)} bordereau(x) facturé(s) par FedEx mais absent(s) du CSV Shipment Detail — {detail}"
                      + f" — total HT manquant ~ {total_manquant:.2f} EUR (probable envoi d'un mois antérieur jamais reporté, à vérifier avec le pôle transport).")

    # Filtrage "Droits & Taxes" (seule regle certaine, cf. docstring index.js) : une facture
    # dont TOUTES les lignes ont un fret nul et une taxe non nulle est un ajustement
    # douanier pur -- exclue de Shipment Detail.
    factures_fret = {}
    factures_taxe = {}
    if i_invoice_num is not None:
        for r in rows:
            nf = str(r[i_invoice_num] or "").strip()
            if not nf:
                continue
            fret = float(str(r[i_fret]).replace(",", ".")) if i_fret is not None and r[i_fret] not in (None, "") else 0.0
            taxe = float(str(r[i_taxe]).replace(",", ".")) if i_taxe is not None and r[i_taxe] not in (None, "") else 0.0
            factures_fret[nf] = factures_fret.get(nf, 0.0) + fret
            factures_taxe[nf] = factures_taxe.get(nf, 0.0) + taxe
    factures_droits_taxes = {nf for nf in factures_fret if factures_fret[nf] == 0 and factures_taxe.get(nf, 0) != 0}
    if factures_droits_taxes and i_invoice_num is not None:
        avant = len(rows)
        rows = [r for r in rows if str(r[i_invoice_num] or "").strip() not in factures_droits_taxes]
        print(f"Facture(s) 100% Droits & Taxes exclue(s) : {sorted(factures_droits_taxes)} ({avant - len(rows)} ligne(s)).")
        n = len(rows)

    # Facture absente des PDF portail fournis (regle utilisateur 2026-08-20, meme logique que
    # index.js) : si au moins un PDF est fourni, un numero de facture du CSV brut absent de
    # TOUS les PDF fournis (numero lu dans leur CONTENU, cf. pdf_invoice_nums plus haut) est
    # tres probablement deja facture un mois precedent -- exclue automatiquement (confirme sur
    # le cas reel 634313590 de juin 2026).
    if pdf_invoice_nums and i_invoice_num is not None:
        # factures_fret contient TOUTES les factures vues initialement (y compris celles deja
        # exclues ci-dessus par la regle Droits&Taxes) -- exclure factures_droits_taxes pour
        # ne pas lister deux fois la meme facture dans des messages differents (elle est deja
        # sortie de "rows", donc deja absente de facon certaine, pas besoin de la re-signaler
        # comme "sans PDF").
        factures_sans_pdf = {nf for nf in factures_fret if nf not in pdf_invoice_nums} - factures_droits_taxes
        if factures_sans_pdf:
            avant = len(rows)
            rows = [r for r in rows if str(r[i_invoice_num] or "").strip() not in factures_sans_pdf]
            print(f"Facture(s) absente(s) des PDF portail fournis (probablement deja facturee(s) un mois precedent) exclue(s) : {sorted(factures_sans_pdf)} ({avant - len(rows)} ligne(s)) -- a confirmer avec le pole transport.")
            n = len(rows)

    # Mois cible = mois majoritaire (meme piege que Chronopost/TNT).
    date_validite_serial = None
    if i_mois is not None or i_invoice_date is not None:
        import datetime as _dt
        from collections import Counter
        compte = Counter()
        for r in rows:
            v = str(r[i_mois] or "").strip() if i_mois is not None else ""
            if not re.match(r"^\d{6}$", v) and i_invoice_date is not None:
                m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", str(r[i_invoice_date] or "").strip())
                if m:
                    v = f"{m.group(3)}{int(m.group(1)):02d}"
            if re.match(r"^\d{6}$", v):
                compte[v] += 1
        if compte:
            mois_maj = compte.most_common(1)[0][0]
            annee, mois = int(mois_maj[:4]), int(mois_maj[4:6])
            EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
            date_validite_serial = (_dt.datetime(annee, mois, 1) - EXCEL_EPOCH).days

    # E/P : export WMS "expeditions_brut" m/m-1 -- repli "particulier" si tracking absent
    # (meme defaut le plus sur que le carrier Node, cf. index.js).
    ep_map = load_brut_ep(brut_paths) if brut_paths else {}
    if not brut_paths:
        print("AVERTISSEMENT: aucun export 'expeditions_brut' fourni -> toutes les lignes classées 'particulier' par défaut (E/P).")
    n_ep_defaut = 0
    ep_values = []
    for r in rows:
        t = str(r[i_tracking] or "").strip() if i_tracking is not None else ""
        ep = ep_map.get(t)
        if not ep:
            ep = "particulier"
            n_ep_defaut += 1
        ep_values.append(ep)
    if n_ep_defaut:
        print(f"{n_ep_defaut} ligne(s) sans correspondance dans l'export WMS (E/P) — classée(s) 'particulier' par défaut.")

    data_brut = [[coerce(v) for v in r] for r in rows]
    pdfs = pdf_infos_precoce
    pdf_total_du_par_facture = {p["numero_facture"]: p["total_du"] for p in pdfs if p["numero_facture"] and p["total_du"] is not None}

    import win32com.client as win32
    xlUp = -4162
    xlDatabase = 1
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) "Shipment Detail" : purge + collage donnees brutes a partir de la colonne F
        #    (A-E = calculees : Montant/Clients/Categorie/E-P/Mode d'envoi).
        ws = wb.Sheets("Shipment Detail")
        if ws.AutoFilterMode:  # meme piege que Chronopost/TNT : AutoFilter actif -> ecritures ignorees
            ws.AutoFilterMode = False
        FIRST_RAW_COL = 6  # colonne F
        LAST_RAW_COL = FIRST_RAW_COL - 1 + len(header)
        LAST_CALC_COL = 5  # A->E
        oldLast = ws.Cells(ws.Rows.Count, FIRST_RAW_COL).End(xlUp).Row
        newLast = 1 + n
        maxLast = max(oldLast, newLast, 2)

        purgeUntil = maxLast + 200
        retry(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(purgeUntil, max(LAST_RAW_COL, LAST_CALC_COL))).ClearContents())

        # "Numéro de suivi" : TOUJOURS numerique dans le brut reel FedEx (jamais de zero de
        # tete constate sur 3607 lignes juin 2026) -- pas de force NumberFormat texte ici,
        # contrairement a TNT (tracking 16 chiffres avec zeros de tete).
        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, LAST_RAW_COL)).__setattr__("Value", data_brut))
        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, max(LAST_RAW_COL, LAST_CALC_COL))).ClearContents())

        # D (E/P) ecrite en VALEUR LITERALE (calculee ci-dessus via l'export WMS) -- PAS une
        # formule, reproduit l'etat reel du fichier de juin 2026 (100% litteral sur 3607/3607
        # lignes, jamais de XLOOKUP vers un fichier externe). B (Clients) laissee vide
        # (saisie manuelle post-generation, meme regle transversale que BLS/Chronopost/
        # Colissimo). C (Categorie) laissee vide (purge suffit, toujours vide en reel).
        col_d_vals = [[v] for v in ep_values]
        retry(lambda: ws.Range(ws.Cells(2, 4), ws.Cells(newLast, 4)).__setattr__("Value", col_d_vals))

        col_service = col_letter(FIRST_RAW_COL - 1 + i_service) if i_service is not None else None
        col_bk = col_letter(FIRST_RAW_COL - 1 + i_fret) if i_fret is not None else None
        col_bn = col_letter(FIRST_RAW_COL - 1 + i_remise) if i_remise is not None else None

        if col_bk and col_bn:
            formula_a = f"={col_bk}{{row}}+{col_bn}{{row}}"
            retry(lambda t=formula_a: ws.Range(ws.Cells(2, 1), ws.Cells(newLast, 1)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))
        else:
            print("AVERTISSEMENT: 'Montant des frais de transport'/'remise' introuvable(s) -> colonne A (Montant) non reconstruite.")

        if col_service:
            formula_e = f'=IF({col_service}{{row}}="FedEx Priority","ST","FICP")'
            retry(lambda t=formula_e: ws.Range(ws.Cells(2, 5), ws.Cells(newLast, 5)).__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))
        else:
            print("AVERTISSEMENT: 'Description du service' introuvable -> colonne E (Mode d'envoi) non reconstruite.")

        ws.Range(ws.Cells(1, 1), ws.Cells(newLast, max(LAST_RAW_COL, LAST_CALC_COL))).AutoFilter()

        # 2) TCD / Bilan factures / Bilan clients : PivotCache redirige vers la vraie plage
        #    large de 'Shipment Detail' (cache du modele fige sur une plage etroite --
        #    'Shipment Detail'!C1:C70 ou C1:C2, meme piege que DPD/Geodis/Mondial Relay/
        #    Chronopost/TNT).
        newRange = ws.Range(ws.Cells(1, 1), ws.Cells(newLast, max(LAST_RAW_COL, LAST_CALC_COL)))
        for sheet_name in ("TCD", "Bilan factures", "Bilan clients"):
            wsPivot = wb.Sheets(sheet_name)
            for i in range(1, wsPivot.PivotTables().Count + 1):
                pt = wsPivot.PivotTables(i)
                src = str(pt.PivotCache().SourceData)
                if "Shipment Detail" not in src:
                    print(f"TCD '{pt.Name}' ({sheet_name}) : source '{src}' n'est pas 'Shipment Detail' -> config préservée.")
                    continue
                newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=newRange)
                pt.ChangePivotCache(newCache)
                print(f"TCD '{pt.Name}' ({sheet_name}) : PivotCache redirigé vers {newRange.Address}.")
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        xl.Calculate()

        # 3) "TCD" colonnes manuelles E/F (etendues jusqu'a la vraie derniere ligne native,
        #    colonne A = Shipment Tracking Number).
        wsTcd = wb.Sheets("TCD")
        lastTcd = wsTcd.Cells(wsTcd.Rows.Count, 1).End(xlUp).Row
        if lastTcd >= 2:
            formulas_tcd = {
                5: "=ROUNDUP(D{row}*0.453592,1)",  # E Poids arrondis sup (kg)
                6: "=IF(_xlfn.XLOOKUP(A{row},'Shipment Detail'!N:N,'Shipment Detail'!D:D)=\"entreprise\",\"E\",\"P\")",  # F E/P
            }
            for col_idx, tmpl in formulas_tcd.items():
                retry(lambda c=col_idx, t=tmpl: wsTcd.Range(wsTcd.Cells(2, c), wsTcd.Cells(lastTcd, c))
                      .__setattr__("Formula", [[t.format(row=r)] for r in range(2, lastTcd + 1)]))
            print(f"'TCD' : colonnes E/F étirées jusqu'à la ligne {lastTcd}.")

        # 4) "Bilan factures" colonne D (PDF) : remplie AUTOMATIQUEMENT avec le "Total dû EUR"
        #    extrait de chaque PDF fourni (regle utilisateur 2026-08-20) -- appariee par numero
        #    de facture (colonne A, valeur native du TCD) au lieu d'une saisie manuelle. Colonne
        #    E (Ecart=C-D) etiree jusqu'a la derniere ligne native (colonne A = Invoice Number).
        #    Facture sans PDF correspondant : colonne D laissee vide (saisie manuelle de repli).
        wsBf = wb.Sheets("Bilan factures")
        lastBf = wsBf.Cells(wsBf.Rows.Count, 1).End(xlUp).Row
        if lastBf >= 4:  # en-tete sur la ligne 3 (constate sur le modele de juin)
            n_pdf_ecrits = 0
            for r in range(4, lastBf + 1):
                # Colonne A = valeur NATIVE du TCD (RowField Invoice Number) -- win32com la
                # renvoie en float (ex. 634323394.0), jamais en texte, contrairement au numero
                # de facture extrait du PDF (string via regex) -- BUG TROUVE 2026-08-20 :
                # comparaison str("634323394.0") vs "634323394" echouait systematiquement
                # (0/10 factures appariees). Normalise en entier des deux cotes.
                rawFacture = wsBf.Cells(r, 1).Value
                try:
                    numFacture = str(int(rawFacture))
                except (TypeError, ValueError):
                    numFacture = str(rawFacture or "").strip()
                totalDu = pdf_total_du_par_facture.get(numFacture)
                if totalDu is not None:
                    wsBf.Cells(r, 4).Value = totalDu
                    n_pdf_ecrits += 1
            retry(lambda: wsBf.Range(wsBf.Cells(4, 5), wsBf.Cells(lastBf, 5))
                  .__setattr__("Formula", [[f"=C{r}-D{r}"] for r in range(4, lastBf + 1)]))
            print(f"'Bilan factures' : colonne D (PDF) renseignée automatiquement pour {n_pdf_ecrits}/{lastBf - 3} facture(s) (Total dû EUR extrait du PDF), colonne E (Ecart) étirée jusqu'à la ligne {lastBf}.")

        # 5) "Import ERP" : formules PAR LIGNE (XLOOKUP tracking, PAS de decalage fixe comme
        #    TNT) -- nombre de lignes = nombre de trackings du TCD.
        wsImp = wb.Sheets("Import ERP")
        if wsImp.AutoFilterMode:
            wsImp.AutoFilterMode = False
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 6).End(xlUp).Row
        LAST_COL_IMPORT = 22
        newLastImp = max(lastTcd, 2)

        formulas_import = {
            1: '=IF(OR(M{row}="FICP",M{row}="IE", M{row}="RE", M{row}="International", M{row}="Europe"),"FEDEX INTERNATIONAL","FEDEX FRANCE")',  # A Transporteur
            6: "=TCD!A{row}",  # F N° Tracking
            8: "=TCD!F{row}",  # H E/P
            9: "=TCD!B{row}",  # I Pays
            10: '=IF(COUNTIF(Zoning!B:B,I{row})=0,"inconnu",_xlfn.XLOOKUP(I{row},Zoning!B:B,Zoning!C:C))',  # J Zone
            11: "=_xlfn.XLOOKUP(F{row},'Shipment Detail'!N:N,'Shipment Detail'!AB:AB)",  # K Nbr Colis
            12: "=TCD!E{row}",  # L Poids
            13: '=IF(J{row}="France","ST",_xlfn.XLOOKUP(F{row},\'Shipment Detail\'!N:N,\'Shipment Detail\'!E:E,""))',  # M mode envoi
            14: '=IF(OR(I{row}="FR",I{row}="DE",I{row}="IT",I{row}="ES",I{row}="BE",I{row}="NL",I{row}="LU",I{row}="PT",I{row}="AT",I{row}="IE",I{row}="FI",I{row}="SE",I{row}="DK",I{row}="PL",I{row}="CZ",I{row}="SK",I{row}="HU",I{row}="RO",I{row}="BG",I{row}="HR",I{row}="SI",I{row}="EE",I{row}="LV",I{row}="LT",I{row}="GR",I{row}="CY",I{row}="MT"),0.2,0)',  # N TVA
            15: "=IF(_xlfn.XLOOKUP(F{row},'Shipment Detail'!N:N,'Shipment Detail'!BM:BM)=0,\"\",_xlfn.XLOOKUP(F{row},'Shipment Detail'!N:N,'Shipment Detail'!BM:BM))",  # O Droits et taxes
            18: "=IF(_xlfn.XLOOKUP(F{row},'Shipment Detail'!N:N,'Shipment Detail'!AV:AV)>60,10,\"\")",  # R Colis volumineux
            20: "=TCD!C{row}",  # T Frêt
        }
        for col_idx, tmpl in formulas_import.items():
            retry(lambda c=col_idx, t=tmpl: wsImp.Range(wsImp.Cells(2, c), wsImp.Cells(newLastImp, c))
                  .__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLastImp + 1)]))

        # DateValidite (colonne B) : valeur ligne 2, puis "=B{n-1}" chaine (modele exact).
        if date_validite_serial is not None:
            wsImp.Cells(2, 2).Value = date_validite_serial
        else:
            print("AVERTISSEMENT: mois de facturation introuvable -> 'Import ERP'!B2 non mise à jour, reste celle du modèle.")
        if newLastImp > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(3, 2), wsImp.Cells(newLastImp, 2))
                  .__setattr__("Formula", [["=B{prev}".format(prev=r - 1)] for r in range(3, newLastImp + 1)]))

        if newLastImp < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLastImp + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT)).ClearContents())
        print(f"'Import ERP' : formules reconstruites jusqu'à la ligne {newLastImp}.")

        # 5bis) Supplements "surplus" (Zones eloignees Q/17, Colis volumineux R/18, Plus-value
        #       BtoC U/21) : ECRITS EN VALEUR par-dessus les formules ci-dessus (pas de formule
        #       Excel possible, la donnee n'existe que dans le texte du PDF) -- Colis volumineux
        #       ADDITIONNE le montant PDF a la valeur deja posee par la formule >60cm=10EUR
        #       (2 mecanismes independants, confirme pole transport 2026-08-20, cf. mail "Oubli
        #       facturation surplus Fedex janvier a mai 2026").
        if supplements_par_tracking and newLastImp >= 2:
            n_maj = 0
            for r in range(2, newLastImp + 1):
                tracking = str(wsImp.Cells(r, 6).Value or "").strip()  # F N° Tracking
                if tracking.endswith(".0"):
                    tracking = tracking[:-2]
                supp = supplements_par_tracking.get(tracking)
                if not supp:
                    continue
                if supp["ZonesEloignees"]:
                    wsImp.Cells(r, 17).Value = supp["ZonesEloignees"]
                if supp["ColisVolumineuxPdf"]:
                    existant = wsImp.Cells(r, 18).Value
                    existant = existant if isinstance(existant, (int, float)) else 0
                    wsImp.Cells(r, 18).Value = round(existant + supp["ColisVolumineuxPdf"], 2)
                if supp["PlusValueB2C"]:
                    wsImp.Cells(r, 21).Value = supp["PlusValueB2C"]
                n_maj += 1
            print(f"'Import ERP' : suppléments PDF (Zones éloignées/Colis volumineux/Plus-value BtoC) appliqués sur {n_maj} ligne(s) — cf. mail pôle transport 2026-08-20.")

        wsImp.Range(wsImp.Cells(1, 1), wsImp.Cells(max(newLastImp, 2), LAST_COL_IMPORT)).AutoFilter()
        xl.Calculate()

        # 6) Reconciliation PDF (info console -- colonne D "PDF" de 'Bilan factures' reste a
        #    la saisie manuelle, comme les autres transporteurs sans onglet de controle dedie).
        for p in pdfs:
            if p["total_du"] is not None:
                print(f"PDF {p['file']}{' (n° ' + p['numero_facture'] + ')' if p['numero_facture'] else ''} : "
                      f"Total dû = {p['total_du']:.2f} EUR (à comparer/coller en 'Bilan factures'!D).")
            else:
                print(f"PDF {p['file']} : 'Total dû' introuvable, ignoré.")

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)

        # 7) Export "Import ERP" en VALEURS (pas en JS recalcule a part) : remontee pole
        # transport 2026-08-24 ("le fichier import CSV deconne alors que la feuille Import CSV
        # du fichier de facture est correcte -> copier-coller depuis le fichier de la facture,
        # coller en valeur"). BUG TROUVE 2026-08-24 (cf. finaliser_colissimo.py, meme fix) :
        # lire Range.Value AVANT Save()/Close(), meme juste apres un Calculate(), peut renvoyer
        # des cellules dans un etat transitoire (valeur d'erreur COM brute, ou une ligne vide)
        # qui ne correspond PAS a ce que contient le fichier une fois sauvegarde sur disque.
        # Fix : fermer le classeur (deja fait ci-dessus), le ROUVRIR en lecture, et lire
        # seulement alors -- meme etat que ce que le pole transport voit en ouvrant le fichier.
        wbRead = None
        try:
            # wbRead ferme dans le finally imbrique ci-dessous, quoi qu'il arrive -- cf.
            # finaliser_colissimo.py (meme bug/fix, 2026-08-24) : une exception entre
            # l'ouverture et la fermeture laissait un EXCEL.EXE orphelin, verrouillant le
            # fichier de sortie pour toute regeneration suivante du meme mois.
            wbRead = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=True))
            wsImpRead = wbRead.Sheets("Import ERP")
            impLast = newLastImp
            if impLast >= 2:
                values = retry(lambda: wsImpRead.Range(wsImpRead.Cells(2, 1), wsImpRead.Cells(impLast, LAST_COL_IMPORT)).Value)
                # BUG TROUVE 2026-08-24 : le TCD (PivotTable, RowField "Shipment Tracking
                # Number") contient un item "(blank)" LEGITIME quand le CSV brut a au moins 1
                # ligne sans tracking (ex. ajustement/note de credit sans envoi associe) -> une
                # ligne "Import ERP" a Tracking vide se glissait dans l'export. Le carrier Node
                # (index.js) exclut deja ce cas explicitement (`if (!tracking) continue;`) --
                # meme filtre applique ici pour rester coherent (pas une vraie ligne facturable).
                # Filet de securite residuel : une cellule encore en erreur COM (marshaling
                # ponctuel, type int au lieu de str/float/datetime/None) est ecrite vide plutot
                # que comme un nombre absurde.
                values = [row for row in values if str(row[5]).strip() and str(row[5]).strip() != "(vide)"]
                n_err = 0
                def clean(v):
                    nonlocal n_err
                    if isinstance(v, int) and not isinstance(v, bool) and v < -1000000:
                        n_err += 1
                        return ""
                    return "" if v is None else v
                export_path = os.path.splitext(sortie)[0] + "_import_valeurs.csv"
                with open(export_path, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f, delimiter=";")
                    for row in values:
                        w.writerow([clean(v) for v in row])
                if n_err:
                    print(f"AVERTISSEMENT: {n_err} cellule(s) en erreur COM lors de l'export Import ERP (valeurs) -- laissees vides, a verifier.")
                print(f"EXPORT_IMPORT_VALEURS:{export_path}")
        except Exception as e:
            print("Export Import ERP (valeurs) ignore :", e)
        finally:
            if wbRead is not None:
                try:
                    wbRead.Close(SaveChanges=False)
                except Exception:
                    pass

        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
