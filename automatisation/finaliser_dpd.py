#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINALISEUR DPD — produit "AAAA_MM_Facture DPD.xlsx" A L'IDENTIQUE du fichier
fait a la main (7 feuilles : Zoning, Bilan factures, Facture DPD, Import ERP,
Bilan clients, Tarifs achat DPD, Suppléments DPD ; memes formules, memes TCD,
meme mise en forme), en partant du fichier existant comme MODELE.

Meme principe que finaliser_geodis.py : les colonnes brutes DPD ("complement_
facture...") se decalent d'un mois a l'autre (colonne ajoutee/retiree), donc
les postes calcules de "Facture DPD" (colonnes A->N) ne sont PAS colles avec
des formules a positions fixes -- on colle l'entete recu TEL QUEL (avec le
style bleu du modele) a partir de la colonne O, PUIS on reconstruit les
formules A->N en cherchant CHAQUE colonne par NOM dans cet entete
(postes_from_columns du config.json Node, source unique de verite, partagee
avec facturation-app/src/carriers/dpd/config.json).

Deux notions distinctes de "frais de dossier" (cf. config.json Node) :
  - poste M ("Frais dossier reel") = colonne brute 'Frais de tenue de compte'
    TELLE QUELLE (deja repartie par DPD dans son propre fichier au sein d'une
    meme facture) -> entre dans "Total GO" (E=SUM(G:M)+SUM(A:C)), utilise pour
    la reconciliation PDF (Bilan factures : Somme de Total GO).
  - poste N ("Frais dossier modifie") = ROUNDUP(SUM(M:M)/COUNT(M:M),2) : le
    total de TOUT le mois reparti egalement sur toutes les lignes -> utilise
    UNIQUEMENT dans Import ERP.Fret (='Facture DPD'!I+K+N), pas dans Total GO.

Colonne F ("Clients") de "Facture DPD" : vide dans le modele (le TCD "Bilan
clients" affiche un cache fige/obsolete, source reelle = No de compte d'apres
le cache pivot) -> reconstruite ici comme =T{row} (No de compte) pour que le
TCD redevienne vivant apres RefreshAll.

Necessite : Windows + Excel + pywin32 + openpyxl.
Usage :
  python finaliser_dpd.py "<modele.xlsx>" "<sortie.xlsx>" "<entree1>" [<entree2>...] [--pdf <pdf1> [<pdf2>...]]
"""
import sys, os, shutil, re, csv, json, zipfile

FIRST_RAW_COL = 15  # colonne O : debut des donnees brutes DPD dans "Facture DPD"
HEADER_FILL_ARGB = None  # determine depuis le modele (cf. read_header_style)

# Colonnes de "Facture DPD" -> cle postes_from_columns (config.json Node).
# B (Assurance) n'a pas de formule dans le modele (toujours vide) -> ignoree ici.
POSTE_COLS = {
    "A": "DroitsTaxes", "C": "ZonesEloignees", "G": "ColisVolumineux", "H": "Adresses",
    "I": "Fret", "J": "BtoC", "K": "Retour", "L": "Gazole",
}
# Colonnes a formule FIXE (pas de reconstruction par nom) : D,E,M,N calculees
# a partir d'autres colonnes de Facture DPD elle-meme (pas des colonnes brutes).
# E (Total GO) : le modele clone (juin 2026) utilise SUM(G:M)+SUM(A:C) --
# c'est-a-dire M (Frais dossier REEL, brut par facture). Mais le fichier fait a
# la main de JUILLET 2026 utilise SUM(G:L)+SUM(A:C)+N -- N (Frais dossier
# MODIFIE, reparti egalement sur tout le mois) A LA PLACE de M. Vrai changement
# de logique metier confirme entre les 2 mois (pas une erreur de reconstruction
# de colonnes) -> on suit la version la plus recente (juillet), pas celle du
# modele clone.
FIXED_FORMULAS = {
    "D": "=E{row}-L{row}",                                  # Total hors GO
    "E": "=SUM(G{row}:L{row})+SUM(A{row}:C{row})+N{row}",    # Total GO (formule juillet 2026, PAS celle du modele juin)
    "F": "=T{row}",                                          # Clients (= No de compte, reactive le TCD Bilan clients)
    "M": "=BI{row}",                                         # Frais dossier reel (reconstruit par nom plus bas, PAS fixe -> voir main())
    "N": "=ROUNDUP($DI$1/COUNT(M:M),2)",                     # Frais dossier modifie
}


def load_dpd_config():
    """Lit config.json depuis le carrier Node (source unique de verite, deja
    utilise par src/carriers/dpd/index.js -> pas de mapping duplique/desync)."""
    cfg_path = os.path.join(os.path.dirname(__file__), "..", "facturation-app",
                             "src", "carriers", "dpd", "config.json")
    with open(cfg_path, encoding="utf-8") as f:
        return json.load(f)


def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def compare_key(h):
    """Cle de comparaison ROBUSTE aux variantes cosmetiques vues entre exports
    DPD (accents, points/virgules) -- meme logique que Geodis."""
    s = normalize_header(h).lower()
    s = s.translate(str.maketrans("éèêëàâäùûüôöîïç", "eeeeaaauuuooiic"))
    s = re.sub(r"[.,]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def is_xlsx(path):
    """Detection par contenu (magic bytes ZIP), pas par extension : les
    fichiers uploades (multer) arrivent sans extension du tout."""
    try:
        return zipfile.is_zipfile(path)
    except Exception:
        return False


def read_xlsx_rows(path):
    """openpyxl refuse d'ouvrir un fichier sans extension .xlsx (verifiee sur le
    NOM, pas le contenu) -- les fichiers uploades (multer) arrivent SANS
    extension du tout -> copie temporaire renommee si besoin (meme piege deja
    rencontre et corrige dans finaliser_geodis.py)."""
    import openpyxl, tempfile
    load_path = path
    tmp_path = None
    if not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copyfile(path, tmp.name)
        load_path = tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(load_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def read_csv_rows(path):
    with open(path, encoding="latin-1", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    if not rows:
        return None
    header = [normalize_header(h) for h in rows[0]]
    data = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    return header, data


def read_input(paths):
    """Concatene toutes les sources (CSV ou XLSX, un fichier par n° facture/
    sous-compte DPD) -> (header, data_rows). Suppose meme structure de colonnes
    pour tous les fichiers d'un meme mois (verifie/reclasse par nom ensuite).
    Chaque fichier source contient, en fin de bloc, des lignes de SOUS-TOTAL
    ('Total service'/'Total expéditions', valeurs agregees genre '3,39' dans la
    colonne No de compte) sans 'N° Colis' -- confirme sur le fichier reel (colonne
    N Colis vide sur ces lignes). Exclues ici (comme deja fait cote carrier
    Node, index.js: .filter(rec => rec.tracking)), sinon elles polluent les
    postes calcules (Total GO) et les TCD."""
    header = None
    all_rows = []
    for p in paths:
        result = read_xlsx_rows(p) if is_xlsx(p) else read_csv_rows(p)
        if result is None:
            raise RuntimeError(f"Fichier vide ou illisible : {p}")
        h, rows = result
        if header is None:
            header = h
        ncol = len(header)
        i_colis = next((i for i, name in enumerate(header) if compare_key(name) == compare_key("N° Colis")), None)
        for r in rows:
            r = (list(r) + [None] * ncol)[:ncol]
            if i_colis is not None and not str(r[i_colis] or "").strip():
                continue
            all_rows.append(r)
    return header, all_rows


def coerce(v):
    """Nombre si la valeur est un nombre francais valide (virgule ou point),
    sinon la valeur telle quelle (texte/date/None).
    Exception : un entier avec un zero de tete est une reference/code, pas une
    quantite -- reste texte (meme regle que Geodis, cf. finaliser_geodis.py).
    datetime/date -> NOMBRE SERIEL Excel (jours depuis 1899-12-30) : passer un
    datetime naif directement a une propriete COM Value le fait repasser par le
    fuseau horaire LOCAL de la machine lors de la conversion en PyTime
    (pywin32), ce qui peut faire glisser la date d'un jour (deja constate et
    corrige sur Geodis)."""
    import datetime as _dt
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    if isinstance(v, _dt.datetime):
        return (v - EXCEL_EPOCH).total_seconds() / 86400
    if isinstance(v, _dt.date):
        return (_dt.datetime(v.year, v.month, v.day) - EXCEL_EPOCH).days
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.fullmatch(r"-?0\d+", s):
        return v
    if re.fullmatch(r"-?\d+(,\d+)?", s):
        return float(s.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


def argb_to_com_bgr(argb):
    """'FF3200E6' (ARGB hex, format openpyxl) -> entier BGR attendu par
    Range.Interior.Color / Font.Color en COM (B*65536 + G*256 + R)."""
    r, g, b = int(argb[2:4], 16), int(argb[4:6], 16), int(argb[6:8], 16)
    return b * 65536 + g * 256 + r


def col_letter(idx0):
    import openpyxl
    return openpyxl.utils.get_column_letter(idx0 + 1)


def read_header_style(modele_path):
    """Style (fond + police) de l'entete brute du modele (ligne 1, colonne
    FIRST_RAW_COL), pour le reappliquer sur l'entete du fichier recu -- peut
    differer de Geodis (bleu) selon le classeur DPD."""
    import openpyxl
    wb = openpyxl.load_workbook(modele_path)
    ws = wb["Facture DPD"]
    cell = ws.cell(row=1, column=FIRST_RAW_COL)
    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    font = cell.font.color.rgb if cell.font and cell.font.color else None
    wb.close()
    return (fill if isinstance(fill, str) else "FF3200E6"), (font if isinstance(font, str) else "FFFFFFFF")


def build_poste_formulas(header, dpd_config, first_raw_col):
    """Pour chaque poste ERP a somme de colonnes nommees (POSTE_COLS), cherche
    ses colonnes DANS L'ENTETE REEL recu ce mois-ci -> renvoie
    { "A": "=X{row}+Y{row}+...", ... } avec les BONNES lettres de colonnes pour
    ce fichier. Une colonne introuvable ce mois-ci est ignoree dans la somme
    (poste calcule sur les colonnes presentes) + avertissement, jamais bloquant
    (coherent avec le comportement Node : idx() -> -1, colonne juste ignoree)."""
    postes_from_columns = dpd_config["postes_from_columns"]
    key_by_pos = {compare_key(h): i for i, h in enumerate(header)}
    formulas, missing = {}, {}
    for col, poste_key in POSTE_COLS.items():
        names = postes_from_columns.get(poste_key, [])
        letters, not_found = [], []
        for name in names:
            i = key_by_pos.get(compare_key(name))
            if i is None:
                not_found.append(name)
            else:
                letters.append(col_letter(first_raw_col - 1 + i))
        if letters:
            formulas[col] = "=" + "+".join(f"{L}{{row}}" for L in letters)
        if not_found:
            missing[poste_key] = not_found
    # M ("Frais dossier reel") : colonne UNIQUE 'Frais de tenue de compte', pas une somme.
    frais_col_name = dpd_config.get("frais_dossier_source_col")
    i = key_by_pos.get(compare_key(frais_col_name)) if frais_col_name else None
    if i is not None:
        formulas["M"] = f"={col_letter(first_raw_col - 1 + i)}{{row}}"
    else:
        missing["FraisDossierReel"] = [frais_col_name]
    return formulas, missing


def extract_pdf_totals(pdf_paths):
    """PDF DPD ('Complément de facture' ou 'FACTURE' principale -- la mise en
    page CHANGE d'un mois a l'autre, confirme sur juin vs juillet 2026, donc on
    reste volontairement minimal/tolerant plutot que de coller a un format
    precis) :
    - Cle de rapprochement = 4 DERNIERS CHIFFRES DU N° DE CLIENT ('N° de
      client : 021-009836' -> '9836'), PAS le n° de facture -- ca correspond
      exactement a 'No de compte' dans le classeur (confirme : meme valeur des
      2 cotes). Le n° de facture n'a AUCUN rapport avec 'No de compte'.
    - Montant = premier montant HT trouve sous un des libelles connus ('Total
      EUR H.T', 'TOTAL FACTURE HT') -- recherche tolerante aux espaces/sauts de
      ligne entre le libelle et le nombre (les deux varient selon le mois)."""
    try:
        import pypdf
    except ImportError:
        return []
    TOTAL_LABELS = [r"Total\s*EUR\s*H\.?T", r"TOTAL\s*FACTURE\s*HT"]
    results = []
    for p in pdf_paths:
        try:
            text = "\n".join((page.extract_text() or "") for page in pypdf.PdfReader(p).pages)
        except Exception:
            continue
        text = text.replace("\xa0", " ")
        m_client = re.search(r"N°\s*de\s*client\s*:\s*([\d\-]+)", text, re.IGNORECASE)
        if not m_client:
            continue
        total = None
        for label in TOTAL_LABELS:
            m_total = re.search(label + r"\s*[\r\n]*\s*([\d\s]+[.,]\d{2})", text, re.IGNORECASE)
            if m_total:
                total = float(m_total.group(1).replace(" ", "").replace(",", "."))
                break
        if total is None:
            continue
        last4 = re.sub(r"\D", "", m_client.group(1))[-4:]
        results.append({"file": os.path.basename(p), "last4": last4, "total": total})
    return results


def fill_reconciliation(wb, pdfs):
    """Onglet 'Bilan factures' : colonne 'Montant' (G, saisie a la main dans le
    fichier fait a la main) remplie automatiquement depuis les PDF. Rapproche
    chaque PDF directement par ses 4 derniers chiffres de 'N° de client'
    (extract_pdf_totals) -> ligne du TCD portant ce No de compte en colonne B
    (lecture dynamique du TCD deja rafraichi, cf. section 4 de main()).

    ATTENTION : le tableau 'Factures pdf' (F:I) est saisi/colle A LA MAIN a
    cote du TCD, PAS regenere par le TCD lui-meme -> si le TCD a moins de
    lignes ce mois-ci qu'avant (nombre de comptes distincts different), les
    anciennes valeurs de F:I (ex. colonne G 'Montant' du mois precedent, dans
    le modele clone) restent affichees sur les lignes en trop, y compris sur
    la ligne 'Total general' qui se retrouve alors decalee sur une ancienne
    ligne de donnees -> purge de F:I sur toute la hauteur AVANT de reecrire,
    jusqu'a l'ancienne fin de tableau (constate : modele juin avait 'Montant'
    8694.72EUR pour le compte 10005 en ligne 18, reste affiche en ligne 18
    'Total general' du mois suivant qui n'a que 15 comptes).

    DEUXIEME PIEGE (meme cause, autre symptome) : dans le modele, 'Total
    general' est sa PROPRE ligne, SANS formule F/H/I (juste B/C/D remplis par
    le TCD) -- cf. juin, ligne 20 : F20/G20/H20/I20 tous vides. Mais le TCD de
    juillet a UNE LIGNE DE MOINS (16 comptes au lieu de 17) -> 'Total general'
    remonte sur une ligne qui PORTAIT ENCORE les formules F/H/I d'une ancienne
    ligne de compte du modele (F='=RIGHT(B,4)', I='=C-G') -> avec G vide (pas
    de PDF pour une ligne 'Total general'), I affiche a tort 'C - 0' = le
    total General entier colonne C, au lieu de rester vide comme dans le
    modele. Corrige ici en vidant F/H/I explicitement sur la ligne 'Total
    general' une fois sa position connue (apres refresh TCD)."""
    xlUp = -4162
    ws = wb.Sheets("Bilan factures")
    last = ws.Cells(ws.Rows.Count, 2).End(xlUp).Row  # col B = No de compte
    header_row = None
    total_row = None
    for r in range(1, last + 1):
        v = str(ws.Cells(r, 2).Value or "").strip()
        if v == "Étiquettes de lignes":
            header_row = r
        elif v == "Total général":
            total_row = r
    start = (header_row or 2) + 1

    # Purge F:I (num facture / Montant / HT / Ecart) sur toute la hauteur
    # potentielle de l'ANCIEN tableau (jusqu'a la derniere ligne non vide de F,
    # qui peut deborder au-dela du nouveau TCD si celui-ci a moins de lignes).
    lastF = ws.Cells(ws.Rows.Count, 6).End(xlUp).Row
    if lastF >= start:
        ws.Range(ws.Cells(start, 7), ws.Cells(lastF, 7)).ClearContents()  # colonne G = Montant uniquement (F/H/I sont des formules du modele, deja recalculees par le TCD/formules fixes)

    # Ligne 'Total general' : jamais de F/H/I dans le modele (cf. docstring) --
    # vider explicitement au cas ou elle a herite des formules d'une ancienne
    # ligne de compte (TCD plus court ce mois-ci).
    if total_row is not None:
        ws.Range(ws.Cells(total_row, 6), ws.Cells(total_row, 9)).ClearContents()

    if not pdfs:
        return

    compte_to_row = {}
    for r in range(start, last + 1):
        v = ws.Cells(r, 2).Value
        if v not in (None, "", "(vide)", "Total général"):
            last4 = str(int(v))[-4:] if isinstance(v, (int, float)) else str(v)[-4:]
            compte_to_row[last4] = r

    matched = 0
    for p in pdfs:
        row = compte_to_row.get(p["last4"])
        if row is None:
            print(f"AVERTISSEMENT: facture PDF {p['file']} (client se terminant par {p['last4']}) "
                  f"-> compte introuvable dans Bilan factures, montant non affecté.")
            continue
        ws.Cells(row, 7).Value = p["total"]  # colonne G = Montant
        matched += 1
    print(f"Réconciliation PDF : {matched}/{len(pdfs)} facture(s) rapprochée(s) dans Bilan factures.")


def parse_args(argv):
    """<modele> <sortie> <entree1> [<entree2>...] [--pdf <pdf1> [<pdf2>...]]"""
    modele, sortie = argv[1], argv[2]
    inputs, pdfs, cur = [], [], "in"
    for a in argv[3:]:
        if a == "--pdf":
            cur = "pdf"
        elif cur == "pdf":
            pdfs.append(a)
        else:
            inputs.append(a)
    return modele, sortie, inputs, pdfs


def retry(fn, tries=8, delay=0.6):
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def first_day_of_month_serial(header, rows, date_col_name):
    """1er jour du mois de la premiere 'Date de facture' trouvee, en NOMBRE
    SERIEL Excel -- utilise pour 'Import ERP'!B2 ('Date validite tarif'), une
    valeur FIXE saisie a la main dans le modele (=B2 recopie sur les lignes
    suivantes) qui reste sinon bloquee sur le mois du modele (ex. juin) meme
    quand on traite un autre mois (juillet) -- constate sur 2026_07."""
    import datetime as _dt
    i = next((idx for idx, name in enumerate(header) if compare_key(name) == compare_key(date_col_name)), None)
    if i is None:
        return None
    EXCEL_EPOCH = _dt.datetime(1899, 12, 30)
    for r in rows:
        v = r[i] if i < len(r) else None
        d = None
        if isinstance(v, _dt.datetime):
            d = v
        elif isinstance(v, _dt.date):
            d = _dt.datetime(v.year, v.month, v.day)
        elif v not in (None, ""):
            s = str(v).strip()
            m = re.match(r"^(\d{2})[./](\d{2})[./](\d{4})$", s)
            if m:
                d = _dt.datetime(int(m.group(3)), int(m.group(2)), 1)
        if d is not None:
            first = _dt.datetime(d.year, d.month, 1)
            return (first - EXCEL_EPOCH).days
    return None


def main():
    modele, sortie, inputs, pdf_paths = parse_args(sys.argv)
    shutil.copyfile(modele, sortie)  # on ne touche JAMAIS au modele

    dpd_config = load_dpd_config()
    header, rows = read_input(inputs)
    n = len(rows)
    ncol = len(header)
    data = [[coerce(v) for v in r] for r in rows]
    print(f"Entrée : {n} lignes x {ncol} colonnes")

    date_validite_serial = first_day_of_month_serial(header, rows, dpd_config.get("date_facture_col", "Date de facture"))

    poste_formulas, missing_cols = build_poste_formulas(header, dpd_config, FIRST_RAW_COL)
    for poste_key, names in missing_cols.items():
        print(f"AVERTISSEMENT: colonne(s) introuvable(s) pour le poste {poste_key} "
              f"(ignorée(s) dans la somme, calculé sur les colonnes présentes) : {names}")

    header_fill, header_font = read_header_style(modele)
    pdfs = extract_pdf_totals(pdf_paths) if pdf_paths else []

    import win32com.client as win32
    xlUp = -4162
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    try:
        wb = retry(lambda: xl.Workbooks.Open(os.path.abspath(sortie), UpdateLinks=0, ReadOnly=False))
        if wb is None:
            raise RuntimeError("Excel n'a pas pu ouvrir le fichier (déjà ouvert ? verrouillé ?)")

        # 1) Feuille "Facture DPD" : purge (entete + donnees, colonne O+), collage
        #    entete (avec style du modele) + donnees brutes telles quelles.
        #    ATTENTION : DH1/DI1 (derniere colonne du modele) portent une formule
        #    fixe ('Somme frais de dossier' / =SUM(M:M), lue par N='Frais dossier
        #    modifie') -- PAS des donnees brutes du fichier recu. Si le fichier
        #    recu a moins de colonnes que le modele, la purge (qui va jusqu'a
        #    oldLastCol pour nettoyer un ancien mois plus large) les effacerait
        #    -> sauvegardees avant purge, restaurees juste apres.
        ws = wb.Sheets("Facture DPD")
        oldLastCol = ws.Cells(1, ws.Columns.Count).End(-4159).Column  # xlToLeft
        lastCol = FIRST_RAW_COL + ncol - 1
        oldLast = ws.Cells(ws.Rows.Count, FIRST_RAW_COL).End(xlUp).Row
        newLast = 1 + n
        maxLast = max(oldLast, newLast)
        maxCol = max(oldLastCol, lastCol)
        diCol = oldLastCol  # derniere colonne du modele = DI (Somme frais de dossier / SUM(M:M))
        diHeaderFormula = ws.Cells(1, diCol).Formula if diCol > lastCol else None
        diLabelFormula = ws.Cells(1, diCol - 1).Formula if diCol > lastCol else None
        retry(lambda: ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(maxLast, maxCol)).ClearContents())
        retry(lambda: ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(1, lastCol)).__setattr__("Value", [header]))
        headerRange = ws.Range(ws.Cells(1, FIRST_RAW_COL), ws.Cells(1, lastCol))
        headerRange.Interior.Color = argb_to_com_bgr(header_fill)
        headerRange.Font.Color = argb_to_com_bgr(header_font)
        if diHeaderFormula is not None:
            ws.Cells(1, diCol - 1).Formula = diLabelFormula
            ws.Cells(1, diCol).Formula = diHeaderFormula
        retry(lambda: ws.Range(ws.Cells(2, FIRST_RAW_COL), ws.Cells(newLast, lastCol)).__setattr__("Value", data))

        # 2) Reconstruire les formules des postes calcules (A->N), ligne a ligne,
        #    avec les lettres de colonnes resolues pour CE fichier -- colonnes a
        #    somme nommee (POSTE_COLS) + colonnes a formule fixe (D,E,F,N).
        all_formulas = dict(FIXED_FORMULAS)
        all_formulas.update(poste_formulas)  # A,C,G,H,I,J,K,L,M ecrasent les entrees fixes si presentes
        col_order = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
                     "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14}
        for col, tmpl in all_formulas.items():
            colIdx = col_order[col]
            retry(lambda c=colIdx, t=tmpl: ws.Range(ws.Cells(2, c), ws.Cells(newLast, c))
                  .__setattr__("Formula", [[t.format(row=r)] for r in range(2, newLast + 1)]))
        if newLast < oldLast:
            retry(lambda: ws.Range(ws.Cells(newLast + 1, 1), ws.Cells(oldLast, maxCol)).ClearContents())

        # 3) Onglet "Import ERP" : formules a positions FIXES (referencent les
        #    colonnes A->N de Facture DPD, stables -> pas de reconstruction par
        #    nom necessaire ici, juste etendre par recopie normale. EXCEPTION :
        #    B2 ("Date validite tarif") est une VALEUR FIXE saisie a la main
        #    dans le modele (B3+ = "=B2" recopie) -> mise a jour vers le 1er du
        #    mois traite AVANT le FillDown, sinon elle reste bloquee sur le mois
        #    du modele (ex. juin) meme en traitant un autre mois (constate sur
        #    juillet 2026 : "Date validite tarif" restait a 01/06/2026).
        wsImp = wb.Sheets("Import ERP")
        # colonnes A->V (donnees reelles). PAS jusqu'a X : X1:X4 porte une note
        # manuelle du modele ('Format tracking a modifier' + un exemple avant/
        # apres + 'ATTENTION AUX POIDS A 0'), jamais remplie au-dela de la ligne
        # 4 dans le modele lui-meme -- un FillDown jusqu'a X aurait recopie cet
        # exemple de juin sur TOUTES les lignes de donnees du mois traite.
        LAST_COL_IMPORT_ERP = 22
        oldLastImp = wsImp.Cells(wsImp.Rows.Count, 6).End(xlUp).Row  # col F = N Tracking
        if date_validite_serial is not None:
            wsImp.Cells(2, 2).Value = date_validite_serial
        else:
            print("AVERTISSEMENT: 'Date de facture' introuvable dans le fichier reçu -> "
                  "'Date validité tarif' (Import ERP!B2) non mise à jour, reste celle du modèle.")
        if newLast > 2:
            retry(lambda: wsImp.Range(wsImp.Cells(2, 1), wsImp.Cells(newLast, LAST_COL_IMPORT_ERP)).FillDown())
        if newLast < oldLastImp:
            retry(lambda: wsImp.Range(wsImp.Cells(newLast + 1, 1), wsImp.Cells(oldLastImp, LAST_COL_IMPORT_ERP)).ClearContents())

        # 4) "Bilan factures" / "Bilan clients" : VRAIS TCD Excel dynamiques
        #    (rafraichissables par le pole transport ensuite, clic droit >
        #    Actualiser). Le PivotCache du modele pointe sur une plage FIXE et
        #    etroite (ex. 'Facture DPD'!C1:C20 pour Bilan factures -- vestige
        #    d'un etat anterieur du fichier) -> on la redefinit vers la vraie
        #    plage de donnees de ce mois AVANT de rafraichir, sinon RefreshAll()
        #    seul ne recalcule QUE cette plage figee et laisse des lignes
        #    fantomes/tronquees. IMPORTANT : ChangePivotCache() doit etre appele
        #    juste apres avoir recree la table (pas apres un retrait+reecriture
        #    manuelle du TCD, qui corrompt le RowField de facon non
        #    reproductible -- teste et confirme).
        wsFactureDpd = ws
        newLastPivot = max(newLast, 2)
        xlDatabase = 1
        newRange = wsFactureDpd.Range(wsFactureDpd.Cells(1, 1), wsFactureDpd.Cells(newLastPivot, 20))
        for sheetName in ("Bilan factures", "Bilan clients"):
            wsPivot = wb.Sheets(sheetName)
            for i in range(1, wsPivot.PivotTables().Count + 1):
                pt = wsPivot.PivotTables(i)
                newCache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=newRange)
                pt.ChangePivotCache(newCache)
        wb.RefreshAll()
        try:
            xl.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass

        # 5) Réconciliation PDF (colonne Montant de Bilan factures). Appelee
        #    MEME SANS PDF fourni : purge quand meme les anciennes valeurs
        #    'Montant' laissees par le modele clone (cf. docstring de la
        #    fonction -- sinon une ancienne valeur de mois precedent reste
        #    affichee sur une ligne devenue 'Total general').
        try:
            fill_reconciliation(wb, pdfs)
        except Exception as e:
            print("Réconciliation Bilan factures ignorée :", e)

        xl.Calculate()
        retry(lambda: wb.Save())
        wb.Close(SaveChanges=True)
        print(f"OK -> {sortie}")
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
